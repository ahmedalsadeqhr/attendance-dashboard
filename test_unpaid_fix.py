"""Test the Unpaid Leave formula fix using Feb 2026 data."""
import sys
import os
import traceback

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from attendance_dashboard_streamlit import AttendanceProcessor
from openpyxl import load_workbook
import io

TEST_DIR = r"C:\Users\high tech\Desktop\Feb 2026\New folder"
MASTER_FILE = os.path.join(TEST_DIR, "Master.xlsx")
ATTENDANCE_FILE = os.path.join(TEST_DIR, "21 Jan to 17 Feb.xls")
LEAVE_FILE = os.path.join(TEST_DIR, "Updated Leaves 2026.xlsx")


def run_test():
    print("=" * 60)
    print("  TEST: Unpaid Leave Formula Fix (Feb 2026 data)")
    print("=" * 60)

    proc = AttendanceProcessor()

    # 1. Load master data
    print("\n[1] Loading master data...")
    with open(MASTER_FILE, "rb") as f:
        ok = proc.load_master_data(f, "Master.xlsx")
    print(f"    Loaded {len(proc.employee_mapping)} employees. OK={ok}")
    assert ok, "Failed to load master data"

    # 2. Load leave data
    print("\n[2] Loading leave data...")
    with open(LEAVE_FILE, "rb") as f:
        ok = proc.load_leave_data(f, "Updated Leaves 2026.xlsx")
    print(f"    Loaded {len(proc.leave_records)} leave records. OK={ok}")
    assert ok, "Failed to load leave data"

    # 3. Process attendance
    print("\n[3] Processing attendance...")
    with open(ATTENDANCE_FILE, "rb") as f:
        att_files = [(f, "21 Jan to 17 Feb.xls")]
        proc.process_attendance_files(att_files)
    print(f"    Processed {len(proc.processed_data)} records")
    assert len(proc.processed_data) > 0, "No attendance records processed"

    # 4. Fill leave records
    print("\n[4] Filling leave records...")
    proc.fill_leave_records()
    print(f"    Total records after leave fill: {len(proc.processed_data)}")

    # 5. Generate report
    print("\n[5] Generating Excel report...")
    excel_bytes = proc.create_excel_report()
    assert excel_bytes, "No report generated"

    # 6. Inspect the Penalties sheet for Unpaid Leave formula
    print("\n[6] Checking Penalties sheet - Column T (Unpaid Leave Count)...")
    wb = load_workbook(io.BytesIO(excel_bytes))
    penalties_ws = wb["Penalties"]

    errors = []
    checked = 0
    for row in range(5, penalties_ws.max_row + 1):
        cell = penalties_ws.cell(row, 20)  # Column T
        val = cell.value
        if val is None:
            continue
        val_str = str(val)
        if "COUNTIF" not in val_str:
            continue

        checked += 1
        # The formula should NOT contain "Unpaid leave" (lowercase)
        if '"Unpaid leave"' in val_str:
            crm = penalties_ws.cell(row, 1).value
            errors.append(f"    Row {row} (CRM={crm}): STILL has 'Unpaid leave' term!\n      Formula: {val_str}")

        # The formula SHOULD have "Unpaid Leave" and "Unpaid Leave (BD)"
        if '"Unpaid Leave"' not in val_str:
            crm = penalties_ws.cell(row, 1).value
            errors.append(f"    Row {row} (CRM={crm}): Missing 'Unpaid Leave' term!")

        if '"Unpaid Leave (BD)"' not in val_str:
            crm = penalties_ws.cell(row, 1).value
            errors.append(f"    Row {row} (CRM={crm}): Missing 'Unpaid Leave (BD)' term!")

    print(f"    Checked {checked} formula cells in Column T")

    if checked > 0:
        # Show a sample formula
        for row in range(5, penalties_ws.max_row + 1):
            cell = penalties_ws.cell(row, 20)
            if cell.value and "COUNTIF" in str(cell.value):
                print(f"    Sample formula (row {row}): {cell.value}")
                break

    if errors:
        print("\n  ERRORS FOUND:")
        for e in errors:
            print(e)
        print("\n  [X] TEST FAILED")
        return False
    else:
        print("\n  [+] All Unpaid Leave formulas are correct (no duplicate lowercase term)")

    # 7. Also check all sheet names are present
    print("\n[7] Verifying report sheets...")
    expected_sheets = ["Summary Report", "Individual Analytics", "Alerts & Warnings", "Penalties", "Duplicates"]
    for sheet in expected_sheets:
        present = sheet in wb.sheetnames
        status = "[+]" if present else "[X]"
        print(f"    {status} {sheet}: {'present' if present else 'MISSING'}")
        if not present:
            errors.append(f"Missing sheet: {sheet}")

    # 8. Quick sanity check on Summary sheet
    print("\n[8] Summary Report sanity check...")
    summary_ws = wb["Summary Report"]
    crm_count = summary_ws.max_row - 3  # header rows
    col_count = summary_ws.max_column
    print(f"    {crm_count} employees, {col_count} columns (dates + 3 header cols)")

    # Save the report for manual inspection
    output_path = os.path.join(TEST_DIR, "TEST_OUTPUT_unpaid_fix.xlsx")
    with open(output_path, "wb") as f:
        f.write(excel_bytes)
    print(f"\n    Report saved to: {output_path}")

    wb.close()

    if errors:
        print("\n  [X] TEST FAILED")
        return False
    else:
        print("\n  [+] ALL CHECKS PASSED")
        return True


if __name__ == "__main__":
    try:
        passed = run_test()
        sys.exit(0 if passed else 1)
    except Exception as e:
        print(f"\n  UNEXPECTED ERROR: {e}")
        traceback.print_exc()
        sys.exit(1)
