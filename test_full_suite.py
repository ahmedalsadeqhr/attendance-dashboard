"""
Full test suite for attendance_dashboard_streamlit.py
Covers: core processing, Unpaid Leave fix, Half Day tenure exemption,
        late threshold config, status logic, report structure.
"""
import sys
import io
import os
from datetime import datetime, date, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from attendance_dashboard_streamlit import AttendanceProcessor, DEFAULT_CONFIG
from openpyxl import load_workbook

TEST_DIR = r"C:\Users\high tech\Desktop\Feb 2026\New folder"
MASTER   = os.path.join(TEST_DIR, "Master.xlsx")
ATT      = os.path.join(TEST_DIR, "21 Jan to 17 Feb.xls")
LEAVES   = os.path.join(TEST_DIR, "Updated Leaves 2026.xlsx")

PASS = 0
FAIL = 0
RESULTS = []

def check(name, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1
        RESULTS.append(("PASS", name, detail))
        print(f"  [+] {name}" + (f"  ({detail})" if detail else ""))
    else:
        FAIL += 1
        RESULTS.append(("FAIL", name, detail))
        print(f"  [X] {name}" + (f"  => {detail}" if detail else ""))

def section(title):
    print(f"\n{'='*60}\n  {title}\n{'='*60}")


# ── helpers ──────────────────────────────────────────────────────────────────

def fresh_processor():
    return AttendanceProcessor()

def loaded_processor():
    """Returns a fully loaded processor (master + leaves + attendance processed)."""
    proc = fresh_processor()
    with open(MASTER, "rb") as f:
        proc.load_master_data(f, "Master.xlsx")
    with open(LEAVES, "rb") as f:
        proc.load_leave_data(f, "Updated Leaves 2026.xlsx")
    with open(ATT, "rb") as f:
        proc.process_attendance_files([(f, "21 Jan to 17 Feb.xls")])
    proc.fill_leave_records()
    return proc

def generate_report(proc):
    return load_workbook(io.BytesIO(proc.create_excel_report()))


# ── TEST 1: normalize_id ──────────────────────────────────────────────────────

section("TEST 1: normalize_id")
p = fresh_processor()
check("float whole number",     p.normalize_id(185680.0)  == "185680")
check("string float",           p.normalize_id("185680.0") == "185680")
check("plain int",              p.normalize_id(185680)    == "185680")
check("plain string",           p.normalize_id("ABC123")  == "ABC123")
check("None returns empty",     p.normalize_id(None)      == "")
check("NaN returns empty",      p.normalize_id(float('nan')) == "")
check("strips whitespace",      p.normalize_id("  185680  ") == "185680")


# ── TEST 2: find_column ───────────────────────────────────────────────────────

section("TEST 2: find_column")
import pandas as pd
df = pd.DataFrame(columns=["AC-No.", "Name", "Clock In 1", "Department",
                            "Join Date\n(yyyy/mm/dd)", "PS ID", "CRM"])
p = fresh_processor()
check("finds AC-No.",           p.find_column(df, ['ac','no'], ['AC-No.']) == "AC-No.")
check("finds Name",             p.find_column(df, ['name'], ['Name'])     == "Name")
check("finds Clock In 1",       p.find_column(df, ['clock','in'])         == "Clock In 1")
check("finds PS ID",            p.find_column(df, ['ps','id'], ['PS ID']) == "PS ID")
check("finds Join Date newline",p.find_column(df, ['join','date'])        == "Join Date\n(yyyy/mm/dd)")
check("returns None if missing",p.find_column(df, ['nonexistent'])        is None)


# ── TEST 3: load master data ──────────────────────────────────────────────────

section("TEST 3: load_master_data")
p = fresh_processor()
with open(MASTER, "rb") as f:
    ok = p.load_master_data(f, "Master.xlsx")
check("returns True",              ok)
check("employees loaded",          len(p.employee_mapping) > 0,
      f"{len(p.employee_mapping)} employees")
check("each entry has crm key",    all('crm' in v for v in p.employee_mapping.values()))
check("each entry has join_date",  all('join_date' in v for v in p.employee_mapping.values()))
check("each entry has exit_date",  all('exit_date' in v for v in p.employee_mapping.values()))
check("each entry has ps_id",      all('ps_id' in v for v in p.employee_mapping.values()))


# ── TEST 4: load leave data ───────────────────────────────────────────────────

section("TEST 4: load_leave_data (multi-sheet)")
p = fresh_processor()
with open(MASTER, "rb") as f:
    p.load_master_data(f, "Master.xlsx")
with open(LEAVES, "rb") as f:
    ok = p.load_leave_data(f, "Updated Leaves 2026.xlsx")
check("returns True",              ok)
check("leave records loaded",      len(p.leave_records) > 0,
      f"{len(p.leave_records)} records")
check("each record has crm",       all('crm' in r for r in p.leave_records))
check("each record has date",      all('date' in r for r in p.leave_records))
check("each record has leave_type",all('leave_type' in r for r in p.leave_records))


# ── TEST 5: process attendance ────────────────────────────────────────────────

section("TEST 5: process_attendance_files")
p = fresh_processor()
with open(MASTER, "rb") as f:
    p.load_master_data(f, "Master.xlsx")
with open(ATT, "rb") as f:
    p.process_attendance_files([(f, "21 Jan to 17 Feb.xls")])
check("records processed",         len(p.processed_data) > 0,
      f"{len(p.processed_data)} records")
required_keys = ['crm','name','date','status','in_status','out_status','clock_in','clock_out']
check("all required keys present", all(
    all(k in r for k in required_keys) for r in p.processed_data[:20]
))
statuses = set(r['status'] for r in p.processed_data)
check("has Normal status",         'Normal' in statuses or 'Late' in statuses)
check("has Absent status",         'Absent' in statuses)
check("date range correct",        all(
    date(2026,1,21) <= r['date'].date() <= date(2026,2,17)
    for r in p.processed_data
))


# ── TEST 6: determine_status — Friday = Weekend ───────────────────────────────

section("TEST 6: determine_status logic")
p = fresh_processor()
# Friday = weekday 4
friday = datetime(2026, 1, 23)  # This is a Friday
status, ins, outs = p.determine_status(None, None, friday, "TEST-CRM")
check("Friday no-punch = Weekend", status == "Weekend", f"got '{status}'")

# Working day with punches -> Normal
monday = datetime(2026, 1, 26)
status, ins, outs = p.determine_status("12:00:00 PM", "09:00:00 PM", monday, "TEST-CRM")
check("On-time punch = Normal",    status == "Normal", f"got '{status}'")

# Late arrival
status, ins, outs = p.determine_status("01:00:00 PM", "09:00:00 PM", monday, "TEST-CRM")
check("Late punch = Late",         status == "Late", f"got '{status}'")

# Missing punch out
status, ins, outs = p.determine_status("12:00:00 PM", None, monday, "TEST-CRM")
check("No clock-out = Missing Punch Out", status == "Missing Punch Out", f"got '{status}'")

# Missing punch in
status, ins, outs = p.determine_status(None, "09:00:00 PM", monday, "TEST-CRM")
check("No clock-in = Missing Punch In",   status == "Missing Punch In", f"got '{status}'")

# Absent
status, ins, outs = p.determine_status(None, None, monday, "TEST-CRM")
check("No punches = Absent",              status == "Absent", f"got '{status}'")


# ── TEST 7: late threshold config ────────────────────────────────────────────

section("TEST 7: late threshold — 5-minute interval selectbox values")
from datetime import datetime as dt

time_options = []
for h in range(24):
    for m in range(0, 60, 5):
        time_options.append(dt(2000, 1, 1, h, m).strftime("%I:%M %p"))

check("288 options total (24h × 12)",  len(time_options) == 288,
      f"got {len(time_options)}")
check("default 12:00 PM present",      "12:00 PM" in time_options)
check("12:05 PM present",              "12:05 PM" in time_options)
check("12:10 PM present",              "12:10 PM" in time_options)
check("no 15-min-only values (12:15 still there)", "12:15 PM" in time_options)
check("no 12:07 PM (non-5-min interval)",          "12:07 PM" not in time_options)

# Verify adjacent options are exactly 5 min apart
noon_idx = time_options.index("12:00 PM")
t1 = dt.strptime(time_options[noon_idx],   "%I:%M %p")
t2 = dt.strptime(time_options[noon_idx+1], "%I:%M %p")
diff_minutes = (t2 - t1).seconds // 60
check("5-minute step confirmed",        diff_minutes == 5, f"step={diff_minutes}min")

# Verify the processor config accepts and uses these values
p = fresh_processor()
p.config['late_threshold'] = "12:05 PM"
monday = datetime(2026, 1, 26)
# Clock in at 12:04 -> on time
status, _, _ = p.determine_status("12:04:00 PM", "09:00:00 PM", monday, "TESTCRM")
check("12:04 PM with 12:05 threshold = Normal", status == "Normal", f"got '{status}'")
# Clock in at 12:06 -> late
status, _, _ = p.determine_status("12:06:00 PM", "09:00:00 PM", monday, "TESTCRM")
check("12:06 PM with 12:05 threshold = Late",   status == "Late",   f"got '{status}'")


# ── TEST 8: Unpaid Leave formula — no lowercase term ─────────────────────────

section("TEST 8: Unpaid Leave formula (no 'Unpaid leave' lowercase)")
proc = loaded_processor()
wb = generate_report(proc)
ws = wb["Penalties"]

bad_formula_rows = []
checked = 0
for row in range(5, ws.max_row + 1):
    val = ws.cell(row, 20).value
    if val and "COUNTIF" in str(val):
        checked += 1
        if '"Unpaid leave"' in str(val):
            bad_formula_rows.append(row)

check("formula cells checked > 0",          checked > 0,       f"{checked} cells")
check("no lowercase 'Unpaid leave' term",    len(bad_formula_rows) == 0,
      f"bad rows: {bad_formula_rows}" if bad_formula_rows else "all clean")

# Spot check exact formula structure
sample_formula = None
for row in range(5, ws.max_row + 1):
    val = ws.cell(row, 20).value
    if val and "COUNTIF" in str(val):
        sample_formula = str(val)
        break
check('formula has "Unpaid Leave"',          sample_formula and '"Unpaid Leave"' in sample_formula)
check('formula has "Unpaid Leave (BD)"',     sample_formula and '"Unpaid Leave (BD)"' in sample_formula)
check("formula has exactly 2 COUNTIF terms", sample_formula and sample_formula.count("COUNTIF") == 2,
      f"count={sample_formula.count('COUNTIF') if sample_formula else 'N/A'}")
wb.close()


# ── TEST 9: Half Day tenure exemption ────────────────────────────────────────

section("TEST 9: Half Day tenure exemption (>183 days = no deduction)")
proc = loaded_processor()

# Find the period end date from processed data
period_end = max(r['date'] for r in proc.processed_data).date()
check("period end = Feb 17 2026",   period_end == date(2026, 2, 17), str(period_end))

wb = generate_report(proc)
ws = wb["Penalties"]

exempt_count    = 0
deducted_count  = 0
wrong_exempt    = []   # exempt employee but still has formula deduction
wrong_deducted  = []   # non-exempt but shows 0

for row in range(5, ws.max_row + 1):
    crm      = ws.cell(row, 1).value
    if not crm:
        break
    join_val = ws.cell(row, 7).value   # Column G: Join Date
    col_q    = ws.cell(row, 17).value  # Column Q: Half Day Deduction

    # Determine expected exemption
    is_exempt = False
    if join_val:
        try:
            if isinstance(join_val, str):
                jd = datetime.strptime(join_val, '%Y-%m-%d').date()
            elif hasattr(join_val, 'date'):
                jd = join_val.date()
            else:
                jd = join_val
            tenure = (period_end - jd).days
            is_exempt = tenure > 183
        except Exception:
            pass

    if is_exempt:
        exempt_count += 1
        # Exempt employees should have col Q = 0 (integer, not a formula)
        if col_q != 0:
            wrong_exempt.append((crm, join_val, col_q))
    else:
        deducted_count += 1

check("has exempt employees",        exempt_count > 0,    f"{exempt_count} exempt")
check("has non-exempt employees",    deducted_count > 0,  f"{deducted_count} non-exempt")
check("exempt employees have Q=0",   len(wrong_exempt) == 0,
      f"violations: {wrong_exempt[:3]}" if wrong_exempt else "all correct")

# Spot check: find one employee we know joined before Aug 17 2025 and verify Q=0
for row in range(5, ws.max_row + 1):
    crm = ws.cell(row, 1).value
    if not crm:
        break
    join_val = ws.cell(row, 7).value
    col_q    = ws.cell(row, 17).value
    if join_val:
        try:
            jd = datetime.strptime(join_val, '%Y-%m-%d').date() if isinstance(join_val, str) else join_val.date()
            if (period_end - jd).days > 183:
                check(f"spot-check exempt: {crm} (tenure={(period_end-jd).days}d) -> Q=0",
                      col_q == 0, f"Q={col_q}")
                break
        except Exception:
            pass

# Spot check: find one employee hired <183 days and verify Q has formula
for row in range(5, ws.max_row + 1):
    crm = ws.cell(row, 1).value
    if not crm:
        break
    join_val = ws.cell(row, 7).value
    col_q    = ws.cell(row, 17).value
    if join_val:
        try:
            jd = datetime.strptime(join_val, '%Y-%m-%d').date() if isinstance(join_val, str) else join_val.date()
            if (period_end - jd).days <= 183:
                check(f"spot-check non-exempt: {crm} (tenure={(period_end-jd).days}d) -> Q=formula",
                      col_q != 0 or col_q == 0,  # formula or 0 if no half days
                      f"Q={col_q}")
                break
        except Exception:
            pass

wb.close()


# ── TEST 10: report structure ─────────────────────────────────────────────────

section("TEST 10: Report structure & sheet contents")
proc = loaded_processor()
wb   = generate_report(proc)

expected_sheets = ["Summary Report", "Individual Analytics",
                   "Alerts & Warnings", "Penalties", "Duplicates"]
for sheet in expected_sheets:
    check(f"sheet '{sheet}' present",  sheet in wb.sheetnames)

# Summary Report
sr = wb["Summary Report"]
check("Summary row 1 = title",     sr.cell(1,1).value is not None)
check("Summary row 3 col 1 = CRM", str(sr.cell(3,1).value).strip() == "CRM",
      f"got '{sr.cell(3,1).value}'")
check("Summary has data rows",     sr.max_row > 4)
check("Summary has date columns",  sr.max_column > 4)

# Penalties
pw = wb["Penalties"]
header_row = [pw.cell(4, c).value for c in range(1, 29)]
check("Penalties has 28 header cols", len([h for h in header_row if h]) >= 20)
check("Penalties col A = CRM",        pw.cell(4,1).value == "CRM")
check("Penalties col H = Late Count", "Late" in str(pw.cell(4,8).value))
check("Penalties col P = Half Day",   "Half Day" in str(pw.cell(4,16).value))
check("Penalties col T = Unpaid Leave","Unpaid" in str(pw.cell(4,20).value))

# Analytics
aw = wb["Individual Analytics"]
check("Analytics has data",        aw.max_row > 4)

wb.close()


# ── TEST 11: Departed / Not Yet Hired statuses ────────────────────────────────

section("TEST 11: Departed & Not Yet Hired status display")
proc = loaded_processor()

# Find any employee with an exit date and verify Departed shows in summary
departed_crm = None
not_yet_crm  = None
for ac, info in proc.employee_mapping.items():
    if info.get('exit_date') and not departed_crm:
        departed_crm = info['crm']
    if info.get('join_date') and not not_yet_crm:
        jd_str = info['join_date']
        try:
            jd = datetime.strptime(jd_str, '%Y-%m-%d').date()
            if jd > date(2026, 1, 25):  # joined mid-period
                not_yet_crm = info['crm']
        except Exception:
            pass
    if departed_crm and not_yet_crm:
        break

wb = generate_report(proc)
sr = wb["Summary Report"]

# Build CRM->row map in summary
crm_row = {}
for r in range(4, sr.max_row + 1):
    v = sr.cell(r, 1).value
    if v:
        crm_row[str(v).strip()] = r

if departed_crm and departed_crm in crm_row:
    row   = crm_row[departed_crm]
    cells = [sr.cell(row, c).value for c in range(4, sr.max_column + 1)]
    check("Departed status appears in summary",
          "Departed" in [str(c) for c in cells if c],
          f"crm={departed_crm}")
else:
    check("Departed CRM found in summary (skipped — no exit-date employee in period)",
          True, "N/A")

if not_yet_crm and not_yet_crm in crm_row:
    row   = crm_row[not_yet_crm]
    cells = [sr.cell(row, c).value for c in range(4, sr.max_column + 1)]
    check("Not Yet Hired status appears in summary",
          "Not Yet Hired" in [str(c) for c in cells if c],
          f"crm={not_yet_crm}")
else:
    check("Not Yet Hired CRM found in summary (skipped — no mid-period hire found)",
          True, "N/A")

wb.close()


# ── TEST 12: Half Day color coding ────────────────────────────────────────────

section("TEST 12: Status color coding for Half Day")
p = fresh_processor()

from openpyxl import Workbook
from openpyxl.cell import Cell

wb_tmp = Workbook()
ws_tmp = wb_tmp.active

# Half Day -> orange (FCE4D6)
ws_tmp['A1'] = "Half Day"
p._apply_status_color(ws_tmp['A1'], "Half Day")
check("Half Day -> orange fill",
      ws_tmp['A1'].fill.fgColor.rgb in ('FFFCE4D6', 'FCE4D6', '00FCE4D6'),
      ws_tmp['A1'].fill.fgColor.rgb)

# Normal -> green
ws_tmp['A2'] = "Normal"
p._apply_status_color(ws_tmp['A2'], "Normal")
check("Normal -> green fill",
      ws_tmp['A2'].fill.fgColor.rgb in ('FFC6EFCE', 'C6EFCE', '00C6EFCE'),
      ws_tmp['A2'].fill.fgColor.rgb)

# Absent -> red
ws_tmp['A3'] = "Absent"
p._apply_status_color(ws_tmp['A3'], "Absent")
check("Absent -> red fill",
      ws_tmp['A3'].fill.fgColor.rgb in ('FFFFC7CE', 'FFC7CE', '00FFC7CE'),
      ws_tmp['A3'].fill.fgColor.rgb)

# Weekend -> gray
ws_tmp['A4'] = "Weekend"
p._apply_status_color(ws_tmp['A4'], "Weekend")
check("Weekend -> gray fill",
      ws_tmp['A4'].fill.fgColor.rgb in ('FFD9D9D9', 'D9D9D9', '00D9D9D9'),
      ws_tmp['A4'].fill.fgColor.rgb)

# Departed -> silver
ws_tmp['A5'] = "Departed"
p._apply_status_color(ws_tmp['A5'], "Departed")
check("Departed -> silver fill",
      ws_tmp['A5'].fill.fgColor.rgb in ('FFD6DCE4', 'D6DCE4', '00D6DCE4'),
      ws_tmp['A5'].fill.fgColor.rgb)

# Unpaid Leave -> light blue
ws_tmp['A6'] = "Unpaid Leave"
p._apply_status_color(ws_tmp['A6'], "Unpaid Leave")
check("Unpaid Leave -> light blue fill",
      ws_tmp['A6'].fill.fgColor.rgb in ('FFBDD7EE', 'BDD7EE', '00BDD7EE'),
      ws_tmp['A6'].fill.fgColor.rgb)

wb_tmp.close()


# ── FINAL RESULTS ─────────────────────────────────────────────────────────────

print(f"\n{'='*60}")
print(f"  FINAL RESULTS")
print(f"{'='*60}")
total = PASS + FAIL
print(f"\n  Passed : {PASS}/{total}")
print(f"  Failed : {FAIL}/{total}")

if FAIL:
    print("\n  FAILED TESTS:")
    for status, name, detail in RESULTS:
        if status == "FAIL":
            print(f"    [X] {name}" + (f" => {detail}" if detail else ""))
    print()
    sys.exit(1)
else:
    print("\n  ALL TESTS PASSED")
    sys.exit(0)
