"""
Attendance Dashboard Desktop Application v2.2
Professional attendance processing with graphical interface
PRODUCTION-READY VERSION - All bugs fixed and enhanced
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import os
import sys
import json
import re
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path


def read_excel_file(filepath):
    """Read Excel file with automatic engine detection for .xls and .xlsx files."""
    filepath_str = str(filepath).lower()
    if filepath_str.endswith('.xls') and not filepath_str.endswith('.xlsx'):
        # Old .xls format - use xlrd engine
        return pd.read_excel(filepath, engine='xlrd')
    else:
        # .xlsx format - use default openpyxl engine
        return pd.read_excel(filepath)


def setup_logger():
    """Setup application logger with file and console handlers."""
    log_dir = Path.home() / ".attendance_dashboard" / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / "app.log"

    logger = logging.getLogger('AttendanceDashboard')
    logger.setLevel(logging.DEBUG)

    # File handler (rotating, max 5MB, keep 3 backups)
    fh = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=3)
    fh.setLevel(logging.DEBUG)

    # Console handler
    ch = logging.StreamHandler()
    ch.setLevel(logging.WARNING)

    # Formatter
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)

    logger.addHandler(fh)
    logger.addHandler(ch)

    return logger


class AppConfig:
    """Manage application configuration with persistence."""

    DEFAULT_CONFIG = {
        'work_start_time': '12:00 PM',  # Updated per policy
        'work_end_time': '9:00 PM',     # Updated per policy
        'late_threshold': '12:00 PM',   # Clock-in after this time is considered "Late"
        'off_days': [4],  # Friday (0=Monday, 4=Friday)
        'alert_thresholds': {
            'high_absences': 3,
            'frequent_late': 5,
            'missing_punches': 5,
            'low_attendance_rate': 75.0
        },
        'penalties': {
            'late_1st': 100,             # EGP - 1st late occurrence
            'late_2nd': 200,             # EGP - 2nd late occurrence
            'late_3rd': 500,             # EGP - 3rd late occurrence
            'late_4th_plus': 500,        # EGP - 4th and subsequent
            'missing_punch_deduction': 0.5,   # days - per occurrence after threshold
            'missing_punch_threshold': 3,     # trigger deduction after this count
            'missing_punch_warning_threshold': 6,  # trigger warning after this count
            'early_departure_deduction': 0.5,     # days
            'absence_deduction': 2,               # days per absence
            'currency': 'EGP'
        }
    }

    def __init__(self):
        self.config_file = Path.home() / ".attendance_dashboard" / "config.json"
        self.config = self.load_config()

    def load_config(self):
        """Load config from file or create default."""
        if self.config_file.exists():
            try:
                with open(self.config_file, 'r') as f:
                    loaded = json.load(f)
                    # Merge with defaults to ensure all keys exist
                    config = self.DEFAULT_CONFIG.copy()
                    config.update(loaded)
                    return config
            except Exception as e:
                print(f"Error loading config: {e}. Using defaults.")
        return self.DEFAULT_CONFIG.copy()

    def save_config(self):
        """Save config to file."""
        try:
            self.config_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.config_file, 'w') as f:
                json.dump(self.config, f, indent=2)
        except Exception as e:
            print(f"Error saving config: {e}")

    def get(self, key, default=None):
        """Get config value."""
        return self.config.get(key, default)

    def set(self, key, value):
        """Set config value and save."""
        self.config[key] = value
        self.save_config()


class AttendanceDashboard:
    # Color scheme - 51Talk brand colors
    COLORS = {
        'bg': '#F5F5F5',
        'card': '#FFFFFF',
        'primary': '#00A0DC',           # 51Talk Blue
        'primary_hover': '#0088BC',
        'primary_disabled': '#9CA3AF',
        'accent': '#FFD700',            # 51Talk Yellow
        'success': '#10B981',
        'error': '#EF4444',
        'text': '#1F2937',
        'text_secondary': '#6B7280',
        'border': '#E5E7EB',
    }

    def __init__(self):
        # Set appearance mode and theme
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        # Create main window
        self.root = ctk.CTk()
        self.root.title("Attendance Dashboard v2.2")
        self.root.geometry("600x700")
        self.root.resizable(False, False)

        # Initialize logger and config
        self.logger = setup_logger()
        self.logger.info("Application started")
        self.config = AppConfig()

        # Data storage
        self.master_file = None
        self.attendance_files = []
        self.leave_file = None
        self.employee_mapping = {}
        self.leave_records = []
        self.processed_data = []
        self.conflict_records = []  # Track leave vs attendance conflicts
        self.last_directory = str(Path.home() / "Documents")

        # Filter variables
        self.filter_frame = None
        self.department_vars = {}
        self.crm_vars = {}
        self.available_departments = []
        self.available_crms = []
        self.filter_status = None
        self.dept_select_all_var = None
        self.crm_select_all_var = None

        # Create UI - new order for better hierarchy
        self.create_header()
        self.create_generate_section()
        self.create_upload_section()
        self.create_filter_section()
        self.create_progress_section()
        self.create_status_section()
        self.create_footer()
        self.setup_keyboard_shortcuts()

        self.logger.info("UI initialized successfully")

    def create_header(self):
        """Create minimal header with title and settings"""
        header_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        header_frame.pack(fill='x', padx=20, pady=(15, 10))

        # Title on left
        title = ctk.CTkLabel(
            header_frame,
            text="Attendance Dashboard",
            font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
            text_color=self.COLORS['text']
        )
        title.pack(side='left')

        # Version badge
        version = ctk.CTkLabel(
            header_frame,
            text="v2.2",
            font=ctk.CTkFont(size=10),
            fg_color=self.COLORS['border'],
            text_color=self.COLORS['text_secondary'],
            corner_radius=4,
            padx=8,
            pady=2
        )
        version.pack(side='left', padx=(10, 0))

        # Settings button on right
        settings_btn = ctk.CTkButton(
            header_frame,
            text="Settings",
            font=ctk.CTkFont(size=12),
            fg_color=self.COLORS['card'],
            text_color=self.COLORS['text_secondary'],
            hover_color=self.COLORS['border'],
            corner_radius=6,
            width=80,
            height=30,
            command=self.show_settings_dialog
        )
        settings_btn.pack(side='right')

    def create_generate_section(self):
        """Create prominent generate button at top"""
        gen_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        gen_frame.pack(fill='x', padx=20, pady=(5, 15))

        self.generate_btn = ctk.CTkButton(
            gen_frame,
            text="GENERATE REPORT",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            fg_color=self.COLORS['primary_disabled'],
            hover_color=self.COLORS['primary_hover'],
            text_color='white',
            corner_radius=8,
            height=50,
            command=self.generate_report,
            state='disabled'
        )
        self.generate_btn.pack(fill='x')

        # Shortcut hint
        hint = ctk.CTkLabel(
            gen_frame,
            text="Ctrl+G",
            font=ctk.CTkFont(size=10),
            text_color=self.COLORS['text_secondary']
        )
        hint.pack(pady=(5, 0))

    def create_upload_section(self):
        """Create compact file selection rows"""
        # Section container with white background and rounded corners
        container = ctk.CTkFrame(self.root, fg_color=self.COLORS['card'], corner_radius=10)
        container.pack(fill='x', padx=20, pady=(0, 10))
        self.upload_container = container  # Store reference for filter section

        # Section title
        title_frame = ctk.CTkFrame(container, fg_color="transparent")
        title_frame.pack(fill='x', padx=15, pady=(15, 10))

        ctk.CTkLabel(
            title_frame,
            text="Select Files",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=self.COLORS['text']
        ).pack(side='left')

        # File rows
        files_frame = ctk.CTkFrame(container, fg_color="transparent")
        files_frame.pack(fill='x', padx=15, pady=(0, 15))

        # Master Data row
        self.master_status = self._create_file_row(
            files_frame, "Master Data", "Required", self.select_master_data, 0
        )

        # Attendance row
        self.attendance_status = self._create_file_row(
            files_frame, "Attendance", "Required", self.select_attendance_files, 1
        )

        # Leave Sheet row
        self.leave_status = self._create_file_row(
            files_frame, "Leave Sheet", "Optional", self.select_leave_sheet, 2
        )

    def _create_file_row(self, parent, label_text, hint_text, command, row):
        """Create a single compact file selection row"""
        row_frame = ctk.CTkFrame(parent, fg_color="transparent")
        row_frame.pack(fill='x', pady=6)

        # Label
        label = ctk.CTkLabel(
            row_frame,
            text=label_text,
            font=ctk.CTkFont(size=12),
            text_color=self.COLORS['text'],
            width=100,
            anchor='w'
        )
        label.pack(side='left')

        # Browse button with accent styling
        browse_btn = ctk.CTkButton(
            row_frame,
            text="Browse",
            font=ctk.CTkFont(size=11),
            fg_color=self.COLORS['bg'],
            text_color=self.COLORS['text'],
            hover_color=self.COLORS['border'],
            corner_radius=6,
            width=80,
            height=28,
            command=command
        )
        browse_btn.pack(side='left', padx=(5, 15))

        # Status label
        status = ctk.CTkLabel(
            row_frame,
            text=hint_text,
            font=ctk.CTkFont(size=11),
            text_color=self.COLORS['text_secondary'],
            anchor='w'
        )
        status.pack(side='left', fill='x', expand=True)

        return status

    def create_filter_section(self):
        """Create filter UI section (initially hidden)"""
        # Main filter container - initially hidden
        self.filter_frame = ctk.CTkFrame(self.root, fg_color=self.COLORS['card'], corner_radius=10)
        # Don't pack yet - will be shown after master data is loaded

        # Title row with Reset button
        title_frame = ctk.CTkFrame(self.filter_frame, fg_color="transparent")
        title_frame.pack(fill='x', padx=15, pady=(15, 10))

        ctk.CTkLabel(
            title_frame,
            text="Filter Employees",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=self.COLORS['text']
        ).pack(side='left')

        reset_btn = ctk.CTkButton(
            title_frame,
            text="Reset",
            font=ctk.CTkFont(size=11),
            fg_color=self.COLORS['bg'],
            text_color=self.COLORS['text'],
            hover_color=self.COLORS['border'],
            corner_radius=6,
            width=60,
            height=24,
            command=self.reset_filters
        )
        reset_btn.pack(side='right')

        # Filter content frame
        content_frame = ctk.CTkFrame(self.filter_frame, fg_color="transparent")
        content_frame.pack(fill='x', padx=15, pady=(0, 10))

        # Department filter column
        dept_frame = ctk.CTkFrame(content_frame, fg_color=self.COLORS['bg'], corner_radius=8)
        dept_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))

        self.dept_header = ctk.CTkLabel(
            dept_frame,
            text="Department (0)",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=self.COLORS['text']
        )
        self.dept_header.pack(pady=(8, 5))

        # Select All checkbox for departments
        self.dept_select_all_var = ctk.BooleanVar(value=True)
        dept_select_all = ctk.CTkCheckBox(
            dept_frame,
            text="Select All",
            variable=self.dept_select_all_var,
            command=self.toggle_all_departments,
            checkbox_width=18,
            checkbox_height=18,
            font=ctk.CTkFont(size=11),
            fg_color=self.COLORS['primary']
        )
        dept_select_all.pack(anchor='w', padx=10, pady=2)

        # Scrollable frame for department checkboxes
        self.dept_scroll = ctk.CTkScrollableFrame(
            dept_frame,
            fg_color="transparent",
            height=80
        )
        self.dept_scroll.pack(fill='both', expand=True, padx=5, pady=(0, 8))

        # CRM filter column
        crm_frame = ctk.CTkFrame(content_frame, fg_color=self.COLORS['bg'], corner_radius=8)
        crm_frame.pack(side='left', fill='both', expand=True, padx=(5, 0))

        self.crm_header = ctk.CTkLabel(
            crm_frame,
            text="CRM (0)",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=self.COLORS['text']
        )
        self.crm_header.pack(pady=(8, 5))

        # Select All checkbox for CRMs
        self.crm_select_all_var = ctk.BooleanVar(value=True)
        crm_select_all = ctk.CTkCheckBox(
            crm_frame,
            text="Select All",
            variable=self.crm_select_all_var,
            command=self.toggle_all_crms,
            checkbox_width=18,
            checkbox_height=18,
            font=ctk.CTkFont(size=11),
            fg_color=self.COLORS['primary']
        )
        crm_select_all.pack(anchor='w', padx=10, pady=2)

        # Scrollable frame for CRM checkboxes
        self.crm_scroll = ctk.CTkScrollableFrame(
            crm_frame,
            fg_color="transparent",
            height=80
        )
        self.crm_scroll.pack(fill='both', expand=True, padx=5, pady=(0, 8))

        # Filter status label
        self.filter_status = ctk.CTkLabel(
            self.filter_frame,
            text="",
            font=ctk.CTkFont(size=10),
            text_color=self.COLORS['text_secondary']
        )
        self.filter_status.pack(pady=(0, 10))

    def populate_filters(self):
        """Populate filter checkboxes with departments and CRMs from master data"""
        if not self.employee_mapping:
            return

        # Extract unique departments and CRMs
        self.available_departments = sorted(set(
            info['department'] for info in self.employee_mapping.values()
            if info['department']
        ))
        self.available_crms = sorted(set(
            info['crm'] for info in self.employee_mapping.values()
            if info['crm']
        ))

        # Clear existing checkboxes
        for widget in self.dept_scroll.winfo_children():
            widget.destroy()
        for widget in self.crm_scroll.winfo_children():
            widget.destroy()

        self.department_vars = {}
        self.crm_vars = {}

        # Create department checkboxes
        for dept in self.available_departments:
            var = ctk.BooleanVar(value=True)
            self.department_vars[dept] = var
            cb = ctk.CTkCheckBox(
                self.dept_scroll,
                text=dept[:20] + "..." if len(dept) > 20 else dept,
                variable=var,
                command=self.update_filter_counts,
                checkbox_width=16,
                checkbox_height=16,
                font=ctk.CTkFont(size=10),
                fg_color=self.COLORS['primary']
            )
            cb.pack(anchor='w', pady=1)

        # Create CRM checkboxes
        for crm in self.available_crms:
            var = ctk.BooleanVar(value=True)
            self.crm_vars[crm] = var
            cb = ctk.CTkCheckBox(
                self.crm_scroll,
                text=crm[:18] + "..." if len(crm) > 18 else crm,
                variable=var,
                command=self.update_filter_counts,
                checkbox_width=16,
                checkbox_height=16,
                font=ctk.CTkFont(size=10),
                fg_color=self.COLORS['primary']
            )
            cb.pack(anchor='w', pady=1)

        # Update headers
        self.dept_header.configure(text=f"Department ({len(self.available_departments)})")
        self.crm_header.configure(text=f"CRM ({len(self.available_crms)})")

        # Reset select all checkboxes
        self.dept_select_all_var.set(True)
        self.crm_select_all_var.set(True)

        # Show filter frame
        self.filter_frame.pack(fill='x', padx=20, pady=(0, 10), after=self.upload_container)

        # Update counts
        self.update_filter_counts()

    def toggle_all_departments(self):
        """Toggle all department checkboxes"""
        select_all = self.dept_select_all_var.get()
        for var in self.department_vars.values():
            var.set(select_all)
        self.update_filter_counts()

    def toggle_all_crms(self):
        """Toggle all CRM checkboxes"""
        select_all = self.crm_select_all_var.get()
        for var in self.crm_vars.values():
            var.set(select_all)
        self.update_filter_counts()

    def update_filter_counts(self):
        """Update the filter status label showing selected counts"""
        # Get selected departments and CRMs
        selected_depts = {dept for dept, var in self.department_vars.items() if var.get()}
        selected_crms = {crm for crm, var in self.crm_vars.items() if var.get()}

        # Count employees that match filters
        filtered_count = 0
        total_count = len(self.employee_mapping)

        for ac_no, info in self.employee_mapping.items():
            dept = info.get('department', '')
            crm = info.get('crm', '')

            # Employee must match selected department AND be in selected CRMs
            dept_match = not self.available_departments or not dept or dept in selected_depts
            crm_match = not self.available_crms or not crm or crm in selected_crms

            if dept_match and crm_match:
                filtered_count += 1

        self.filter_status.configure(text=f"Showing {filtered_count} of {total_count} employees")

        # Update select all checkboxes state
        all_depts_selected = all(var.get() for var in self.department_vars.values()) if self.department_vars else True
        all_crms_selected = all(var.get() for var in self.crm_vars.values()) if self.crm_vars else True

        self.dept_select_all_var.set(all_depts_selected)
        self.crm_select_all_var.set(all_crms_selected)

    def get_filtered_employee_mapping(self):
        """Return employee mapping filtered by selected departments and CRMs"""
        if not self.department_vars and not self.crm_vars:
            return self.employee_mapping

        selected_depts = {dept for dept, var in self.department_vars.items() if var.get()}
        selected_crms = {crm for crm, var in self.crm_vars.items() if var.get()}

        filtered = {}
        for ac_no, info in self.employee_mapping.items():
            dept = info.get('department', '')
            crm = info.get('crm', '')

            # Employee must match selected department AND be in selected CRMs
            dept_match = not self.available_departments or not dept or dept in selected_depts
            crm_match = not self.available_crms or not crm or crm in selected_crms

            if dept_match and crm_match:
                filtered[ac_no] = info

        return filtered

    def reset_filters(self):
        """Reset all filter checkboxes to selected"""
        for var in self.department_vars.values():
            var.set(True)
        for var in self.crm_vars.values():
            var.set(True)
        self.dept_select_all_var.set(True)
        self.crm_select_all_var.set(True)
        self.update_filter_counts()

    def create_progress_section(self):
        """Create compact progress section"""
        progress_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        progress_frame.pack(fill='x', padx=20, pady=(5, 5))

        # Progress bar with CTk styling
        self.progress_bar = ctk.CTkProgressBar(
            progress_frame,
            progress_color=self.COLORS['primary'],
            fg_color=self.COLORS['border'],
            corner_radius=5,
            height=8
        )
        self.progress_bar.pack(fill='x')
        self.progress_bar.set(0)

        self.progress_label = ctk.CTkLabel(
            progress_frame,
            text="Ready",
            font=ctk.CTkFont(size=11),
            text_color=self.COLORS['text_secondary']
        )
        self.progress_label.pack(pady=(6, 0))

    def create_status_section(self):
        """Create compact status log section"""
        # Container with rounded corners
        container = ctk.CTkFrame(self.root, fg_color=self.COLORS['card'], corner_radius=10)
        container.pack(fill='both', expand=True, padx=20, pady=(5, 10))

        # Title
        title_frame = ctk.CTkFrame(container, fg_color="transparent")
        title_frame.pack(fill='x', padx=12, pady=(10, 5))

        ctk.CTkLabel(
            title_frame,
            text="Log",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=self.COLORS['text_secondary']
        ).pack(side='left')

        # Scrollable textbox
        self.status_text = ctk.CTkTextbox(
            container,
            font=ctk.CTkFont(family="Consolas", size=11),
            fg_color=self.COLORS['bg'],
            text_color=self.COLORS['text'],
            corner_radius=8,
            height=120
        )
        self.status_text.pack(fill='both', expand=True, padx=12, pady=(0, 12))

        self.log_message("Ready to process attendance files.")

    def create_footer(self):
        """Create footer with Clear All button"""
        footer_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        footer_frame.pack(fill='x', padx=20, pady=(0, 15))

        clear_btn = ctk.CTkButton(
            footer_frame,
            text="Clear All",
            font=ctk.CTkFont(size=11),
            fg_color="transparent",
            text_color=self.COLORS['text_secondary'],
            hover_color=self.COLORS['border'],
            corner_radius=6,
            width=80,
            height=28,
            command=self.clear_all
        )
        clear_btn.pack(side='right')

    def setup_keyboard_shortcuts(self):
        """Setup all keyboard shortcuts"""
        self.root.bind('<Control-g>', lambda e: self._safe_generate())
        self.root.bind('<Control-G>', lambda e: self._safe_generate())
        self.root.bind('<Control-comma>', lambda e: self.show_settings_dialog())
        self.root.bind('<Control-r>', lambda e: self.clear_all())
        self.root.bind('<Control-R>', lambda e: self.clear_all())
        self.root.bind('<Control-m>', lambda e: self.select_master_data())
        self.root.bind('<Control-M>', lambda e: self.select_master_data())
        self.root.bind('<Control-a>', lambda e: self.select_attendance_files())
        self.root.bind('<Control-A>', lambda e: self.select_attendance_files())
        self.root.bind('<Control-l>', lambda e: self.select_leave_sheet())
        self.root.bind('<Control-L>', lambda e: self.select_leave_sheet())

    def _safe_generate(self):
        """Generate report only if button is enabled"""
        if self.generate_btn.cget('state') != 'disabled':
            self.generate_report()

    def log_message(self, message, level='info'):
        """Add a message to the status log"""
        timestamp = datetime.now().strftime("%H:%M")

        if level == 'success':
            prefix = "[OK]"
        elif level == 'error':
            prefix = "[ERR]"
        elif level == 'warning':
            prefix = "[!]"
        else:
            prefix = "[i]"

        self.status_text.insert('end', f"{timestamp} {prefix} {message}\n")
        self.status_text.see('end')
        self.root.update_idletasks()

    def show_settings_dialog(self):
        """Show settings configuration dialog"""
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Application Settings")
        dialog.geometry("480x580")
        dialog.transient(self.root)
        dialog.grab_set()

        # Main scrollable container
        main_frame = ctk.CTkScrollableFrame(dialog, fg_color="transparent")
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)

        # Work Hours Section
        ctk.CTkLabel(main_frame, text="Work Hours", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(10, 10))

        work_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        work_frame.pack(pady=5, fill='x')

        # Start Time
        row1 = ctk.CTkFrame(work_frame, fg_color="transparent")
        row1.pack(fill='x', pady=3)
        ctk.CTkLabel(row1, text="Start Time:", width=120).pack(side='left')
        self.start_entry = ctk.CTkEntry(row1, width=120)
        self.start_entry.insert(0, self.config.get('work_start_time'))
        self.start_entry.pack(side='left', padx=5)
        ctk.CTkLabel(row1, text="(e.g., 9:00 AM)", text_color=self.COLORS['text_secondary']).pack(side='left')

        # End Time
        row2 = ctk.CTkFrame(work_frame, fg_color="transparent")
        row2.pack(fill='x', pady=3)
        ctk.CTkLabel(row2, text="End Time:", width=120).pack(side='left')
        self.end_entry = ctk.CTkEntry(row2, width=120)
        self.end_entry.insert(0, self.config.get('work_end_time'))
        self.end_entry.pack(side='left', padx=5)
        ctk.CTkLabel(row2, text="(e.g., 5:00 PM)", text_color=self.COLORS['text_secondary']).pack(side='left')

        # Late Threshold
        row3 = ctk.CTkFrame(work_frame, fg_color="transparent")
        row3.pack(fill='x', pady=3)
        ctk.CTkLabel(row3, text="Late Threshold:", width=120).pack(side='left')
        self.late_entry = ctk.CTkEntry(row3, width=120)
        self.late_entry.insert(0, self.config.get('late_threshold', '12:00 PM'))
        self.late_entry.pack(side='left', padx=5)
        ctk.CTkLabel(row3, text="(Clock-in after = Late)", text_color=self.COLORS['text_secondary']).pack(side='left')

        # OFF Days Section
        ctk.CTkLabel(main_frame, text="OFF Days (Weekly)", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(20, 10))

        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        off_days = self.config.get('off_days', [4])
        day_vars = {}

        days_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        days_frame.pack(pady=5)

        for i, day in enumerate(days):
            var = ctk.BooleanVar(value=(i in off_days))
            day_vars[i] = var
            ctk.CTkCheckBox(days_frame, text=day, variable=var,
                           checkbox_width=20, checkbox_height=20,
                           fg_color=self.COLORS['primary']).pack(anchor='w', pady=2)

        # Alert Thresholds Section
        ctk.CTkLabel(main_frame, text="Alert Thresholds", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(20, 10))

        thresholds = self.config.get('alert_thresholds', {})
        threshold_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        threshold_frame.pack(pady=5, fill='x')

        # High Absences
        t_row1 = ctk.CTkFrame(threshold_frame, fg_color="transparent")
        t_row1.pack(fill='x', pady=3)
        ctk.CTkLabel(t_row1, text="High Absences (days):", width=180).pack(side='left')
        absences_entry = ctk.CTkEntry(t_row1, width=80)
        absences_entry.insert(0, str(thresholds.get('high_absences', 3)))
        absences_entry.pack(side='left')

        # Frequent Late
        t_row2 = ctk.CTkFrame(threshold_frame, fg_color="transparent")
        t_row2.pack(fill='x', pady=3)
        ctk.CTkLabel(t_row2, text="Frequent Late (count):", width=180).pack(side='left')
        late_threshold_entry = ctk.CTkEntry(t_row2, width=80)
        late_threshold_entry.insert(0, str(thresholds.get('frequent_late', 5)))
        late_threshold_entry.pack(side='left')

        # Missing Punches
        t_row3 = ctk.CTkFrame(threshold_frame, fg_color="transparent")
        t_row3.pack(fill='x', pady=3)
        ctk.CTkLabel(t_row3, text="Missing Punches (count):", width=180).pack(side='left')
        punches_entry = ctk.CTkEntry(t_row3, width=80)
        punches_entry.insert(0, str(thresholds.get('missing_punches', 5)))
        punches_entry.pack(side='left')

        # Low Attendance Rate
        t_row4 = ctk.CTkFrame(threshold_frame, fg_color="transparent")
        t_row4.pack(fill='x', pady=3)
        ctk.CTkLabel(t_row4, text="Low Attendance Rate (%):", width=180).pack(side='left')
        rate_entry = ctk.CTkEntry(t_row4, width=80)
        rate_entry.insert(0, str(thresholds.get('low_attendance_rate', 75.0)))
        rate_entry.pack(side='left')

        # Save button
        def save_settings():
            try:
                self.config.set('work_start_time', self.start_entry.get())
                self.config.set('work_end_time', self.end_entry.get())
                self.config.set('late_threshold', self.late_entry.get())

                selected_days = [i for i, var in day_vars.items() if var.get()]
                self.config.set('off_days', selected_days)

                self.config.set('alert_thresholds', {
                    'high_absences': int(absences_entry.get()),
                    'frequent_late': int(late_threshold_entry.get()),
                    'missing_punches': int(punches_entry.get()),
                    'low_attendance_rate': float(rate_entry.get())
                })

                self.logger.info("Settings saved successfully")
                messagebox.showinfo("Success", "Settings saved successfully!")
                dialog.destroy()
            except Exception as e:
                self.logger.error(f"Error saving settings: {e}", exc_info=True)
                messagebox.showerror("Error", f"Failed to save settings:\n{str(e)}")

        ctk.CTkButton(main_frame, text="Save Settings", command=save_settings,
                     fg_color=self.COLORS['success'], hover_color='#0ea572',
                     font=ctk.CTkFont(size=13, weight="bold"),
                     corner_radius=8, height=40).pack(pady=25)

    def select_master_data(self):
        """Handle master data file selection"""
        file_path = filedialog.askopenfilename(
            title="Select Master Data File",
            initialdir=self.last_directory,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if file_path:
            self.master_file = file_path
            self.last_directory = os.path.dirname(file_path)  # Remember directory
            filename = os.path.basename(file_path)

            # Load master data first to get count
            self.load_master_data()

            # Populate filter checkboxes
            self.populate_filters()

            # Show employee count in status
            emp_count = len(self.employee_mapping)
            self.master_status.configure(
                text=f"✓ {filename} ({emp_count})",
                text_color=self.COLORS['success']
            )
            self.log_message(f"Master data loaded: {filename} - {emp_count} employees", 'success')
            self.logger.info(f"Master data file selected: {file_path} ({emp_count} employees)")

            self.update_generate_button()

    def select_attendance_files(self):
        """Handle attendance files selection"""
        file_paths = filedialog.askopenfilenames(
            title="Select Attendance Files (Multiple Selection Allowed)",
            initialdir=self.last_directory,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if file_paths:
            self.attendance_files = list(file_paths)
            self.last_directory = os.path.dirname(file_paths[0])  # Remember directory
            count = len(self.attendance_files)

            # Try to extract date range from files
            date_range_str = self._get_attendance_date_range()

            if date_range_str:
                self.attendance_status.configure(
                    text=f"✓ {count} file(s) · {date_range_str}",
                    text_color=self.COLORS['success']
                )
                self.log_message(f"{count} attendance file(s) loaded - {date_range_str}", 'success')
            else:
                self.attendance_status.configure(
                    text=f"✓ {count} file(s) loaded",
                    text_color=self.COLORS['success']
                )
                self.log_message(f"{count} attendance file(s) loaded", 'success')

            self.logger.info(f"Attendance files selected: {count} files")
            self.update_generate_button()

    def _get_attendance_date_range(self):
        """Extract date range from attendance files for display"""
        try:
            all_dates = []
            for file_path in self.attendance_files[:3]:  # Check first 3 files max for speed
                df = read_excel_file(file_path)
                if 'Date' in df.columns:
                    dates = pd.to_datetime(df['Date'], errors='coerce').dropna()
                    all_dates.extend(dates.tolist())

            if all_dates:
                min_date = min(all_dates)
                max_date = max(all_dates)
                return f"{min_date.strftime('%d-%b')} to {max_date.strftime('%d-%b %Y')}"
        except Exception as e:
            self.logger.debug(f"Could not extract date range: {e}")
        return None

    def select_leave_sheet(self):
        """Handle leave sheet file selection"""
        file_path = filedialog.askopenfilename(
            title="Select Leave Sheet File (Optional)",
            initialdir=self.last_directory,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if file_path:
            self.leave_file = file_path
            self.last_directory = os.path.dirname(file_path)  # Remember directory
            filename = os.path.basename(file_path)

            # Load leave data first to get count
            self.load_leave_data()

            # Show leave record count
            leave_count = len(self.leave_records) if hasattr(self, 'leave_records') else 0
            self.leave_status.configure(
                text=f"✓ {filename} ({leave_count})",
                text_color=self.COLORS['success']
            )
            self.log_message(f"Leave sheet loaded: {filename} - {leave_count} records", 'success')
            self.logger.info(f"Leave sheet file selected: {file_path} ({leave_count} records)")

    def update_generate_button(self):
        """Enable/disable generate button based on required files"""
        if self.master_file and len(self.attendance_files) > 0:
            self.generate_btn.configure(
                state='normal',
                fg_color=self.COLORS['primary']
            )
        else:
            self.generate_btn.configure(
                state='disabled',
                fg_color=self.COLORS['primary_disabled']
            )

    def find_column(self, df, search_terms, exact_matches=None):
        """
        Find column in dataframe using multiple search strategies.

        Args:
            df: DataFrame to search
            search_terms: List of terms to search for (e.g., ['ac', 'no'])
            exact_matches: Optional list of exact column names to try first

        Returns:
            Column name or None
        """
        # Strategy 1: Try exact matches first
        if exact_matches:
            for exact in exact_matches:
                if exact in df.columns:
                    return exact

        # Strategy 2: Try case-insensitive exact match
        if exact_matches:
            df_cols_lower = {str(c).lower(): c for c in df.columns}
            for exact in exact_matches:
                if exact.lower() in df_cols_lower:
                    return df_cols_lower[exact.lower()]

        # Strategy 3: Search for all terms in column name
        for col in df.columns:
            col_lower = str(col).lower()
            if all(term.lower() in col_lower for term in search_terms):
                return col

        # Strategy 4: Search for any term (less strict)
        for col in df.columns:
            col_lower = str(col).lower()
            if any(term.lower() in col_lower for term in search_terms):
                return col

        return None

    def normalize_id(self, value):
        """Normalize ID values by converting floats to integers before string conversion.

        This fixes the mismatch between float PS IDs (185680.0) and int AC-No (185680).
        """
        if pd.isna(value):
            return ""
        # If it's a float that represents a whole number, convert to int first
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        # Try to convert string floats like "185680.0" to int
        try:
            float_val = float(value)
            if float_val.is_integer():
                return str(int(float_val)).strip()
        except (ValueError, TypeError):
            pass
        return str(value).strip()

    def validate_master_data(self, df):
        """Validate master data has required columns."""
        # Check for required columns (case-insensitive)
        df_cols_lower = [str(c).lower() for c in df.columns]

        # Check for ID column (AC-No. OR PS ID)
        has_id_col = any('ac' in col and 'no' in col for col in df_cols_lower) or \
                     any('ps' in col and 'id' in col for col in df_cols_lower) or \
                     any('psid' in col for col in df_cols_lower)

        # Check for CRM column
        has_crm = any('crm' in col for col in df_cols_lower)

        # Check for Name column
        has_name = any('name' in col for col in df_cols_lower)

        missing = []
        if not has_id_col:
            missing.append('AC-No. or PS ID')
        if not has_crm:
            missing.append('CRM')
        if not has_name:
            missing.append('Name')

        if missing:
            raise ValueError(f"Master data missing required columns: {', '.join(missing)}")

        # Check for empty file
        if len(df) == 0:
            raise ValueError("Master data file is empty")

        return True

    def validate_attendance_file(self, df, filename):
        """Validate attendance file has required columns."""
        required_patterns = [
            ('ac', 'no'),  # AC-No.
            ('name',),     # Name
            ('clock', 'in'),  # Clock In columns
        ]

        df_cols_lower = [str(c).lower() for c in df.columns]

        for pattern in required_patterns:
            if not any(all(p in col for p in pattern) for col in df_cols_lower):
                raise ValueError(f"Attendance file '{filename}' missing columns matching pattern: {pattern}")

        if len(df) == 0:
            raise ValueError(f"Attendance file '{filename}' is empty")

        return True

    def validate_leave_sheet(self, df):
        """Validate leave sheet has required structure."""
        # Check if it's matrix format (has date columns)
        date_cols = [c for c in df.columns if self._is_date_column(c)]

        if len(date_cols) > 0:
            # Matrix format - need CRM column
            df_cols_lower = [str(c).lower() for c in df.columns]
            if not any('crm' in col for col in df_cols_lower):
                raise ValueError("Leave sheet (matrix format) missing CRM column")
        else:
            # Vertical format - need CRM and Date columns
            df_cols_lower = [str(c).lower() for c in df.columns]
            has_crm = any('crm' in col for col in df_cols_lower)
            has_date = any('date' in col for col in df_cols_lower)

            if not has_crm or not has_date:
                raise ValueError("Leave sheet (vertical format) missing CRM or Date column")

        if len(df) == 0:
            raise ValueError("Leave sheet is empty")

        return True

    def _is_date_column(self, col_name):
        """Check if column name looks like a date."""
        # Check if it's a datetime object directly
        if isinstance(col_name, datetime):
            return True
        if hasattr(col_name, 'date'):  # pandas Timestamp
            return True

        col_str = str(col_name)
        # Match formats like "1-Jan", "01-Jan", "1/1", "Jan-1", "2026-01-01", etc.
        date_patterns = [
            r'\d{1,2}[-/]\w{3}',  # 1-Jan
            r'\w{3}[-/]\d{1,2}',  # Jan-1
            r'\d{1,2}[-/]\d{1,2}',  # 1/1
            r'\d{4}[-/]\d{2}[-/]\d{2}',  # 2026-01-01
        ]
        return any(re.search(pattern, col_str) for pattern in date_patterns)

    def load_master_data(self):
        """Load and process master data file"""
        try:
            self.log_message("Loading master data...")
            df = read_excel_file(self.master_file)

            # Validate file
            self.validate_master_data(df)

            # Find columns using robust method
            # Support both AC-No. and PS ID column names
            ac_col = self.find_column(df, ['ac', 'no'], ['AC-No.', 'Ac-No.', 'AC No', 'PS ID', 'PS Id', 'PSID'])
            if not ac_col:
                ac_col = self.find_column(df, ['ps', 'id'], ['PS ID', 'PS Id', 'PSID'])
            crm_col = self.find_column(df, ['crm'], ['CRM'])
            name_col = self.find_column(df, ['name'], ['Name'])
            dept_col = self.find_column(df, ['department', 'dept'], ['Department', 'Dept'])
            pos_col = self.find_column(df, ['position', 'title'], ['Position', 'Title'])
            # National ID column - support common variations including typo "Idnetity"
            national_id_col = self.find_column(df, ['identity', 'idnetity', 'national', 'nid'],
                                               ['Idnetity Number', 'Identity Number', 'National ID', 'NID'])
            # Vendor column (Column B)
            vendor_col = self.find_column(df, ['vendor'], ['Vendor'])
            # PS ID column (Column C) - separate from AC-No lookup
            ps_id_col = self.find_column(df, ['ps', 'id'], ['PS ID', 'PS Id', 'PSID', 'PS-ID'])
            # Join Date column (Column M) - support many variations including newlines
            join_date_col = None
            for col in df.columns:
                col_clean = str(col).replace('\n', ' ').lower().strip()
                if 'join date' in col_clean or 'joining date' in col_clean:
                    join_date_col = col
                    break
            if not join_date_col:
                join_date_col = self.find_column(df, ['join', 'joining', 'hired'],
                                                 ['Join Date', 'Joining Date', 'Date of Joining', 'JoinDate',
                                                  'Hire Date', 'HireDate', 'Date Joined', 'DOJ'])

            # Exit Date column - for resigned employees
            exit_date_col = None
            for col in df.columns:
                col_clean = str(col).replace('\n', ' ').lower().strip()
                if any(term in col_clean for term in ['exit date', 'resign', 'end date', 'termination', 'leaving']):
                    exit_date_col = col
                    break
            if not exit_date_col:
                exit_date_col = self.find_column(df, ['exit', 'resign', 'end', 'termination', 'leaving'],
                                                 ['Exit Date', 'Resignation Date', 'End Date', 'Termination Date',
                                                  'Leaving Date', 'Last Day', 'Last Working Day'])

            if not all([ac_col, crm_col, name_col]):
                self.logger.warning("Some required columns may be missing in master data")
                self.log_message("Warning: Some required columns may be missing", 'warning')

            # Build employee mapping
            self.employee_mapping = {}
            for idx, row in df.iterrows():
                ac_no = self.normalize_id(row[ac_col]) if ac_col else ""
                if ac_no and ac_no != 'nan':
                    # Handle join date formatting - support multiple formats
                    join_date_val = ""
                    if join_date_col and pd.notna(row[join_date_col]):
                        try:
                            raw_val = row[join_date_col]
                            # Check if it's already a datetime
                            if isinstance(raw_val, (datetime, pd.Timestamp)):
                                join_date_val = raw_val.strftime('%Y-%m-%d')
                            # Check if it's a number (Excel serial date) - use numpy check
                            elif hasattr(raw_val, '__int__') and not isinstance(raw_val, str):
                                # Excel serial date: days since 1899-12-30
                                excel_epoch = datetime(1899, 12, 30)
                                jd = excel_epoch + timedelta(days=int(raw_val))
                                join_date_val = jd.strftime('%Y-%m-%d')
                            else:
                                # Try parsing as string
                                jd = pd.to_datetime(raw_val, dayfirst=True)
                                join_date_val = jd.strftime('%Y-%m-%d')
                        except Exception as e:
                            self.logger.debug(f"Error parsing join date: {e}")
                            join_date_val = str(row[join_date_col]).strip()

                    # Handle exit date formatting - same logic as join_date
                    exit_date_val = ""
                    if exit_date_col and pd.notna(row[exit_date_col]):
                        try:
                            raw_val = row[exit_date_col]
                            # Check if it's already a datetime
                            if isinstance(raw_val, (datetime, pd.Timestamp)):
                                exit_date_val = raw_val.strftime('%Y-%m-%d')
                            # Check if it's a number (Excel serial date)
                            elif hasattr(raw_val, '__int__') and not isinstance(raw_val, str):
                                # Excel serial date: days since 1899-12-30
                                excel_epoch = datetime(1899, 12, 30)
                                ed = excel_epoch + timedelta(days=int(raw_val))
                                exit_date_val = ed.strftime('%Y-%m-%d')
                            else:
                                # Try parsing as string
                                ed = pd.to_datetime(raw_val, dayfirst=True)
                                exit_date_val = ed.strftime('%Y-%m-%d')
                        except Exception as e:
                            self.logger.debug(f"Error parsing exit date: {e}")
                            exit_date_val = str(row[exit_date_col]).strip()

                    self.employee_mapping[ac_no] = {
                        'crm': str(row[crm_col]).strip() if crm_col and pd.notna(row[crm_col]) else "",
                        'name': str(row[name_col]).strip() if name_col and pd.notna(row[name_col]) else "",
                        'department': str(row[dept_col]).strip() if dept_col and pd.notna(row[dept_col]) else "",
                        'position': str(row[pos_col]).strip() if pos_col and pd.notna(row[pos_col]) else "",
                        'national_id': str(row[national_id_col]).strip() if national_id_col and pd.notna(row[national_id_col]) else "",
                        'vendor': str(row[vendor_col]).strip() if vendor_col and pd.notna(row[vendor_col]) else "",
                        'ps_id': self.normalize_id(row[ps_id_col]) if ps_id_col and pd.notna(row[ps_id_col]) else "",
                        'join_date': join_date_val,
                        'exit_date': exit_date_val
                    }

            self.log_message(f"Loaded {len(self.employee_mapping)} employee records", 'success')
            self.logger.info(f"Loaded {len(self.employee_mapping)} employee records from master data")

        except Exception as e:
            self.logger.error(f"Error loading master data: {e}", exc_info=True)
            self.log_message(f"Error loading master data: {str(e)}", 'error')
            messagebox.showerror("Error", f"Failed to load master data:\n{str(e)}")

    def load_leave_data(self):
        """Load and process leave data file"""
        try:
            self.log_message("Loading leave data...")
            self.leave_records = []

            # Check for multi-sheet workbook (new format with Master + Jan-Dec sheets)
            xlsx = pd.ExcelFile(self.leave_file)
            sheet_names = xlsx.sheet_names
            month_sheets = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

            # Check if this is the new multi-sheet format
            has_month_sheets = any(month in sheet_names for month in month_sheets)

            if has_month_sheets:
                # New multi-sheet format (Master + monthly sheets)
                self.log_message("Detected multi-sheet format (monthly sheets)")
                self.logger.info("Leave sheet format: multi-sheet (Jan-Dec)")

                # Determine which months to load based on user's selected month
                # Load current month and previous month (for payroll cycles spanning months)
                relevant_months = []
                if hasattr(self, '_user_month_year') and self._user_month_year:
                    year, month = self._user_month_year
                    # Get current month name and previous month name
                    current_month_name = month_sheets[month - 1]  # month is 1-12, list is 0-11
                    prev_month = month - 1 if month > 1 else 12
                    prev_month_name = month_sheets[prev_month - 1]
                    relevant_months = [prev_month_name, current_month_name]
                    self.log_message(f"Loading relevant months: {relevant_months}")
                else:
                    # Load all months if no month specified
                    relevant_months = month_sheets

                for sheet_name in sheet_names:
                    if sheet_name in relevant_months:
                        self.log_message(f"Processing sheet: {sheet_name}")
                        df = pd.read_excel(xlsx, sheet_name=sheet_name)
                        self.parse_leave_matrix(df)

                self.log_message(f"Loaded {len(self.leave_records)} leave records from monthly sheets", 'success')
            else:
                # Single sheet format (old format)
                df = read_excel_file(self.leave_file)

                # Validate file
                self.validate_leave_sheet(df)

                # Detect format (matrix vs vertical)
                # Check for datetime columns or date-like string columns
                has_date_columns = any(
                    isinstance(col, datetime) or
                    hasattr(col, 'date') or  # pandas Timestamp
                    ('-' in str(col) and any(month in str(col).lower() for month in ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']))
                    for col in df.columns
                )

                if has_date_columns:
                    # Matrix format
                    self.log_message("Detected matrix format (dates as columns)")
                    self.logger.info("Leave sheet format: matrix")
                    self.parse_leave_matrix(df)
                else:
                    # Vertical format
                    self.log_message("Detected vertical format")
                    self.logger.info("Leave sheet format: vertical")
                    self.parse_leave_vertical(df)

                self.log_message(f"Loaded {len(self.leave_records)} leave records", 'success')

            self.logger.info(f"Loaded {len(self.leave_records)} leave records")

        except Exception as e:
            self.logger.error(f"Error loading leave data: {e}", exc_info=True)
            self.log_message(f"Error loading leave data: {str(e)}", 'error')
            messagebox.showerror("Error", f"Failed to load leave data:\n{str(e)}")

    def parse_leave_matrix(self, df):
        """Parse leave data in matrix format - optimized for performance"""
        current_year = datetime.now().year

        # Find CRM column using robust method
        crm_col = self.find_column(df, ['crm'], ['CRM'])

        if not crm_col:
            self.logger.warning("CRM column not found in leave sheet")
            self.log_message("CRM column not found in leave sheet", 'warning')
            return

        month_map = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }

        # Pre-identify date columns for efficiency
        date_columns = []
        for col in df.columns:
            leave_date = None

            # Handle datetime column names directly
            if isinstance(col, datetime):
                leave_date = col
            elif hasattr(col, 'date'):  # pandas Timestamp
                leave_date = datetime(col.year, col.month, col.day)
            else:
                # Try parsing string date formats
                col_str = str(col)
                if '-' in col_str:
                    parts = col_str.split('-')
                    if len(parts) == 2:
                        try:
                            day = int(parts[0])
                            month_str = parts[1].lower()[:3]
                            if month_str in month_map:
                                month = month_map[month_str]
                                leave_date = datetime(current_year, month, day)
                        except:
                            pass

            if leave_date:
                date_columns.append((col, leave_date))

        if not date_columns:
            self.logger.warning("No date columns found in leave sheet")
            return

        # Process rows efficiently
        crm_values = df[crm_col].astype(str).str.strip()

        for idx, crm in enumerate(crm_values):
            if not crm or crm == 'nan' or crm.lower() == 'crm':
                continue

            row = df.iloc[idx]

            # Only check identified date columns
            for col, leave_date in date_columns:
                try:
                    leave_type = row[col]
                    if pd.notna(leave_type):
                        leave_str = str(leave_type).strip()
                        if leave_str:
                            self.leave_records.append({
                                'crm': crm,
                                'date': leave_date,
                                'leave_type': leave_str
                            })
                except:
                    pass

    def parse_leave_vertical(self, df):
        """Parse leave data in vertical format"""
        # Find columns using robust method
        crm_col = self.find_column(df, ['crm', 'employee'], ['CRM', 'Employee CRM'])
        date_col = self.find_column(df, ['date'], ['Date'])
        type_col = self.find_column(df, ['type', 'leave'], ['Leave Type', 'Type'])

        if not all([crm_col, date_col]):
            self.logger.warning("Required columns not found in leave sheet")
            self.log_message("Required columns not found in leave sheet", 'warning')
            return

        for idx, row in df.iterrows():
            crm = str(row[crm_col]).strip()
            if not crm or crm == 'nan':
                continue

            try:
                leave_date = pd.to_datetime(row[date_col])
                leave_type = str(row[type_col]).strip() if type_col and pd.notna(row[type_col]) else "Leave"

                self.leave_records.append({
                    'crm': crm,
                    'date': leave_date,
                    'leave_type': leave_type
                })
            except Exception as e:
                self.logger.debug(f"Error parsing leave record at row {idx}: {e}")

    def prompt_for_month_year(self):
        """Prompt user to enter month/year for attendance files."""
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Select Month and Year")
        dialog.geometry("380x280")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.focus_force()  # Bring to front

        ctk.CTkLabel(dialog, text="Cannot detect date from filename.",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(pady=(25, 8))
        ctk.CTkLabel(dialog, text="Please enter Month and Year for attendance files:").pack(pady=8)

        frame = ctk.CTkFrame(dialog, fg_color="transparent")
        frame.pack(pady=15)

        # Month row
        month_row = ctk.CTkFrame(frame, fg_color="transparent")
        month_row.pack(fill='x', pady=8)
        ctk.CTkLabel(month_row, text="Month (1-12):", width=100).pack(side='left')
        month_entry = ctk.CTkEntry(month_row, width=100)
        month_entry.insert(0, str(datetime.now().month))
        month_entry.pack(side='left', padx=5)

        # Year row
        year_row = ctk.CTkFrame(frame, fg_color="transparent")
        year_row.pack(fill='x', pady=8)
        ctk.CTkLabel(year_row, text="Year:", width=100).pack(side='left')
        year_entry = ctk.CTkEntry(year_row, width=100)
        year_entry.insert(0, str(datetime.now().year))
        year_entry.pack(side='left', padx=5)

        result = [None]

        def on_ok():
            try:
                month = int(month_entry.get())
                year = int(year_entry.get())
                if 1 <= month <= 12 and 2020 <= year <= 2030:
                    result[0] = (year, month)
                    self.logger.info(f"User specified month/year: {year}-{month}")
                    dialog.destroy()
                else:
                    messagebox.showerror("Invalid Input", "Month must be 1-12, Year must be 2020-2030")
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter valid numbers")

        # OK Button - more prominent
        ctk.CTkButton(dialog, text="OK", command=on_ok,
                     fg_color=self.COLORS['primary'], hover_color=self.COLORS['primary_hover'],
                     corner_radius=8, width=120, height=40,
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=20)

        dialog.wait_window()
        return result[0]

    def extract_date_from_filename(self, filename):
        """Extract date from filename with multiple strategies."""
        try:
            # Strategy 1: Try to extract full date (YYYY-MM-DD format)
            date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', filename)
            if date_match:
                year, month, day = date_match.groups()
                date = datetime(int(year), int(month), int(day))
                self.logger.debug(f"Extracted date from {filename}: {date} (Strategy 1)")
                return date

            # Strategy 2: Try DD-MM-YYYY or DD_MM_YYYY
            date_match = re.search(r'(\d{1,2})[-_](\d{1,2})[-_](\d{4})', filename)
            if date_match:
                day, month, year = date_match.groups()
                date = datetime(int(year), int(month), int(day))
                self.logger.debug(f"Extracted date from {filename}: {date} (Strategy 2)")
                return date

            # Strategy 3: Extract day only, use current month/year
            day_match = re.search(r'(\d{1,2})_report', filename)
            if day_match:
                day = int(day_match.group(1))
                now = datetime.now()
                date = datetime(now.year, now.month, day)
                self.logger.debug(f"Extracted date from {filename}: {date} (Strategy 3 - current month)")
                return date

            # Strategy 4: Prompt user for month/year on first file
            if not hasattr(self, '_user_month_year'):
                self._user_month_year = self.prompt_for_month_year()

            if self._user_month_year:
                day_str = filename.split('_')[0]
                day = int(day_str)
                year, month = self._user_month_year
                date = datetime(year, month, day)
                self.logger.debug(f"Extracted date from {filename}: {date} (Strategy 4 - user input)")
                return date

            # Fallback to current date
            self.logger.warning(f"Could not extract date from {filename}, using current date")
            return datetime.now()

        except Exception as e:
            self.logger.error(f"Error parsing date from {filename}: {e}", exc_info=True)
            return datetime.now()

    def generate_report(self):
        """Main function to generate the attendance report"""
        try:
            # Validate filter selection - ensure at least one employee is selected
            filtered_mapping = self.get_filtered_employee_mapping()
            if not filtered_mapping:
                messagebox.showwarning(
                    "No Employees Selected",
                    "No employees match the current filter selection.\n\nPlease adjust your Department/CRM filters or click 'Reset' to select all employees."
                )
                return

            self.progress_label.configure(text="Processing... Please wait...", text_color='#ff6400')
            self.progress_bar.set(0.1)  # Show initial progress
            self.root.update()

            # Log filter status
            filter_count = len(filtered_mapping)
            total_count = len(self.employee_mapping)
            if filter_count < total_count:
                self.log_message(f"Generating report for {filter_count} of {total_count} employees (filtered)")

            # Process all files
            self.log_message("=" * 50)
            self.log_message("Starting report generation...")
            self.logger.info("=== Report generation started ===")
            self.process_attendance_files()

            if self.leave_records:
                self.fill_leave_records()

            # Generate Excel reports
            self.log_message("Generating Excel reports...")
            output_file = self.create_excel_report()

            self.progress_bar.set(1.0)  # Complete
            self.progress_label.configure(
                text="✅ Processing complete! Report saved.",
                text_color=self.COLORS['success']
            )

            self.log_message("=" * 50)
            self.log_message(f"Report saved: {output_file}", 'success')
            self.logger.info(f"Report generated successfully: {output_file}")

            # Ask to open file
            response = messagebox.askyesno(
                "Success",
                f"Report generated successfully!\n\nSaved as:\n{output_file}\n\nWould you like to open it now?"
            )

            if response:
                try:
                    os.startfile(output_file)
                    self.logger.info(f"Opened report file: {output_file}")
                except Exception as e:
                    self.logger.error(f"Error opening file: {e}", exc_info=True)
                    messagebox.showwarning("Warning", f"File saved but could not be opened automatically:\n{str(e)}")

        except Exception as e:
            self.progress_bar.set(0)  # Reset
            self.progress_label.configure(text="❌ Error occurred", text_color=self.COLORS['error'])
            self.logger.error(f"Error generating report: {e}", exc_info=True)
            self.log_message(f"Error: {str(e)}", 'error')
            messagebox.showerror("Error", f"Failed to generate report:\n{str(e)}")

    def process_attendance_files(self):
        """Process all attendance files"""
        self.processed_data = []
        skipped_files = 0

        total_files = len(self.attendance_files)
        for i, file_path in enumerate(self.attendance_files, 1):
            filename = os.path.basename(file_path)
            self.log_message(f"Processing file {i}/{total_files}: {filename}")
            self.logger.info(f"Processing attendance file {i}/{total_files}: {file_path}")

            try:
                df = read_excel_file(file_path)
                self.validate_attendance_file(df, filename)
                file_date = self.extract_date_from_filename(filename)

                records = self.process_single_attendance_file(df, file_date)
                self.processed_data.extend(records)
                self.logger.info(f"Successfully processed {filename}: {len(records)} records")

            except Exception as e:
                self.logger.error(f"Error processing {filename}: {e}", exc_info=True)
                self.log_message(f"⚠️ Skipped {filename} due to error: {str(e)}", 'warning')
                skipped_files += 1

        if skipped_files > 0:
            self.log_message(f"Skipped {skipped_files} file(s) due to errors", 'warning')
            self.logger.warning(f"Skipped {skipped_files} files during processing")

        self.log_message(f"Total records processed: {len(self.processed_data)}", 'success')
        self.logger.info(f"Total records processed: {len(self.processed_data)}")

    def process_single_attendance_file(self, df, file_date):
        """Process a single attendance file"""
        records = []

        # Get filtered employee mapping based on user filter selections
        filtered_mapping = self.get_filtered_employee_mapping()

        # Find columns using robust method
        ac_col = self.find_column(df, ['ac', 'no'], ['AC-No.', 'Ac-No.', 'AC No'])
        emp_col = self.find_column(df, ['emp', 'no'], ['Emp No.', 'Employee No'])
        name_col = self.find_column(df, ['name'], ['Name'])
        date_col = self.find_column(df, ['date'], ['Date'])

        for idx, row in df.iterrows():
            ac_no = self.normalize_id(row[ac_col]) if ac_col and pd.notna(row[ac_col]) else ""
            if not ac_no or ac_no == 'nan':
                continue

            # Get date from row if available, otherwise use file_date
            record_date = file_date
            if date_col and pd.notna(row[date_col]):
                try:
                    record_date = pd.to_datetime(row[date_col])
                    if hasattr(record_date, 'to_pydatetime'):
                        record_date = record_date.to_pydatetime()
                except Exception as e:
                    self.logger.debug(f"Error parsing date from row: {e}")
                    record_date = file_date

            # Get clock in/out times
            clock_in = self.get_first_clock_in(row)
            clock_out = self.get_last_clock_out(row)

            # Get employee info - try multiple matching strategies (use filtered mapping)
            emp_info = None
            crm = "Unknown"
            name = ""
            dept = ""
            position = ""

            # Strategy 1: Match by AC-No./PS ID (only include if in filtered mapping)
            if ac_no in filtered_mapping:
                emp_info = filtered_mapping[ac_no]

            # Strategy 2: Match by name if AC-No. didn't match (only in filtered mapping)
            if not emp_info and name_col and pd.notna(row[name_col]):
                row_name = str(row[name_col]).strip().lower()
                for key, info in filtered_mapping.items():
                    if info['name'].lower() == row_name:
                        emp_info = info
                        break

            if emp_info:
                crm = emp_info['crm']
                name = emp_info['name']
                dept = emp_info['department']
                position = emp_info['position']
            else:
                # Skip employees not in filtered mapping (when filters are active)
                if self.department_vars or self.crm_vars:
                    continue  # Skip unmatched employees when filters are applied
                name = str(row[name_col]).strip() if name_col and pd.notna(row[name_col]) else (
                    str(row[emp_col]).strip() if emp_col and pd.notna(row[emp_col]) else "Unknown"
                )
                # Use name as CRM fallback for unmatched employees
                crm = name if name != "Unknown" else f"ID-{ac_no}"
                self.logger.debug(f"Employee AC-No {ac_no} not found in master data, using name: {name}")

            # Determine status with CRM to check for leaves
            status, in_status, out_status = self.determine_status(
                clock_in, clock_out, record_date, crm
            )

            records.append({
                'ac_no': ac_no,
                'crm': crm,
                'name': name,
                'department': dept,
                'position': position,
                'date': record_date,
                'day': record_date.strftime('%A'),
                'clock_in': clock_in,
                'clock_out': clock_out,
                'in_status': in_status,
                'out_status': out_status,
                'status': status
            })

        return records

    def get_first_clock_in(self, row):
        """Get first clock in time from row"""
        for i in range(1, 6):
            col_name = f'Clock In {i}'
            if col_name in row.index and pd.notna(row[col_name]):
                return row[col_name]
        return None

    def get_last_clock_out(self, row):
        """Get last clock out time from row"""
        for i in range(5, 0, -1):
            col_name = f'Clock Out {i}'
            if col_name in row.index and pd.notna(row[col_name]):
                return row[col_name]
        return None

    def determine_status(self, clock_in, clock_out, date, crm):
        """Determine attendance status"""
        # Check day of week
        day_of_week = date.weekday()  # Monday = 0, Sunday = 6

        # Check if employee has resigned (before all other checks)
        for ac_no, info in self.employee_mapping.items():
            if info.get('crm') == crm:
                exit_date_str = info.get('exit_date', '')
                if exit_date_str:
                    try:
                        exit_date = datetime.strptime(exit_date_str, '%Y-%m-%d').date()
                        if date.date() > exit_date:
                            return "Resigned", "Resigned", "Resigned"
                    except Exception:
                        pass
                break

        # Check if it's an OFF day FIRST (before leaves - ignore leaves on OFF days)
        off_days = self.config.get('off_days', [4])  # Default: Friday
        if day_of_week in off_days:
            if not clock_in and not clock_out:
                return "Weekend", "Weekend", "Weekend"
            else:
                return "Worked on Day Off", "Worked", "Worked"

        # Check for leave (only on working days, not OFF days)
        for leave in self.leave_records:
            if (leave['crm'] == crm and
                leave['date'].date() == date.date()):
                return leave['leave_type'], "On Leave", "On Leave"

        # Check for absent
        if not clock_in and not clock_out:
            return "Absent", "No Clock In", "No Clock Out"

        # Check for missing punch - differentiate between In and Out
        if not clock_in:
            return "Missing Punch In", "No Clock In", "..."

        if not clock_out:
            # Check if late
            try:
                if isinstance(clock_in, str):
                    clock_in_time = datetime.strptime(clock_in, "%I:%M:%S %p").time()
                else:
                    clock_in_time = clock_in

                late_threshold_str = self.config.get('late_threshold', '12:00 PM')
                late_threshold = datetime.strptime(late_threshold_str, "%I:%M %p").time()
                in_status = "Late" if clock_in_time > late_threshold else "On Time"
            except Exception as e:
                self.logger.debug(f"Error parsing clock in time: {e}")
                in_status = "..."

            return "Missing Punch Out", in_status, "No Clock Out"

        # Determine if late
        try:
            if isinstance(clock_in, str):
                clock_in_time = datetime.strptime(clock_in, "%I:%M:%S %p").time()
            else:
                clock_in_time = clock_in

            late_threshold_str = self.config.get('late_threshold', '12:00 PM')
            late_threshold = datetime.strptime(late_threshold_str, "%I:%M %p").time()
            is_late = clock_in_time > late_threshold

            if is_late:
                return "Late", "Late", "On Time"
            else:
                return "Normal", "On Time", "On Time"
        except Exception as e:
            self.logger.debug(f"Error determining late status: {e}")
            return "Normal", "On Time", "On Time"

    def fill_leave_records(self):
        """Apply leave records to attendance data - ONLY within attendance date range"""
        self.log_message("Applying leave records to attendance data...")
        self.logger.info("Applying leave records to attendance data")

        # Get OFF days from config (skip adding leave records for these days)
        off_days = self.config.get('off_days', [4])  # Default: Friday

        # Get the attendance date range - ONLY process leaves within this range
        if not self.processed_data:
            self.log_message("No processed data, skipping leave fill")
            return

        all_attendance_dates = [record['date'] for record in self.processed_data]
        min_date = min(all_attendance_dates).date()
        max_date = max(all_attendance_dates).date()
        self.log_message(f"Attendance date range: {min_date} to {max_date}")
        self.logger.info(f"Attendance date range: {min_date} to {max_date}")

        # *** FIX: Use FILTERED employee mapping to respect filter selections ***
        filtered_mapping = self.get_filtered_employee_mapping()

        # Get valid CRMs from FILTERED mapping (not all master data)
        valid_crms = set(info['crm'] for info in filtered_mapping.values())
        self.log_message(f"Valid CRMs from filtered mapping: {len(valid_crms)}")

        # Build a lookup: (crm, date) -> index in processed_data
        record_index = {}
        for idx, record in enumerate(self.processed_data):
            key = (record['crm'], record['date'].date())
            record_index[key] = idx

        # Build employee info lookup from FILTERED mapping
        crm_to_emp_info = {}
        for ac_no, info in filtered_mapping.items():
            crm_to_emp_info[info['crm']] = info

        added_count = 0
        updated_count = 0
        skipped_friday = 0
        skipped_out_of_range = 0
        skipped_no_master = 0

        # Build leave lookup by (crm, date) for efficiency
        leave_lookup = {}
        for leave in self.leave_records:
            crm = leave['crm']
            leave_date = leave['date'].date()

            # Skip if CRM not in master data
            if crm not in valid_crms:
                skipped_no_master += 1
                continue

            # Skip if leave date is OUTSIDE attendance date range
            if leave_date < min_date or leave_date > max_date:
                skipped_out_of_range += 1
                continue

            # Skip leave records on OFF days (Friday) - they should show as Weekend
            if leave['date'].weekday() in off_days:
                skipped_friday += 1
                continue

            key = (crm, leave_date)
            # Store the leave (if multiple leaves on same day, keep the last one)
            leave_lookup[key] = leave

        self.log_message(f"Leaves to apply: {len(leave_lookup)}")

        # Clear previous conflict records before processing
        self.conflict_records = []

        # Now apply leaves
        for key, leave in leave_lookup.items():
            crm, leave_date = key

            if key in record_index:
                # UPDATE existing record with leave status
                idx = record_index[key]
                existing_record = self.processed_data[idx]
                existing_status = existing_record['status']

                # Track conflict BEFORE updating - if employee had actual attendance but leave applied
                if existing_status not in ['Weekend', 'Resigned', 'Absent']:
                    self.conflict_records.append({
                        'crm': crm,
                        'name': existing_record.get('name', ''),
                        'date': leave['date'],
                        'attendance_status': existing_status,
                        'leave_type': leave['leave_type'],
                        'clock_in': existing_record.get('clock_in'),
                        'clock_out': existing_record.get('clock_out')
                    })

                # Only update if current status is not already a leave type or is Absent/Missing Punch
                if existing_status in ['Absent', 'Missing Punch In', 'Missing Punch Out', 'Late', 'Normal']:
                    self.processed_data[idx]['status'] = leave['leave_type']
                    self.processed_data[idx]['in_status'] = 'On Leave'
                    self.processed_data[idx]['out_status'] = 'On Leave'
                    updated_count += 1
            else:
                # ADD new record for this leave
                emp_info = crm_to_emp_info.get(crm)
                if emp_info:
                    self.processed_data.append({
                        'ac_no': '',
                        'crm': crm,
                        'name': emp_info['name'],
                        'department': emp_info['department'],
                        'position': emp_info['position'],
                        'date': leave['date'],
                        'day': leave['date'].strftime('%A'),
                        'clock_in': None,
                        'clock_out': None,
                        'in_status': 'On Leave',
                        'out_status': 'On Leave',
                        'status': leave['leave_type']
                    })
                    added_count += 1

        self.log_message(f"Updated {updated_count} existing records with leave status")
        self.log_message(f"Added {added_count} new leave-only records")
        self.log_message(f"Skipped: {skipped_friday} Fridays, {skipped_out_of_range} out-of-range, {skipped_no_master} not in master")
        if self.conflict_records:
            self.log_message(f"Detected {len(self.conflict_records)} leave vs attendance conflicts", 'warning')
        self.log_message(f"Total records: {len(self.processed_data)}")
        self.logger.info(f"Applied leaves: {updated_count} updated, {added_count} added (skipped {skipped_friday} Fridays, {skipped_out_of_range} out-of-range, {skipped_no_master} not in master). Conflicts: {len(self.conflict_records)}. Total: {len(self.processed_data)}")

    def calculate_penalties(self):
        """Calculate penalties per employee based on attendance policy"""
        self.log_message("Calculating penalties...")
        self.logger.info("Calculating penalties based on attendance policy")

        penalties_config = self.config.get('penalties', {})
        currency = penalties_config.get('currency', 'EGP')

        # Group records by employee (CRM)
        employee_stats = {}

        for record in self.processed_data:
            crm = record['crm']
            status = record['status']

            if crm not in employee_stats:
                # Look up employee info from employee_mapping by CRM
                national_id = ''
                vendor = ''
                ps_id = ''
                join_date = ''
                for ac_no, emp_info in self.employee_mapping.items():
                    if emp_info.get('crm') == crm:
                        national_id = emp_info.get('national_id', '')
                        vendor = emp_info.get('vendor', '')
                        ps_id = emp_info.get('ps_id', '')
                        join_date = emp_info.get('join_date', '')
                        break

                employee_stats[crm] = {
                    'name': record.get('name', ''),
                    'department': record.get('department', ''),
                    'national_id': national_id,
                    'vendor': vendor,
                    'ps_id': ps_id,
                    'join_date': join_date,
                    'late_count': 0,
                    'missing_punch_count': 0,
                    'absence_count': 0,
                    'early_departure_count': 0,
                    'total_days': 0,
                    'working_days': 0
                }

            employee_stats[crm]['total_days'] += 1

            # Count by status type
            if status == 'Late':
                employee_stats[crm]['late_count'] += 1
                employee_stats[crm]['working_days'] += 1
            elif status == 'Missing Punch':
                employee_stats[crm]['missing_punch_count'] += 1
                employee_stats[crm]['working_days'] += 1
            elif status == 'Absent':
                employee_stats[crm]['absence_count'] += 1
            elif status in ['Normal', 'Present']:
                employee_stats[crm]['working_days'] += 1
            # Weekend and Leave don't count as working days for penalty

        # Calculate penalties for each employee
        penalties_summary = {}

        for crm, stats in employee_stats.items():
            late_count = stats['late_count']
            missing_count = stats['missing_punch_count']
            absence_count = stats['absence_count']

            # Calculate late penalties (cumulative: 1st=100, 2nd=200, 3rd=500, 4th+=500)
            late_penalty = 0
            late_warnings = 0
            for i in range(late_count):
                if i == 0:
                    late_penalty += penalties_config.get('late_1st', 100)
                elif i == 1:
                    late_penalty += penalties_config.get('late_2nd', 200)
                elif i == 2:
                    late_penalty += penalties_config.get('late_3rd', 500)
                    late_warnings += 1  # Warning after 3rd
                else:
                    late_penalty += penalties_config.get('late_4th_plus', 500)

            # Calculate missing punch deductions
            missing_threshold = penalties_config.get('missing_punch_threshold', 3)
            missing_warning_threshold = penalties_config.get('missing_punch_warning_threshold', 6)
            missing_deduction_rate = penalties_config.get('missing_punch_deduction', 0.5)

            missing_deduction = 0
            missing_warnings = 0
            if missing_count > missing_threshold:
                missing_deduction = (missing_count - missing_threshold) * missing_deduction_rate
            if missing_count > missing_warning_threshold:
                missing_warnings = 1

            # Calculate absence deductions (2 days per absence)
            absence_deduction_rate = penalties_config.get('absence_deduction', 2)
            absence_deduction = absence_count * absence_deduction_rate
            absence_warnings = 1 if absence_count > 0 else 0

            # Total calculations
            total_penalty_egp = late_penalty
            total_deduction_days = missing_deduction + absence_deduction
            total_warnings = late_warnings + missing_warnings + absence_warnings

            penalties_summary[crm] = {
                'name': stats['name'],
                'department': stats['department'],
                'national_id': stats.get('national_id', ''),
                'vendor': stats.get('vendor', ''),
                'ps_id': stats.get('ps_id', ''),
                'join_date': stats.get('join_date', ''),
                'late_count': late_count,
                'late_penalty': late_penalty,
                'missing_punch_count': missing_count,
                'missing_deduction': missing_deduction,
                'absence_count': absence_count,
                'absence_deduction': absence_deduction,
                'total_penalty_egp': total_penalty_egp,
                'total_deduction_days': total_deduction_days,
                'total_warnings': total_warnings,
                'working_days': stats['working_days'],
                'total_days': stats['total_days']
            }

        self.penalties_data = penalties_summary
        self.log_message(f"Calculated penalties for {len(penalties_summary)} employees", 'success')
        return penalties_summary

    def create_excel_report(self):
        """Create the final Excel report"""
        # Create output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"Attendance_Report_{timestamp}.xlsx"

        # ✅ FIX: Always save to Documents folder (works for both exe and script mode)
        docs_folder = Path.home() / "Documents"
        docs_folder.mkdir(parents=True, exist_ok=True)
        output_file = docs_folder / output_file
        output_file = str(output_file)

        self.logger.info(f"Creating Excel report: {output_file}")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Calculate penalties first
        self.calculate_penalties()

        # Create Summary Report
        self.create_summary_sheet(wb)

        # Create Individual Analytics
        self.create_analytics_sheet(wb)

        # Create Alerts sheet
        self.create_alerts_sheet(wb)

        # Create Penalties sheet
        self.create_penalties_sheet(wb)

        # Create Duplicates sheet (leave vs attendance conflicts)
        self.create_duplicates_sheet(wb)

        # Save workbook
        wb.save(output_file)
        self.logger.info(f"Excel report saved successfully: {output_file}")

        return output_file

    def create_summary_sheet(self, wb):
        """Create the summary report sheet with dropdown justification options"""
        ws = wb.create_sheet("Summary Report", 0)

        # Get unique CRMs
        crms = sorted(set(r['crm'] for r in self.processed_data))

        # Get date range and generate ALL dates (including Fridays/weekends)
        attendance_dates = [r['date'] for r in self.processed_data]
        if attendance_dates:
            min_date = min(attendance_dates)
            max_date = max(attendance_dates)

            # Generate all dates in range
            dates = []
            current = min_date
            while current <= max_date:
                dates.append(current)
                current = current + timedelta(days=1)
        else:
            dates = sorted(set(r['date'] for r in self.processed_data))

        self.logger.info(f"Creating summary sheet: {len(crms)} employees, {len(dates)} dates (full range)")

        # Build matrix
        matrix = {crm: {date: '' for date in dates} for crm in crms}

        for record in self.processed_data:
            matrix[record['crm']][record['date']] = record['status']

        # Title
        ws.merge_cells('A1:' + get_column_letter(len(dates) + 3) + '1')
        title_cell = ws['A1']
        title_cell.value = "📊 Enhanced Attendance Summary Report"
        title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Headers
        row = 3
        ws.cell(row, 1, "CRM")
        ws.cell(row, 2, "Normal Days")
        ws.cell(row, 3, "Abnormal Days")

        for i, date in enumerate(dates, start=4):
            ws.cell(row, i, date.strftime("%d-%b"))

        # Style headers
        for col in range(1, len(dates) + 4):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        # Get OFF days from config (used for Weekend detection)
        off_days = self.config.get('off_days', [4])  # Default: Friday

        # Define justification options for dropdown (including (BD) backdated variants)
        justification_options = [
            "Normal",
            "Late (Approved)",
            "Late",
            "Absent",
            "Missing Punch In",
            "Missing Punch In (Justified)",
            "Missing Punch Out",
            "Missing Punch Out (Justified)",
            "Early Departure (Approved)",
            "Early Departure",
            "Half Day",
            "Early Leave (HD)",
            "Sick Leave",
            "Annual Leave",
            "Casual Leave",
            "Marriage Leave",
            "Paternity Leave",
            "Maternity Leave",
            "Bereavement Leave",
            "Military Call Leave",
            "Unpaid Leave",
            "Weekend",
            "Resigned",
            # Backdated variants (BD) - for leaves transferred from previous months
            "Annual Leave (BD)",
            "Casual Leave (BD)",
            "Sick Leave (BD)",
            "Unpaid Leave (BD)",
            "Half Day (BD)",
            "Early Leave (HD) (BD)",
            "Early Departure (BD)",
            "Marriage Leave (BD)",
            "Paternity Leave (BD)",
            "Maternity Leave (BD)",
            "Bereavement Leave (BD)",
            # Refund variants - reversal of previous penalties
            "Annual Leave (Refund)",
            "Sick Leave (Refund)",
            "Half Day (Refund)"
        ]

        # Create data validation for dropdown
        justification_list = ",".join(justification_options)
        dv = DataValidation(
            type="list",
            formula1=f'"{justification_list}"',
            allow_blank=True,
            showDropDown=False  # False means show the dropdown arrow
        )
        dv.error = "Please select a valid justification from the list"
        dv.errorTitle = "Invalid Entry"
        dv.prompt = "Select attendance status/justification"
        dv.promptTitle = "Attendance Status"
        ws.add_data_validation(dv)

        # Build CRM → join_date lookup for "Not Yet Hired" detection
        crm_join_dates = {}
        for ac_no, emp_info in self.employee_mapping.items():
            crm_val = emp_info.get('crm', '')
            jd = emp_info.get('join_date', '')
            if crm_val and jd:
                try:
                    crm_join_dates[crm_val] = datetime.strptime(jd, '%Y-%m-%d').date() if isinstance(jd, str) else jd.date() if hasattr(jd, 'date') else jd
                except (ValueError, TypeError):
                    pass

        # Data rows
        row = 4
        first_data_row = row
        for crm in crms:
            ws.cell(row, 1, crm)

            # Count normal days (includes Weekend and Worked on Day Off as expected)
            # Friday (OFF day) ALWAYS counts as Weekend (normal) - ignore any leave status
            normal_statuses = ['Normal', 'Present', 'Weekend', 'Worked on Day Off', 'Late (Approved)',
                              'Annual Leave', 'Casual Leave', 'Marriage Leave', 'Paternity Leave',
                              'Maternity Leave', 'Bereavement Leave', 'Military Call Leave',
                              'Early Departure (Approved)',
                              'Annual Leave (Refund)', 'Sick Leave (Refund)', 'Half Day (Refund)',
                              'Not Yet Hired']
            normal_count = 0
            for d in dates:
                # Pre-hire dates count as normal (no penalty)
                join_dt = crm_join_dates.get(crm)
                d_date = d.date() if hasattr(d, 'date') else d
                if join_dt and d_date < join_dt:
                    normal_count += 1
                # Friday always counts as normal (Weekend)
                elif d.weekday() in off_days:
                    normal_count += 1
                else:
                    status = matrix[crm][d]
                    if status in normal_statuses:
                        normal_count += 1
            ws.cell(row, 2, normal_count)
            ws.cell(row, 3, len(dates) - normal_count)

            for i, date in enumerate(dates, start=4):
                status = matrix[crm][date]
                day_of_week = date.weekday()  # Monday=0, Friday=4, Sunday=6

                # Check if date is before employee's join date
                join_dt = crm_join_dates.get(crm)
                date_val = date.date() if hasattr(date, 'date') else date
                if join_dt and date_val < join_dt:
                    status = "Not Yet Hired"
                # Friday (OFF day) ALWAYS shows as Weekend - ignore any leave status
                elif day_of_week in off_days:
                    status = "Weekend"
                elif not status:
                    # Blank cell on a working day = Absent
                    status = "Absent"

                # Map old status values to new justification values
                status_mapping = {
                    'Present': 'Normal',
                }
                status = status_mapping.get(status, status)

                cell = ws.cell(row, i, status)

                # Apply color coding based on justification type
                self._apply_status_color(cell, status)

                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            row += 1

        last_data_row = row - 1

        # Apply data validation to all date columns (from column D onwards)
        if dates and crms:
            first_date_col = get_column_letter(4)
            last_date_col = get_column_letter(len(dates) + 3)
            dv.add(f"{first_date_col}{first_data_row}:{last_date_col}{last_data_row}")

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        for i in range(4, len(dates) + 4):
            ws.column_dimensions[get_column_letter(i)].width = 18

    def _apply_status_color(self, cell, status):
        """Apply color coding to a cell based on status/justification"""
        # Light Green - Refund leaves (Refund) - check first before other rules
        if '(Refund)' in status:
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        # Light Purple - Backdated leaves (BD)
        elif '(BD)' in status:
            cell.fill = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
        # Green - No deduction statuses
        elif status in ['Normal', 'Present', 'Late (Approved)', 'Annual Leave', 'Casual Leave',
                      'Marriage Leave', 'Paternity Leave', 'Maternity Leave', 'Bereavement Leave',
                      'Military Call Leave', 'Early Departure (Approved)']:
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        # Yellow - Late with deduction
        elif status == 'Late':
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        # Red - Absent
        elif status == 'Absent':
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        # Pink - Missing Punch types
        elif 'Missing Punch' in status:
            cell.fill = PatternFill(start_color='FFD9E6', end_color='FFD9E6', fill_type='solid')
        # Light Blue - Leave types with deduction
        elif status in ['Sick Leave', 'Unpaid Leave', 'Unpaid leave']:
            cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        # Orange - Half day / Early departure / Early Leave (HD)
        elif status in ['Early Departure', 'Half Day', 'Early Leave (HD)', 'Early Leave (HD) (BD)']:
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        # Gray - Weekend
        elif status == 'Weekend':
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        # Light Gray - Resigned
        elif status == 'Resigned':
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        # Light Silver - Not Yet Hired
        elif status == 'Not Yet Hired':
            cell.fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
            cell.font = Font(color='808080', italic=True)
        # Light green - Worked on Day Off
        elif status == 'Worked on Day Off':
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        # Default light blue for other leave types
        elif 'Leave' in status:
            cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

    def create_analytics_sheet(self, wb):
        """Create individual analytics sheet"""
        ws = wb.create_sheet("Individual Analytics", 1)

        # Calculate analytics
        analytics = {}
        for crm in set(r['crm'] for r in self.processed_data):
            records = [r for r in self.processed_data if r['crm'] == crm]
            total = len(records)
            normal = sum(1 for r in records if r['status'] == 'Normal')
            late = sum(1 for r in records if r['status'] == 'Late')
            absent = sum(1 for r in records if r['status'] == 'Absent')
            missing = sum(1 for r in records if r['status'] == 'Missing Punch')

            analytics[crm] = {
                'name': records[0]['name'],
                'dept': records[0]['department'],
                'position': records[0]['position'],
                'total': total,
                'normal': normal,
                'late': late,
                'absent': absent,
                'missing': missing,
                'attendance_rate': round((total - absent) / total * 100, 1) if total > 0 else 0,
                'punctuality_rate': round(normal / total * 100, 1) if total > 0 else 0
            }

        self.logger.info(f"Creating analytics sheet: {len(analytics)} employees")

        # Title
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = "📈 Individual Employee Analytics"
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Headers
        headers = ['CRM', 'Name', 'Department', 'Position', 'Total Days', 'Normal Days',
                  'Late Days', 'Absent Days', 'Missing Punch', 'Attendance Rate %', 'Punctuality Rate %']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        row = 4
        for crm, data in sorted(analytics.items()):
            ws.cell(row, 1, crm)
            ws.cell(row, 2, data['name'])
            ws.cell(row, 3, data['dept'])
            ws.cell(row, 4, data['position'])
            ws.cell(row, 5, data['total'])
            ws.cell(row, 6, data['normal'])
            ws.cell(row, 7, data['late'])
            ws.cell(row, 8, data['absent'])
            ws.cell(row, 9, data['missing'])
            ws.cell(row, 10, f"{data['attendance_rate']}%")
            ws.cell(row, 11, f"{data['punctuality_rate']}%")
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        for col in range(5, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def create_alerts_sheet(self, wb):
        """Create alerts sheet with configurable thresholds"""
        ws = wb.create_sheet("Alerts & Warnings", 2)

        # Get thresholds from config
        thresholds = self.config.get('alert_thresholds', {})
        high_absences = thresholds.get('high_absences', 3)
        frequent_late = thresholds.get('frequent_late', 5)
        missing_punches = thresholds.get('missing_punches', 5)
        low_rate = thresholds.get('low_attendance_rate', 75.0)

        # Calculate analytics for alerts
        analytics = {}
        for crm in set(r['crm'] for r in self.processed_data):
            records = [r for r in self.processed_data if r['crm'] == crm]
            total = len(records)
            late = sum(1 for r in records if r['status'] == 'Late')
            absent = sum(1 for r in records if r['status'] == 'Absent')
            missing = sum(1 for r in records if r['status'] == 'Missing Punch')
            attendance_rate = (total - absent) / total * 100 if total > 0 else 0

            analytics[crm] = {
                'name': records[0]['name'],
                'late': late,
                'absent': absent,
                'missing': missing,
                'rate': attendance_rate
            }

        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "⚠️ Attendance Alerts & Warnings"
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Headers
        headers = ['CRM', 'Name', 'Alert Type', 'Details']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Generate alerts
        row = 4
        alert_count = 0

        for crm, data in sorted(analytics.items()):
            if data['absent'] >= high_absences:
                ws.cell(row, 1, crm)
                ws.cell(row, 2, data['name'])
                ws.cell(row, 3, "High Absences")
                ws.cell(row, 4, f"{data['absent']} absences in period")
                for col in range(1, 5):
                    ws.cell(row, col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                row += 1
                alert_count += 1

            if data['late'] >= frequent_late:
                ws.cell(row, 1, crm)
                ws.cell(row, 2, data['name'])
                ws.cell(row, 3, "Frequent Late Arrivals")
                ws.cell(row, 4, f"{data['late']} late arrivals in period")
                for col in range(1, 5):
                    ws.cell(row, col).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                row += 1
                alert_count += 1

            if data['missing'] >= missing_punches:
                ws.cell(row, 1, crm)
                ws.cell(row, 2, data['name'])
                ws.cell(row, 3, "Missing Punches")
                ws.cell(row, 4, f"{data['missing']} missing punches (possible device issue)")
                for col in range(1, 5):
                    ws.cell(row, col).fill = PatternFill(start_color='FFD9E6', end_color='FFD9E6', fill_type='solid')
                row += 1
                alert_count += 1

            if data['rate'] < low_rate:
                ws.cell(row, 1, crm)
                ws.cell(row, 2, data['name'])
                ws.cell(row, 3, "Low Attendance Rate")
                ws.cell(row, 4, f"Attendance rate: {data['rate']:.1f}%")
                for col in range(1, 5):
                    ws.cell(row, col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                row += 1
                alert_count += 1

        self.logger.info(f"Created alerts sheet: {alert_count} alerts generated")

        if alert_count == 0:
            ws.merge_cells('A5:D5')
            cell = ws['A5']
            cell.value = "✅ No attendance issues detected!"
            cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 50

    def create_penalties_sheet(self, wb):
        """Create the penalties calculation sheet based on attendance policy.
        Uses Excel formulas to link to Summary Report for dynamic updates.
        Supports all justification types with their respective deductions."""
        ws = wb.create_sheet("Penalties", 3)

        currency = self.config.get('penalties', {}).get('currency', 'EGP')
        penalties_config = self.config.get('penalties', {})

        # Get summary sheet info for formula references
        summary_ws = wb['Summary Report']
        summary_last_col = summary_ws.max_column
        summary_last_col_letter = get_column_letter(summary_last_col)

        # Total columns: 28 (A-AB) - Added Refund columns and Net Deduction
        total_cols = 28

        # Title
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        title_cell = ws['A1']
        title_cell.value = f"Attendance Penalties Report ({currency})"
        title_cell.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle with policy reference
        ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
        subtitle = ws['A2']
        subtitle.value = "Based on Attendance and Discipline Policy 2026 - Section 7 (Linked to Summary Report)"
        subtitle.font = Font(size=9, italic=True)
        subtitle.alignment = Alignment(horizontal='center')

        # Headers with justification-based columns (28 columns: A-AB)
        headers = [
            'CRM', 'Name', 'National ID', 'Vendor', 'PS ID', 'Department', 'Join Date',
            'Late Count', f'Late Penalty ({currency})',
            'Missing Punches', 'Punch Ded. (-)',
            'Absences', 'Absence Ded. (-)',
            'Early Dep.', 'Early Dep. Ded. (-)',
            'Half Day', 'Half Day Ded. (-)',
            'Sick Leave', 'Sick Ded. (-)',
            'Unpaid Leave', 'Unpaid Ded. (-)',
            'Refund Count', 'Refund (+days)',
            f'Total Penalty ({currency})', 'Gross Ded. (days)', 'Net Ded. (days)',
            'Warnings', 'Backdated Leaves'
        ]

        header_row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill(start_color='495057', end_color='495057', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        ws.row_dimensions[header_row].height = 40

        # Data rows with formulas linking to Summary Report
        data_row = 5

        # Get penalty config values for formulas
        late_1st = penalties_config.get('late_1st', 100)
        late_2nd = penalties_config.get('late_2nd', 200)
        late_3rd = penalties_config.get('late_3rd', 500)
        late_4th_plus = penalties_config.get('late_4th_plus', 500)
        missing_threshold = penalties_config.get('missing_punch_threshold', 3)
        missing_deduction_rate = penalties_config.get('missing_punch_deduction', 0.5)
        absence_deduction_rate = penalties_config.get('absence_deduction', 2)

        # New deduction rates from justification rules
        early_dep_deduction = 0.5  # Half day
        half_day_deduction = 0.5   # Half day
        sick_leave_deduction = 0.25  # 0.25 days
        unpaid_leave_deduction = 1.0  # 1 day

        # Build a mapping of CRM to row number in Summary Report
        crm_to_summary_row = {}
        for r in range(4, summary_ws.max_row + 1):
            crm_val = summary_ws.cell(r, 1).value
            if crm_val:
                crm_to_summary_row[crm_val] = r

        for crm, data in sorted(self.penalties_data.items()):
            # Get the corresponding row in Summary Report for this CRM
            summary_row = crm_to_summary_row.get(crm, None)
            sr = summary_row  # Shorthand
            rng = f"'Summary Report'!$D${sr}:${summary_last_col_letter}${sr}" if sr else None

            # Column A: CRM
            ws.cell(data_row, 1, crm)

            # Columns B-G: Static employee info
            ws.cell(data_row, 2, data['name'])
            ws.cell(data_row, 3, data.get('national_id', ''))
            ws.cell(data_row, 4, data.get('vendor', ''))
            ws.cell(data_row, 5, data.get('ps_id', ''))
            ws.cell(data_row, 6, data.get('department', ''))
            ws.cell(data_row, 7, data.get('join_date', ''))

            # Column H (8): Late Count - Only counts "Late" (not "Late (Approved)")
            if sr:
                ws.cell(data_row, 8, f'=COUNTIF({rng},"Late")')
            else:
                ws.cell(data_row, 8, data['late_count'])

            # Column I (9): Late Penalty - Cumulative formula
            late_col = "H"
            penalty_formula = (
                f"=IF({late_col}{data_row}=0,0,"
                f"IF({late_col}{data_row}=1,{late_1st},"
                f"IF({late_col}{data_row}=2,{late_1st}+{late_2nd},"
                f"IF({late_col}{data_row}=3,{late_1st}+{late_2nd}+{late_3rd},"
                f"{late_1st}+{late_2nd}+{late_3rd}+({late_col}{data_row}-3)*{late_4th_plus}))))"
            )
            ws.cell(data_row, 9, penalty_formula)

            # Column J (10): Missing Punches - All 4 types
            if sr:
                missing_formula = (
                    f'=COUNTIF({rng},"Missing Punch In")'
                    f'+COUNTIF({rng},"Missing Punch In (Justified)")'
                    f'+COUNTIF({rng},"Missing Punch Out")'
                    f'+COUNTIF({rng},"Missing Punch Out (Justified)")'
                )
                ws.cell(data_row, 10, missing_formula)
            else:
                ws.cell(data_row, 10, data['missing_punch_count'])

            # Column K (11): Punch Deduction (negative)
            ws.cell(data_row, 11, f"=-IF(J{data_row}>{missing_threshold},(J{data_row}-{missing_threshold})*{missing_deduction_rate},0)")

            # Column L (12): Absences
            if sr:
                ws.cell(data_row, 12, f'=COUNTIF({rng},"Absent")')
            else:
                ws.cell(data_row, 12, data['absence_count'])

            # Column M (13): Absence Deduction (negative)
            ws.cell(data_row, 13, f"=-L{data_row}*{absence_deduction_rate}")

            # Column N (14): Early Departure Count (includes BD variant)
            if sr:
                ws.cell(data_row, 14, f'=COUNTIF({rng},"Early Departure")+COUNTIF({rng},"Early Departure (BD)")')
            else:
                ws.cell(data_row, 14, 0)

            # Column O (15): Early Departure Deduction (negative)
            ws.cell(data_row, 15, f"=-N{data_row}*{early_dep_deduction}")

            # Column P (16): Half Day Count (includes BD variant and Early Leave HD)
            if sr:
                ws.cell(data_row, 16, f'=COUNTIF({rng},"Half Day")+COUNTIF({rng},"Half Day (BD)")+COUNTIF({rng},"Early Leave (HD)")+COUNTIF({rng},"Early Leave (HD) (BD)")')
            else:
                ws.cell(data_row, 16, 0)

            # Column Q (17): Half Day Deduction (negative)
            ws.cell(data_row, 17, f"=-P{data_row}*{half_day_deduction}")

            # Column R (18): Sick Leave Count (includes BD variant)
            if sr:
                ws.cell(data_row, 18, f'=COUNTIF({rng},"Sick Leave")+COUNTIF({rng},"Sick Leave (BD)")')
            else:
                ws.cell(data_row, 18, 0)

            # Column S (19): Sick Leave Deduction (negative)
            ws.cell(data_row, 19, f"=-R{data_row}*{sick_leave_deduction}")

            # Column T (20): Unpaid Leave Count (includes BD variant and variations)
            if sr:
                ws.cell(data_row, 20, f'=COUNTIF({rng},"Unpaid Leave")+COUNTIF({rng},"Unpaid Leave (BD)")+COUNTIF({rng},"Unpaid leave")')
            else:
                ws.cell(data_row, 20, 0)

            # Column U (21): Unpaid Leave Deduction (negative)
            ws.cell(data_row, 21, f"=-T{data_row}*{unpaid_leave_deduction}")

            # Column V (22): Refund Count (all refund types)
            if sr:
                ws.cell(data_row, 22, f'=COUNTIF({rng},"Annual Leave (Refund)")+COUNTIF({rng},"Sick Leave (Refund)")+COUNTIF({rng},"Half Day (Refund)")')
            else:
                ws.cell(data_row, 22, 0)

            # Column W (23): Refund Days (positive - adds days back)
            if sr:
                ws.cell(data_row, 23, f'=COUNTIF({rng},"Annual Leave (Refund)")*1+COUNTIF({rng},"Sick Leave (Refund)")*0.75+COUNTIF({rng},"Half Day (Refund)")*0.5')
            else:
                ws.cell(data_row, 23, 0)

            # Column X (24): Total Penalty (only late penalty contributes to EGP)
            ws.cell(data_row, 24, f"=I{data_row}")

            # Column Y (25): Gross Deduction (sum of all negative day deductions)
            ws.cell(data_row, 25, f"=K{data_row}+M{data_row}+O{data_row}+Q{data_row}+S{data_row}+U{data_row}")

            # Column Z (26): Net Deduction (gross deductions + refund credits)
            ws.cell(data_row, 26, f"=Y{data_row}+W{data_row}")

            # Column AA (27): Warnings
            warnings_formula = (
                f"=IF(H{data_row}>=3,1,0)+"  # Late warning
                f"IF(J{data_row}>6,1,0)+"     # Missing punch warning
                f"IF(L{data_row}>0,1,0)"      # Absence warning
            )
            ws.cell(data_row, 27, warnings_formula)

            # Column AB (28): Backdated Leaves - counts all leaves with (BD) suffix
            if sr:
                backdated_formula = f'=COUNTIF({rng},"*(BD)")'
                ws.cell(data_row, 28, backdated_formula)
            else:
                ws.cell(data_row, 28, 0)

            # Apply borders and alignment to all cells
            for col in range(1, total_cols + 1):
                cell = ws.cell(data_row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            data_row += 1

        # Totals row with SUM formulas
        totals_row = data_row + 1
        first_data_row = 5
        last_data_row = data_row - 1

        ws.cell(totals_row, 1, "TOTAL").font = Font(bold=True)

        # Sum formulas for count and deduction columns (28 columns)
        sum_columns = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28]
        for col in sum_columns:
            col_letter = get_column_letter(col)
            ws.cell(totals_row, col, f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")

        # Highlight total penalty and deduction
        ws.cell(totals_row, 24).fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        ws.cell(totals_row, 24).font = Font(bold=True, color='FFFFFF', size=12)
        ws.cell(totals_row, 25).fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        ws.cell(totals_row, 25).font = Font(bold=True, color='FFFFFF', size=12)
        ws.cell(totals_row, 26).fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        ws.cell(totals_row, 26).font = Font(bold=True, color='FFFFFF', size=12)
        ws.cell(totals_row, 27).font = Font(bold=True)

        for col in range(1, total_cols + 1):
            ws.cell(totals_row, col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='double'), bottom=Side(style='thin')
            )

        # Set column widths (28 columns)
        widths = [15, 18, 16, 12, 10, 15, 11, 8, 14, 10, 10, 8, 10, 8, 10, 8, 10, 8, 8, 10, 10, 10, 12, 14, 12, 12, 8, 12]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Apply highlighting to deduction columns
        for r in range(5, data_row):
            # Late Penalty - Yellow
            ws.cell(r, 9).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            # Punch Deduction - Pink
            ws.cell(r, 11).fill = PatternFill(start_color='FFD9E6', end_color='FFD9E6', fill_type='solid')
            # Absence Deduction - Red
            ws.cell(r, 13).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            # Early Departure Deduction - Orange
            ws.cell(r, 15).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            # Half Day Deduction - Orange
            ws.cell(r, 17).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            # Sick Leave Deduction - Light Blue
            ws.cell(r, 19).fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            # Unpaid Leave Deduction - Light Blue
            ws.cell(r, 21).fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            # Refund Count - Light Green
            ws.cell(r, 22).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            # Refund Days - Light Green
            ws.cell(r, 23).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            ws.cell(r, 23).font = Font(bold=True, color='375623')
            # Total Penalty - Bold Red background
            ws.cell(r, 24).fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            ws.cell(r, 24).font = Font(bold=True)
            # Gross Deduction - Bold Red background
            ws.cell(r, 25).fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            ws.cell(r, 25).font = Font(bold=True)
            # Net Deduction - Bold Red background
            ws.cell(r, 26).fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            ws.cell(r, 26).font = Font(bold=True)
            # Warnings - Yellow
            ws.cell(r, 27).fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            ws.cell(r, 27).font = Font(bold=True, color='856404')
            # Backdated Leaves - Light Purple (same as summary sheet)
            ws.cell(r, 28).fill = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
            ws.cell(r, 28).font = Font(bold=True, color='6B3FA0')

        # Add policy legend with all justification deduction rules
        legend_row = totals_row + 3
        ws.cell(legend_row, 1, "Deduction Policy Reference:").font = Font(bold=True)
        legend_row += 1
        ws.cell(legend_row, 1, f"• Late: {currency} {late_1st} (1st), {currency} {late_2nd} (2nd), {currency} {late_3rd}+ (3rd+) + Warning | Late (Approved): No Deduction")
        legend_row += 1
        ws.cell(legend_row, 1, f"• Missing Punch (all types): {missing_deduction_rate} day after {missing_threshold} occurrences, warning after 6")
        legend_row += 1
        ws.cell(legend_row, 1, f"• Absent: {absence_deduction_rate} days | Early Departure/Half Day: {early_dep_deduction} day | Sick Leave: {sick_leave_deduction} day | Unpaid Leave: {unpaid_leave_deduction} day")
        legend_row += 1
        ws.cell(legend_row, 1, "• No Deduction: Normal, Late (Approved), Early Departure (Approved), Annual/Casual/Marriage/Paternity/Maternity/Bereavement/Military Call Leave")
        legend_row += 1
        ws.cell(legend_row, 1, "• REFUND LEAVES: Annual Leave (Refund) +1 day, Sick Leave (Refund) +0.75 days, Half Day (Refund) +0.5 days - reversal of previous penalties")
        ws.cell(legend_row, 1).font = Font(bold=True, color='375623')
        legend_row += 1
        ws.cell(legend_row, 1, "• SIGN CONVENTION: Deductions shown as negative (-), Refunds shown as positive (+). Net Ded. = Gross Ded. + Refunds")
        legend_row += 1
        ws.cell(legend_row, 1, "• BACKDATED LEAVES (BD): Leaves marked with (BD) suffix are transferred from previous months - shown in purple, HR to review manually")
        ws.cell(legend_row, 1).font = Font(bold=True, color='6B3FA0')
        legend_row += 1
        ws.cell(legend_row, 1, "• Note: This sheet is linked to Summary Report - use the dropdown to change status and penalties will auto-update")
        ws.cell(legend_row, 1).font = Font(italic=True, color='0066CC')

        self.logger.info(f"Created penalties sheet with formulas linked to Summary Report (28 columns including Refund and Net Deduction)")

    def create_duplicates_sheet(self, wb):
        """Create the Duplicates sheet showing leave vs attendance conflicts.
        Lists employees who had attendance records but also had leave applied for the same date."""
        ws = wb.create_sheet("Duplicates", 4)

        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "Leave vs Attendance Conflicts"
        title_cell.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells('A2:G2')
        subtitle = ws['A2']
        subtitle.value = "Employees with both attendance and leave on the same date"
        subtitle.font = Font(size=9, italic=True)
        subtitle.alignment = Alignment(horizontal='center')

        # Headers
        headers = ['CRM', 'Name', 'Date', 'Original Attendance Status', 'Leave Type Applied', 'Clock In', 'Clock Out']
        header_row = 4

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        ws.row_dimensions[header_row].height = 25

        # Data rows
        if self.conflict_records:
            data_row = 5
            for conflict in sorted(self.conflict_records, key=lambda x: (x['crm'], x['date'])):
                # Format clock times
                clock_in = conflict.get('clock_in')
                clock_out = conflict.get('clock_out')
                clock_in_str = str(clock_in) if clock_in else '-'
                clock_out_str = str(clock_out) if clock_out else '-'

                # Format date
                date_val = conflict['date']
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%d-%b-%Y')
                else:
                    date_str = str(date_val)

                ws.cell(data_row, 1, conflict['crm'])
                ws.cell(data_row, 2, conflict['name'])
                ws.cell(data_row, 3, date_str)
                ws.cell(data_row, 4, conflict['attendance_status'])
                ws.cell(data_row, 5, conflict['leave_type'])
                ws.cell(data_row, 6, clock_in_str)
                ws.cell(data_row, 7, clock_out_str)

                # Yellow row highlighting
                for col in range(1, 8):
                    cell = ws.cell(data_row, col)
                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                data_row += 1

            # Summary row
            summary_row = data_row + 1
            ws.cell(summary_row, 1, f"Total Conflicts: {len(self.conflict_records)}")
            ws.cell(summary_row, 1).font = Font(bold=True)
        else:
            # No conflicts detected
            ws.merge_cells('A5:G5')
            no_conflict_cell = ws['A5']
            no_conflict_cell.value = "No conflicts detected - All leave records are consistent with attendance data"
            no_conflict_cell.font = Font(size=11, color='28A745')
            no_conflict_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[5].height = 30

        # Set column widths
        widths = [15, 20, 12, 25, 20, 12, 12]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        self.logger.info(f"Created duplicates sheet with {len(self.conflict_records)} conflicts")

    def clear_all(self):
        """Clear all data and reset"""
        response = messagebox.askyesno(
            "Confirm Clear",
            "Are you sure you want to clear all data and start over?"
        )

        if response:
            self.master_file = None
            self.attendance_files = []
            self.leave_file = None
            self.employee_mapping = {}
            self.leave_records = []
            self.processed_data = []
            self.conflict_records = []

            # Clear month/year prompt cache
            if hasattr(self, '_user_month_year'):
                delattr(self, '_user_month_year')

            # Clear filter state and hide filter frame
            self.department_vars = {}
            self.crm_vars = {}
            self.available_departments = []
            self.available_crms = []
            if self.filter_frame:
                self.filter_frame.pack_forget()

            self.master_status.configure(text="Required", text_color=self.COLORS['text_secondary'])
            self.attendance_status.configure(text="Required", text_color=self.COLORS['text_secondary'])
            self.leave_status.configure(text="Optional", text_color=self.COLORS['text_secondary'])
            self.progress_label.configure(text="Ready", text_color=self.COLORS['text_secondary'])

            self.status_text.delete('1.0', 'end')
            self.log_message("All data cleared. Ready to start over.")
            self.logger.info("All data cleared by user")

            self.update_generate_button()


def main():
    """Main entry point for the application"""
    app = AttendanceDashboard()
    app.root.mainloop()


if __name__ == "__main__":
    main()
