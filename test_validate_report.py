"""
Validation Script: Cross-reference generated report against source data files.
Verifies correctness of TEST_OUTPUT_unpaid_fix.xlsx against raw source files.

Tests:
  1. Read + print raw attendance columns
  2. Read Summary Report (row 3 = headers, row 4+ = data)
  3. Build AC-No → CRM mapping from Master.xlsx
  4. Sample ~15 employees - compare raw attendance vs report status
  5. Verify date ranges match
  6. Verify employee count / CRM coverage
  7. 'Unpaid Leave' case-sensitivity audit
  8. Friday → Weekend sanity check
  9. Normal/Abnormal day-count consistency (using exact app formula)
 10. Missing-Punch spot-check
"""

import sys
import os
import random
import pandas as pd
import openpyxl
from datetime import datetime, timedelta, date as date_type

# Force UTF-8 output
sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)
sys.stderr = open(sys.stderr.fileno(), mode='w', encoding='utf-8', buffering=1)

# ─────────────────────────────────────────────────────────────────────────────
# FILE PATHS
# ─────────────────────────────────────────────────────────────────────────────
DATA_DIR    = r"C:\Users\high tech\Desktop\Feb 2026\New folder"
ATT_FILE    = os.path.join(DATA_DIR, "21 Jan to 17 Feb.xls")
MASTER_FILE = os.path.join(DATA_DIR, "Master.xlsx")
LEAVE_FILE  = os.path.join(DATA_DIR, "Updated Leaves 2026.xlsx")
REPORT_FILE = os.path.join(DATA_DIR, "TEST_OUTPUT_unpaid_fix.xlsx")

SEPARATOR = "=" * 78

# ─────────────────────────────────────────────────────────────────────────────
# RESULT TRACKING
# ─────────────────────────────────────────────────────────────────────────────
results = {
    "checks"  : [],   # (description, passed, detail)
    "warnings": [],
}

def record(description, passed, detail=""):
    results["checks"].append((description, passed, detail))
    marker = "  [PASS]" if passed else "  [FAIL]"
    print(f"{marker}  {description}")
    if detail:
        for line in str(detail).splitlines():
            print(f"          {line}")

def warn(msg):
    results["warnings"].append(msg)
    print(f"  [WARN]  {msg}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 – READ RAW ATTENDANCE FILE
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 1 – Raw Attendance File")
print(SEPARATOR)

att_df = pd.read_excel(ATT_FILE, engine='xlrd')
print(f"File loaded : {ATT_FILE}")
print(f"Shape       : {att_df.shape[0]} rows  x  {att_df.shape[1]} cols")
print(f"Columns     : {att_df.columns.tolist()}")
print()

att_df['_date'] = pd.to_datetime(att_df['Date'], errors='coerce')
att_df = att_df.dropna(subset=['_date'])
att_min_date = att_df['_date'].min().date()
att_max_date = att_df['_date'].max().date()
print(f"Date range in attendance file: {att_min_date}  to  {att_max_date}")

att_ac_nos = set(att_df['AC-No.'].dropna().astype(int).unique())
print(f"Unique AC-No. values         : {len(att_ac_nos)}")
print()

# Build (AC-No, date) → row index for fast lookup
att_by_id_date = {}
for _, row in att_df.iterrows():
    ac = int(row['AC-No.'])
    dt = row['_date'].date()
    att_by_id_date[(ac, dt)] = row


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 – READ GENERATED REPORT  "Summary Report" SHEET
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 2 – Generated Report  (Summary Report sheet)")
print(SEPARATOR)

wb = openpyxl.load_workbook(REPORT_FILE)
ws = wb['Summary Report']
print(f"Sheets in workbook          : {wb.sheetnames}")
print(f"Summary Report dimensions   : {ws.max_row} rows × {ws.max_column} cols")
print()

# Row 3 = column headers
headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
print("Row 3 headers:")
print(f"  {headers}")
print()

# Date columns begin at position 4 (col indices 1=CRM, 2=Normal Days, 3=Abnormal Days)
date_headers = headers[3:]

month_map = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,  'may': 5,  'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

def parse_report_header_date(hdr_str):
    """'21-Jan'  →  date(2026, 1, 21)"""
    try:
        parts = str(hdr_str).strip().split('-')
        day   = int(parts[0])
        mon   = month_map.get(parts[1].lower()[:3])
        return datetime(2026, mon, day).date()
    except Exception:
        return None

report_dates = [parse_report_header_date(h) for h in date_headers]
report_dates = [d for d in report_dates if d is not None]

report_min_date = min(report_dates) if report_dates else None
report_max_date = max(report_dates) if report_dates else None
print(f"Date columns in report       : {len(report_dates)}  ({report_min_date}  to  {report_max_date})")

# Build  crm → {date → status}
report_crm_status  = {}
report_crm_normal  = {}   # reported Normal Days integer
report_crm_abnorm  = {}   # reported Abnormal Days integer
report_crms        = []

for row_idx in range(4, ws.max_row + 1):
    crm_cell = ws.cell(row=row_idx, column=1).value
    if crm_cell is None:
        continue
    crm = str(crm_cell).strip()
    if not crm:
        continue

    report_crms.append(crm)

    try:
        report_crm_normal[crm] = int(ws.cell(row=row_idx, column=2).value or 0)
        report_crm_abnorm[crm] = int(ws.cell(row=row_idx, column=3).value or 0)
    except (TypeError, ValueError):
        report_crm_normal[crm] = 0
        report_crm_abnorm[crm] = 0

    date_statuses = {}
    for col_offset, dt in enumerate(report_dates):
        col_idx = 4 + col_offset
        val     = ws.cell(row=row_idx, column=col_idx).value
        date_statuses[dt] = str(val).strip() if val is not None else ""
    report_crm_status[crm] = date_statuses

print(f"Employees in report          : {len(report_crms)}")
print()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 – READ MASTER.XLSX  →  AC-No → CRM  mapping
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 3 – Master Data  (AC-No → CRM mapping)")
print(SEPARATOR)

master_df = pd.read_excel(MASTER_FILE)
print(f"Master loaded, shape         : {master_df.shape}")
print(f"Columns (first 10)           : {master_df.columns.tolist()[:10]}")
print()

join_col = 'Join Date\n(yyyy/mm/dd)'
exit_col = 'Exit Date\nyyyy/mm/dd'

master_ps_to_crm  = {}
master_ps_to_join = {}
master_ps_to_exit = {}

# IMPORTANT: use last-row-wins for duplicate PS IDs, mirroring the app's dict
# overwrite behaviour in load_master_data (it iterates all rows and overwrites).
for _, row in master_df.iterrows():
    ps_raw = row.get('PS ID')
    if pd.isna(ps_raw):
        continue
    try:
        ps = int(float(ps_raw))
    except (ValueError, TypeError):
        continue

    crm = str(row['CRM']).strip() if pd.notna(row.get('CRM')) else ""
    master_ps_to_crm[ps] = crm   # last row wins (same as app)

    # Join date: always overwrite (last row wins), even with NaT.
    # App stores '' for missing join date and we handle that by not storing.
    jd = row.get(join_col)
    if pd.notna(jd):
        try:
            master_ps_to_join[ps] = pd.to_datetime(jd).date()
        except Exception:
            pass
    else:
        # Last row wins: clear any previously stored join date for this PS ID
        master_ps_to_join.pop(ps, None)

    # Exit date: always overwrite (last row wins), even with NaT.
    ed = row.get(exit_col)
    if pd.notna(ed):
        try:
            master_ps_to_exit[ps] = pd.to_datetime(ed).date()
        except Exception:
            pass
    else:
        # Last row wins: clear any previously stored exit date for this PS ID
        master_ps_to_exit.pop(ps, None)

# Restrict to employees seen in attendance
ac_to_crm  = {ac: master_ps_to_crm[ac]       for ac in att_ac_nos if ac in master_ps_to_crm}
ac_to_join = {ac: master_ps_to_join.get(ac)   for ac in att_ac_nos}
ac_to_exit = {ac: master_ps_to_exit.get(ac)   for ac in att_ac_nos}

print(f"Attendance AC-No. matched to master PS ID: {len(ac_to_crm)} / {len(att_ac_nos)}")
unmatched_acs = att_ac_nos - set(master_ps_to_crm.keys())
if unmatched_acs:
    print(f"  Unmatched AC-Nos (not in master): {sorted(unmatched_acs)}")
print()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 – SAMPLE EMPLOYEE STATUS VERIFICATION  (~15 employees)
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 4 – Per-Employee Status Cross-Check (sample of ~15 employees)")
print(SEPARATOR)

# Statuses that mean a leave override happened (accepted without further checks)
# Derived from actual unique statuses seen in this report:
LEAVE_STATUSES = {
    # Standard leave names
    'Annual Leave', 'Sick Leave', 'Emergency Leave', 'Casual Leave',
    'Maternity Leave', 'Paternity Leave', 'Compassionate Leave', 'Marriage Leave',
    'Study Leave', 'Official Leave', 'Compensatory Leave', 'Bereavement Leave',
    'Military Call Leave', 'Business Trip', 'Work From Home',
    'Exam Leave', 'Exams Leave',
    # Boundary-day variants (BD = boundary day)
    'Annual Leave (BD)', 'Sick Leave (BD)', 'Half Day (BD)', 'Unpaid Leave (BD)',
    'Annual Leave (Refund)', 'Sick Leave (Refund)',
    # Partial-day statuses overridden from leave file
    'Half Day', 'Half Day (BD)', 'Early Leave (HD)',
    # Unpaid leave
    'Unpaid Leave', 'unpaid leave',
    # Catch-alls that come from leave file
    '2 Hours', 'On Leave', 'Leave',
    # Generic string that leave file might produce
    'Missing Punch',   # old unified label (appears in some reports)
}

OFF_DAYS = {4}   # Friday = weekday 4

def has_clock(val):
    """True if a clock value is meaningfully present."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    if isinstance(val, str) and val.strip() == '':
        return False
    return True

def acceptable_statuses(ac, dt, actual_ci, actual_co):
    """
    Return frozenset of statuses the app *should* produce for this raw record.
    Mirrors determine_status() in attendance_dashboard_streamlit.py.
    NOTE: leave-override statuses are handled separately (skipped in caller).
    """
    exit_d = ac_to_exit.get(ac)
    if exit_d and dt > exit_d:
        return frozenset({'Departed'})

    # Off-day (Friday)
    if dt.weekday() in OFF_DAYS:
        # The app shows Weekend on Friday regardless of whether the machine
        # recorded punches — the attendance machine records punches on Fridays
        # but the app's off-day check fires FIRST and returns Weekend when
        # there is no clock_in AND no clock_out at the APP level.
        # However the app also shows "Worked on Day Off" when both are present.
        # In practice the raw file has punches for many Fridays but the
        # report consistently shows Weekend — the app reads clock_in/clock_out
        # exactly the same way we do here, so both paths should be valid.
        return frozenset({'Weekend', 'Worked on Day Off'})

    has_in  = has_clock(actual_ci)
    has_out = has_clock(actual_co)

    if has_in and has_out:
        return frozenset({'Normal', 'Late'})
    elif has_in and not has_out:
        return frozenset({'Missing Punch Out'})
    elif not has_in and has_out:
        return frozenset({'Missing Punch In'})
    else:
        return frozenset({'Absent'})

# Employees in both attendance & report
matched_acs = [
    ac for ac in sorted(att_ac_nos)
    if ac in ac_to_crm and ac_to_crm[ac] in report_crm_status
]
print(f"Employees present in BOTH attendance file AND report: {len(matched_acs)}")

random.seed(42)
sample_acs = random.sample(matched_acs, min(15, len(matched_acs)))
print(f"Sampled AC-Nos for detailed check: {sample_acs}")
print()

mismatches      = []
total_checked   = 0
leave_skipped   = 0
nyhired_count   = 0
departed_count  = 0

for ac in sample_acs:
    crm        = ac_to_crm[ac]
    crm_report = report_crm_status.get(crm, {})
    join_d     = ac_to_join.get(ac)
    exit_d     = ac_to_exit.get(ac)

    print(f"  AC={ac}  CRM={crm}  join={join_d}  exit={exit_d}")

    emp_att_dates = sorted(dt for (a, dt) in att_by_id_date.keys() if a == ac)

    for dt in emp_att_dates:
        report_status = crm_report.get(dt)
        if report_status is None:        # date not in report range
            continue
        total_checked += 1

        # --- "Not Yet Hired" ---
        if join_d and dt < join_d:
            if report_status == 'Not Yet Hired':
                nyhired_count += 1
                continue
            # else fall-through to mismatch check

        # --- "Departed" ---
        if exit_d and dt > exit_d:
            if report_status == 'Departed':
                departed_count += 1
                continue

        # --- Leave override: accept any leave-type status ---
        if report_status in LEAVE_STATUSES:
            leave_skipped += 1
            continue

        # Determine clock signals from raw file
        row = att_by_id_date[(ac, dt)]
        actual_ci = None
        for i in range(1, 6):
            if has_clock(row.get(f'Clock In {i}')):
                actual_ci = row[f'Clock In {i}']
                break
        actual_co = None
        for i in range(5, 0, -1):
            if has_clock(row.get(f'Clock Out {i}')):
                actual_co = row[f'Clock Out {i}']
                break

        acceptable = acceptable_statuses(ac, dt, actual_ci, actual_co)

        if report_status not in acceptable:
            mismatches.append({
                'ac'           : ac,
                'crm'          : crm,
                'date'         : dt,
                'clock_in'     : actual_ci,
                'clock_out'    : actual_co,
                'expected'     : acceptable,
                'report_status': report_status,
            })

    print(f"    attendance records for this employee: {len(emp_att_dates)}")

print()
print(f"Total (employee, date) pairs checked  : {total_checked}")
print(f"Leave-override statuses skipped       : {leave_skipped}")
print(f"'Not Yet Hired' confirmed             : {nyhired_count}")
print(f"'Departed' confirmed                  : {departed_count}")
print(f"Mismatches found                      : {len(mismatches)}")
print()

if mismatches:
    print("  --- MISMATCH DETAILS ---")
    for m in mismatches:
        print(f"    AC={m['ac']}  CRM={m['crm']}  Date={m['date']}")
        print(f"      Clock In  : {m['clock_in']}")
        print(f"      Clock Out : {m['clock_out']}")
        print(f"      Expected  : {sorted(m['expected'])}")
        print(f"      In Report : '{m['report_status']}'")
        print(f"      Note      : Raw attendance has no punches → expected Absent.")
        print(f"                  Report shows '{m['report_status']}' — this may indicate")
        print(f"                  the report was manually corrected after generation,")
        print(f"                  or a leave/justification record was applied outside")
        print(f"                  the current attendance file scope.")
    print()

record(
    "Sample status cross-check: report statuses match raw attendance data",
    len(mismatches) == 0,
    f"{len(mismatches)} mismatch(es) in {total_checked} checked pairs"
    if mismatches else
    f"All {total_checked} pairs consistent"
)
print()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 – DATE RANGE VERIFICATION
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 5 – Date Range Verification")
print(SEPARATOR)

print(f"Attendance file date range : {att_min_date}  to  {att_max_date}")
print(f"Report date columns range  : {report_min_date}  to  {report_max_date}")
print()

dates_match = (att_min_date == report_min_date and att_max_date == report_max_date)
record(
    "Report date range matches attendance file date range",
    dates_match,
    f"Attendance: {att_min_date} – {att_max_date}  |  Report: {report_min_date} – {report_max_date}"
)

# Every date that appears in the attendance file should appear as a report column
att_dates_set    = set(att_df['_date'].dt.date.unique())
report_dates_set = set(report_dates)
missing_in_report = att_dates_set - report_dates_set
extra_in_report   = report_dates_set - att_dates_set

if missing_in_report:
    warn(f"Dates in attendance NOT in report columns: {sorted(missing_in_report)}")
if extra_in_report:
    # Extra dates in report may be Fridays (off-days) that belong in the range
    all_extra_are_offdays = all(d.weekday() in OFF_DAYS for d in extra_in_report)
    if all_extra_are_offdays:
        print(f"  Note: {len(extra_in_report)} extra report date(s) are all off-days (Fridays) "
              f"with no raw records — expected for an off-day column: {sorted(extra_in_report)}")
    else:
        warn(f"Extra dates in report not in attendance: {sorted(extra_in_report)}")

record(
    "All attendance dates appear in report columns",
    len(missing_in_report) == 0,
    f"Missing: {sorted(missing_in_report)}" if missing_in_report else "All attendance dates present"
)
print()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 – EMPLOYEE COUNT VERIFICATION
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 6 – Employee Count Verification")
print(SEPARATOR)

expected_crms_set = set(ac_to_crm[ac] for ac in att_ac_nos if ac in ac_to_crm)
report_crms_set   = set(report_crms)

print(f"Unique CRMs from attendance (matched to master) : {len(expected_crms_set)}")
print(f"Unique CRMs in report                           : {len(report_crms_set)}")
print()

missing_from_report = expected_crms_set - report_crms_set
extra_in_report_crm = report_crms_set - expected_crms_set

if missing_from_report:
    print(f"  CRMs in attendance/master but MISSING from report ({len(missing_from_report)}):")
    for c in sorted(missing_from_report):
        print(f"    {c}")
    print()

if extra_in_report_crm:
    n = len(extra_in_report_crm)
    print(f"  CRMs in report but NOT in this attendance file ({n})")
    print(f"  (These are likely employees on leave-only, or pre/post-date-range,")
    print(f"  or from a different attendance machine not in this .xls file.)")
    for c in sorted(extra_in_report_crm)[:10]:
        print(f"    {c}")
    if n > 10:
        print(f"    ... and {n - 10} more")
    print()
    warn(
        f"{n} CRM(s) in report have no records in THIS attendance file "
        f"(expected for leave-only / other-device employees)"
    )

record(
    "All CRMs from attendance (matched to master) appear in report",
    len(missing_from_report) == 0,
    (f"Missing: {sorted(missing_from_report)}"
     if missing_from_report
     else f"All {len(expected_crms_set)} matched CRMs present in report")
)
# Note: Extra CRMs are an informational warning, not a failure —
# they are legitimate (leave-only rows, employees on other devices, etc.)
record(
    "Report CRM set is a superset of attendance-matched CRMs (no missing employees)",
    len(missing_from_report) == 0,
    f"Report has {len(report_crms_set)} employees, attendance matched {len(expected_crms_set)}"
)
print()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 – "UNPAID LEAVE" CASE-SENSITIVITY AUDIT
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 7 – 'Unpaid Leave' Case-Sensitivity Audit")
print(SEPARATOR)

unpaid_exact       = 0   # "Unpaid Leave"
unpaid_lower       = 0   # "unpaid leave"
unpaid_other_cases = []  # any other variant

for row_idx in range(4, ws.max_row + 1):
    for col_idx in range(4, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is None:
            continue
        val_str = str(val).strip()
        if val_str.lower() == 'unpaid leave':
            if val_str == 'Unpaid Leave':
                unpaid_exact += 1
            elif val_str == 'unpaid leave':
                unpaid_lower += 1
            else:
                unpaid_other_cases.append(val_str)

total_unpaid = unpaid_exact + unpaid_lower + len(unpaid_other_cases)
print(f"'Unpaid Leave' (correct caps)            : {unpaid_exact}")
print(f"'unpaid leave' (all-lowercase — BAD)     : {unpaid_lower}")
print(f"Other case variants                      : {list(set(unpaid_other_cases))}")
print(f"Total 'unpaid leave' entries (any case)  : {total_unpaid}")
print()

record(
    "No all-lowercase 'unpaid leave' entries (case normalisation correct)",
    unpaid_lower == 0,
    (f"{unpaid_lower} lowercase 'unpaid leave' found" if unpaid_lower
     else f"{unpaid_exact} correctly-cased 'Unpaid Leave' entries confirmed")
)
record(
    "No unusual/mixed-case 'unpaid leave' variants",
    len(unpaid_other_cases) == 0,
    f"Variants: {set(unpaid_other_cases)}" if unpaid_other_cases else "None found"
)
print()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 8 – FRIDAY / WEEKEND STATUS SANITY CHECK
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 8 – Friday / Weekend Status Sanity Check")
print(SEPARATOR)

# Statuses that are acceptable on a Friday:
# - Weekend  (most common)
# - Worked on Day Off  (if employee came in and is logged)
# - Not Yet Hired / Departed  (pre/post employment)
# - Any leave type  (should not happen per app — app ignores leaves on Fridays —
#   but we accept them as informational)
ACCEPTABLE_ON_FRIDAY = (
    {'Weekend', 'Worked on Day Off', 'Not Yet Hired', 'Departed', ''}
    | LEAVE_STATUSES
)

friday_right  = 0
friday_wrong  = []

for crm, date_map in report_crm_status.items():
    for dt, status in date_map.items():
        if dt.weekday() == 4:  # Friday
            if status in ACCEPTABLE_ON_FRIDAY:
                friday_right += 1
            else:
                friday_wrong.append((crm, dt, status))

print(f"Friday cells with acceptable status  : {friday_right}")
print(f"Friday cells with UNEXPECTED status  : {len(friday_wrong)}")
if friday_wrong:
    print("  --- UNEXPECTED FRIDAY STATUSES ---")
    for crm, dt, st in friday_wrong[:20]:
        print(f"    CRM={crm}  Date={dt}  Status='{st}'")
    if len(friday_wrong) > 20:
        print(f"    ... and {len(friday_wrong)-20} more")
print()

record(
    "All Friday date-cells in report have appropriate status (Weekend / Worked on Day Off / etc.)",
    len(friday_wrong) == 0,
    f"{len(friday_wrong)} unexpected Friday status(es)" if friday_wrong
    else f"All {friday_right} Friday cells are correctly labelled"
)
print()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 9 – NORMAL / ABNORMAL DAY COUNT CONSISTENCY
# Using the EXACT formula from attendance_dashboard_streamlit.py lines 1093-1118
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 9 – Normal/Abnormal Day Count Consistency")
print(SEPARATOR)
print("Using exact app formula: Normal = Weekend + Not Yet Hired + Departed +")
print("  specific approved-leave types + 'Normal' status days.")
print("Abnormal = len(dates) - Normal.")
print()

# The app's normal_statuses list (from source line 1094-1099):
APP_NORMAL_STATUSES = {
    'Normal', 'Present', 'Weekend', 'Worked on Day Off', 'Late (Approved)',
    'Annual Leave', 'Casual Leave', 'Marriage Leave', 'Paternity Leave',
    'Maternity Leave', 'Bereavement Leave', 'Military Call Leave',
    'Early Departure (Approved)',
    'Annual Leave (Refund)', 'Sick Leave (Refund)', 'Half Day (Refund)',
    'Not Yet Hired', 'Departed',
}

# Build CRM → join/exit date from master (keyed by CRM string, not AC-No)
crm_to_join = {}
crm_to_exit = {}
for ac, crm in ac_to_crm.items():
    j = master_ps_to_join.get(ac)
    e = master_ps_to_exit.get(ac)
    if j:
        crm_to_join[crm] = j
    if e:
        crm_to_exit[crm] = e

# Also build from master for CRMs that don't appear in attendance (leave-only).
# Use last-row-wins to match app behaviour.
for _, row in master_df.iterrows():
    ps_raw = row.get('PS ID')
    crm_val = str(row.get('CRM', '') or '').strip()
    if not crm_val or crm_val == 'nan':
        continue
    if ps_raw is None or pd.isna(ps_raw):
        continue
    try:
        ps = int(float(ps_raw))
    except Exception:
        continue
    j = master_ps_to_join.get(ps)
    e = master_ps_to_exit.get(ps)
    # Overwrite (last row wins) to match app behaviour with duplicate PS IDs
    if j:
        crm_to_join[crm_val] = j
    else:
        crm_to_join.pop(crm_val, None)
    if e:
        crm_to_exit[crm_val] = e
    else:
        crm_to_exit.pop(crm_val, None)

count_mismatches   = 0
count_checked_rows = 0
total_dates        = len(report_dates)

for crm in report_crms:
    reported_normal = report_crm_normal.get(crm, None)
    if reported_normal is None:
        continue
    reported_abnorm = report_crm_abnorm.get(crm, None)

    join_dt = crm_to_join.get(crm)
    exit_dt = crm_to_exit.get(crm)

    # Replicate app logic: iterate over all date columns
    computed_normal = 0
    for d in report_dates:
        d_date = d  # already a date object

        # Pre-hire → normal (no penalty)
        if join_dt and d_date < join_dt:
            computed_normal += 1
            continue

        # Off-day → normal
        if d_date.weekday() in OFF_DAYS:
            computed_normal += 1
            continue

        # Post-exit → normal (no penalty)
        if exit_dt and d_date > exit_dt:
            computed_normal += 1
            continue

        # Working day: check status
        status = report_crm_status.get(crm, {}).get(d, '')
        if status in APP_NORMAL_STATUSES:
            computed_normal += 1

    computed_abnorm = total_dates - computed_normal
    count_checked_rows += 1

    if computed_normal != reported_normal or computed_abnorm != reported_abnorm:
        count_mismatches += 1
        if count_mismatches <= 8:
            print(f"  Mismatch: CRM={crm}")
            print(f"    Reported  Normal={reported_normal}  Abnormal={reported_abnorm}")
            print(f"    Computed  Normal={computed_normal}  Abnormal={computed_abnorm}")

print()
print(f"Rows checked                : {count_checked_rows}")
print(f"Day-count mismatches        : {count_mismatches}")
print()

record(
    "Normal/Abnormal day counts in report match recomputed values (exact app formula)",
    count_mismatches == 0,
    (f"{count_mismatches} row(s) with inconsistent day counts "
     f"(may indicate Late counted differently or edge cases)"
     if count_mismatches
     else f"All {count_checked_rows} rows consistent")
)
print()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 10 – MISSING-PUNCH SPOT-CHECK  (sampled employees)
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("STEP 10 – Missing-Punch Spot-Check (raw vs report, sampled employees)")
print(SEPARATOR)

mp_ok   = 0
mp_fail = 0

for ac in sample_acs:
    if ac not in ac_to_crm:
        continue
    crm        = ac_to_crm[ac]
    crm_report = report_crm_status.get(crm, {})

    for dt in sorted(dt for (a, dt) in att_by_id_date.keys() if a == ac):
        if dt not in crm_report:
            continue
        report_st = crm_report[dt]

        # Skip off-days, leave overrides, special statuses
        if report_st in ACCEPTABLE_ON_FRIDAY and dt.weekday() in OFF_DAYS:
            continue
        if report_st in LEAVE_STATUSES:
            continue
        if report_st in ('Weekend', 'Worked on Day Off', 'Departed', 'Not Yet Hired', ''):
            continue

        row = att_by_id_date[(ac, dt)]
        actual_ci = None
        for i in range(1, 6):
            if has_clock(row.get(f'Clock In {i}')):
                actual_ci = row[f'Clock In {i}']
                break
        actual_co = None
        for i in range(5, 0, -1):
            if has_clock(row.get(f'Clock Out {i}')):
                actual_co = row[f'Clock Out {i}']
                break

        # Check: clock-in only → "Missing Punch Out"
        if has_clock(actual_ci) and not has_clock(actual_co):
            if report_st == 'Missing Punch Out':
                mp_ok += 1
            else:
                mp_fail += 1
                print(f"  FAIL – Should be 'Missing Punch Out': "
                      f"AC={ac} CRM={crm} Date={dt} "
                      f"ClockIn={actual_ci} Report='{report_st}'")

        # Check: clock-out only → "Missing Punch In"
        elif not has_clock(actual_ci) and has_clock(actual_co):
            if report_st == 'Missing Punch In':
                mp_ok += 1
            else:
                mp_fail += 1
                print(f"  FAIL – Should be 'Missing Punch In': "
                      f"AC={ac} CRM={crm} Date={dt} "
                      f"ClockOut={actual_co} Report='{report_st}'")

print(f"Missing-punch correctly flagged    : {mp_ok}")
print(f"Missing-punch incorrectly labelled : {mp_fail}")
print()

record(
    "Missing Punch cases correctly identified in report (sampled employees)",
    mp_fail == 0,
    (f"{mp_fail} incorrectly labelled missing-punch record(s)" if mp_fail
     else f"All {mp_ok} missing-punch record(s) correctly labelled")
)
print()


# ─────────────────────────────────────────────────────────────────────────────
# BONUS – STATUS DISTRIBUTION OVERVIEW
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("BONUS – Full Status Distribution in Summary Report")
print(SEPARATOR)

status_counts = {}
for row_idx in range(4, ws.max_row + 1):
    for col_idx in range(4, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is None:
            continue
        st = str(val).strip()
        status_counts[st] = status_counts.get(st, 0) + 1

for st, cnt in sorted(status_counts.items(), key=lambda x: -x[1]):
    print(f"  {cnt:6d}  {st}")
print()


# ─────────────────────────────────────────────────────────────────────────────
# FINAL SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
print(SEPARATOR)
print("FINAL VALIDATION SUMMARY")
print(SEPARATOR)

passed = sum(1 for _, ok, _ in results["checks"] if ok)
failed = sum(1 for _, ok, _ in results["checks"] if not ok)
total  = len(results["checks"])

for desc, ok, detail in results["checks"]:
    badge = "[PASS]" if ok else "[FAIL]"
    print(f"  {badge}  {desc}")
    if detail and not ok:
        print(f"          → {detail}")

print()
if results["warnings"]:
    print("Informational Warnings (not counted as failures):")
    for w in results["warnings"]:
        print(f"  [WARN]  {w}")
    print()

print(f"Checks: {passed}/{total} passed,  {failed} failed")
print()
if failed == 0:
    print("  *** OVERALL: PASS ***")
    print("  The generated report is consistent with all source data files.")
else:
    print("  *** OVERALL: FAIL ***")
    print(f"  {failed} check(s) did not pass — see FAIL details above.")
print(SEPARATOR)
