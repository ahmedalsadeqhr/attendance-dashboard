"""
Attendance Dashboard Streamlit Web Application v2.2
Web-based attendance processing with Streamlit interface
EXACT REPLICA of desktop version report generation
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import io
import re


def read_excel_file(file_obj, filename):
    """Read Excel file with automatic engine detection for .xls and .xlsx files."""
    filename_lower = filename.lower()
    if filename_lower.endswith('.xls') and not filename_lower.endswith('.xlsx'):
        return pd.read_excel(file_obj, engine='xlrd')
    else:
        return pd.read_excel(file_obj)


# Default configuration
DEFAULT_CONFIG = {
    'work_start_time': '12:00 PM',
    'work_end_time': '9:00 PM',
    'late_threshold': '12:00 PM',
    'off_days': [4],  # Friday
    'alert_thresholds': {
        'high_absences': 3,
        'frequent_late': 5,
        'missing_punches': 5,
        'low_attendance_rate': 75.0
    },
    'penalties': {
        'late_1st': 100,
        'late_2nd': 200,
        'late_3rd': 500,
        'late_4th_plus': 500,
        'missing_punch_deduction': 0.5,
        'missing_punch_threshold': 3,
        'missing_punch_warning_threshold': 6,
        'early_departure_deduction': 0.5,
        'absence_deduction': 2,
        'currency': 'EGP'
    }
}


class AttendanceProcessor:
    """Core attendance processing logic - EXACT REPLICA of desktop version"""

    def __init__(self, config=None):
        self.config = config or DEFAULT_CONFIG.copy()
        self.employee_mapping = {}
        self.leave_records = []
        self.processed_data = []
        self.conflict_records = []
        self.penalties_data = {}
        self.logs = []

    def log_message(self, message, level='info'):
        """Add log message"""
        self.logs.append({'level': level, 'message': message})

    def normalize_id(self, value):
        """Normalize ID values by converting floats to integers before string conversion.

        This fixes the mismatch between float PS IDs (185680.0) and int AC-No (185680).
               """
        if pd.isna(value):
            return ""
        # If it's a float that represents a whole number, convert to int first
        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()
        # Try to convert string floats like "185680.0" to int
        try:
            float_val = float(value)
            if float_val.is_integer():
                return str(int(float_val)).strip()
        except (ValueError, TypeError):
            pass
        return str(value).strip()

    def find_column(self, df, search_terms, exact_matches=None):
        """Find column in dataframe using multiple search strategies.

        - 4 strategies.

        Args:
            df: DataFrame to search
            search_terms: List of terms to search for (e.g., ['ac', 'no'])
            exact_matches: Optional list of exact column names to try first

        Returns:
            Column name or None
        """
        # Strategy 1: Try exact matches first
        if exact_matches:
            for exact in exact_matches:
                if exact in df.columns:
                    return exact

        # Strategy 2: Try case-insensitive exact match
        if exact_matches:
            df_cols_lower = {str(c).lower(): c for c in df.columns}
            for exact in exact_matches:
                if exact.lower() in df_cols_lower:
                    return df_cols_lower[exact.lower()]

        # Strategy 3: Search for all terms in column name
        for col in df.columns:
            col_lower = str(col).lower()
            if all(term.lower() in col_lower for term in search_terms):
                return col

        # Strategy 4: Search for any term (less strict)
        for col in df.columns:
            col_lower = str(col).lower()
            if any(term.lower() in col_lower for term in search_terms):
                return col

        return None

    def validate_master_data(self, df):
        """Validate master data has required columns."""
        df_cols_lower = [str(c).lower() for c in df.columns]

        has_id_col = any('ac' in col and 'no' in col for col in df_cols_lower) or \
                     any('ps' in col and 'id' in col for col in df_cols_lower) or \
                     any('psid' in col for col in df_cols_lower)

        has_crm = any('crm' in col for col in df_cols_lower)
        has_name = any('name' in col for col in df_cols_lower)

        missing = []
        if not has_id_col:
            missing.append('AC-No. or PS ID')
        if not has_crm:
            missing.append('CRM')
        if not has_name:
            missing.append('Name')

        if missing:
            raise ValueError(f"Master data missing required columns: {', '.join(missing)}")

        if len(df) == 0:
            raise ValueError("Master data file is empty")

        return True

    def validate_attendance_file(self, df, filename):
        """Validate attendance file has required columns."""
        required_patterns = [
            ('ac', 'no'),
            ('name',),
            ('clock', 'in'),
        ]

        df_cols_lower = [str(c).lower() for c in df.columns]

        for pattern in required_patterns:
            if not any(all(p in col for p in pattern) for col in df_cols_lower):
                raise ValueError(f"Attendance file '{filename}' missing columns matching pattern: {pattern}")

        if len(df) == 0:
            raise ValueError(f"Attendance file '{filename}' is empty")

        return True

    def validate_leave_sheet(self, df):
        """Validate leave sheet has required structure."""
        date_cols = [c for c in df.columns if self._is_date_column(c)]

        if len(date_cols) > 0:
            df_cols_lower = [str(c).lower() for c in df.columns]
            if not any('crm' in col for col in df_cols_lower):
                raise ValueError("Leave sheet (matrix format) missing CRM column")
        else:
            df_cols_lower = [str(c).lower() for c in df.columns]
            has_crm = any('crm' in col for col in df_cols_lower)
            has_date = any('date' in col for col in df_cols_lower)

            if not has_crm or not has_date:
                raise ValueError("Leave sheet (vertical format) missing CRM or Date column")

        if len(df) == 0:
            raise ValueError("Leave sheet is empty")

        return True

    def _is_date_column(self, col_name):
        """Check if column name looks like a date."""
        if isinstance(col_name, datetime):
            return True
        if hasattr(col_name, 'date'):
            return True

        col_str = str(col_name)
        date_patterns = [
            r'\d{1,2}[-/]\w{3}',
            r'\w{3}[-/]\d{1,2}',
            r'\d{1,2}[-/]\d{1,2}',
            r'\d{4}[-/]\d{2}[-/]\d{2}',
        ]
        return any(re.search(pattern, col_str) for pattern in date_patterns)

    def _parse_date_field(self, row, date_col):
        """Parse a date field from a row."""
        if not date_col or pd.isna(row[date_col]):
            return ""
        try:
            raw_val = row[date_col]
            if isinstance(raw_val, (datetime, pd.Timestamp)):
                return raw_val.strftime('%Y-%m-%d')
            elif hasattr(raw_val, '__int__') and not isinstance(raw_val, str):
                excel_epoch = datetime(1899, 12, 30)
                dt = excel_epoch + timedelta(days=int(raw_val))
                return dt.strftime('%Y-%m-%d')
            else:
                dt = pd.to_datetime(raw_val, dayfirst=True)
                return dt.strftime('%Y-%m-%d')
        except Exception:
            return str(row[date_col]).strip()

    def load_master_data(self, file_obj, filename):
        """Load and process master data file -"""
        try:
            self.log_message("Loading master data...")
            df = read_excel_file(file_obj, filename)

            # Validate file
            self.validate_master_data(df)

            # Find columns - EXACT SAME LOGIC as desktop
            ac_col = self.find_column(df, ['ac', 'no'], ['AC-No.', 'Ac-No.', 'AC No', 'PS ID', 'PS Id', 'PSID'])
            if not ac_col:
                ac_col = self.find_column(df, ['ps', 'id'], ['PS ID', 'PS Id', 'PSID'])
            crm_col = self.find_column(df, ['crm'], ['CRM'])
            name_col = self.find_column(df, ['name'], ['Name'])
            dept_col = self.find_column(df, ['department', 'dept'], ['Department', 'Dept'])
            pos_col = self.find_column(df, ['position', 'title'], ['Position', 'Title'])
            national_id_col = self.find_column(df, ['identity', 'idnetity', 'national', 'nid'],
                                               ['Idnetity Number', 'Identity Number', 'National ID', 'NID'])
            vendor_col = self.find_column(df, ['vendor'], ['Vendor'])
            ps_id_col = self.find_column(df, ['ps', 'id'], ['PS ID', 'PS Id', 'PSID', 'PS-ID'])

            # Join Date column - support newlines in column names
            join_date_col = None
            for col in df.columns:
                col_clean = str(col).replace('\n', ' ').lower().strip()
                if 'join date' in col_clean or 'joining date' in col_clean:
                    join_date_col = col
                    break
            if not join_date_col:
                join_date_col = self.find_column(df, ['join', 'joining', 'hired'],
                                                 ['Join Date', 'Joining Date', 'Date of Joining', 'JoinDate',
                                                  'Hire Date', 'HireDate', 'Date Joined', 'DOJ'])

            # Exit Date column
            exit_date_col = None
            for col in df.columns:
                col_clean = str(col).replace('\n', ' ').lower().strip()
                if any(term in col_clean for term in ['exit date', 'resign', 'end date', 'termination', 'leaving']):
                    exit_date_col = col
                    break
            if not exit_date_col:
                exit_date_col = self.find_column(df, ['exit', 'resign', 'end', 'termination', 'leaving'],
                                                 ['Exit Date', 'Resignation Date', 'End Date', 'Termination Date',
                                                  'Leaving Date', 'Last Day', 'Last Working Day'])

            if not all([ac_col, crm_col, name_col]):
                self.log_message("Warning: Some required columns may be missing", 'warning')

            # Build employee mapping
            self.employee_mapping = {}
            for idx, row in df.iterrows():
                ac_no = self.normalize_id(row[ac_col]) if ac_col else ""
                if ac_no and ac_no != 'nan':
                    join_date_val = self._parse_date_field(row, join_date_col)
                    exit_date_val = self._parse_date_field(row, exit_date_col)

                    self.employee_mapping[ac_no] = {
                        'crm': str(row[crm_col]).strip() if crm_col and pd.notna(row[crm_col]) else "",
                        'name': str(row[name_col]).strip() if name_col and pd.notna(row[name_col]) else "",
                        'department': str(row[dept_col]).strip() if dept_col and pd.notna(row[dept_col]) else "",
                        'position': str(row[pos_col]).strip() if pos_col and pd.notna(row[pos_col]) else "",
                        'national_id': str(row[national_id_col]).strip() if national_id_col and pd.notna(row[national_id_col]) else "",
                        'vendor': str(row[vendor_col]).strip() if vendor_col and pd.notna(row[vendor_col]) else "",
                        'ps_id': self.normalize_id(row[ps_id_col]) if ps_id_col and pd.notna(row[ps_id_col]) else "",
                        'join_date': join_date_val,
                        'exit_date': exit_date_val
                    }

            self.log_message(f"Loaded {len(self.employee_mapping)} employee records", 'success')
            return True

        except Exception as e:
            self.log_message(f"Error loading master data: {str(e)}", 'error')
            return False

    def load_leave_data(self, file_obj, filename):
        """Load and process leave data file -"""
        try:
            self.log_message("Loading leave data...")
            self.leave_records = []

            # Check for multi-sheet workbook (new format with Master + Jan-Dec sheets)
            xlsx = pd.ExcelFile(file_obj)
            sheet_names = xlsx.sheet_names
            month_sheets = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

            has_month_sheets = any(month in sheet_names for month in month_sheets)

            if has_month_sheets:
                # New multi-sheet format (Master + monthly sheets)
                self.log_message("Detected multi-sheet format (monthly sheets)")

                # Load all month sheets for web version (no month prompt available)
                for sheet_name in sheet_names:
                    if sheet_name in month_sheets:
                        self.log_message(f"Processing sheet: {sheet_name}")
                        df = pd.read_excel(xlsx, sheet_name=sheet_name)
                        self.parse_leave_matrix(df)

                self.log_message(f"Loaded {len(self.leave_records)} leave records from monthly sheets", 'success')
            else:
                # Single sheet format
                file_obj.seek(0)
                df = read_excel_file(file_obj, filename)

                # Validate file
                self.validate_leave_sheet(df)

                # Detect format (matrix vs vertical)
                has_date_columns = any(
                    isinstance(col, datetime) or
                    hasattr(col, 'date') or
                    ('-' in str(col) and any(month in str(col).lower() for month in
                                             ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                                              'jul', 'aug', 'sep', 'oct', 'nov', 'dec']))
                    for col in df.columns
                )

                if has_date_columns:
                    self.log_message("Detected matrix format (dates as columns)")
                    self.parse_leave_matrix(df)
                else:
                    self.log_message("Detected vertical format")
                    self.parse_leave_vertical(df)

                self.log_message(f"Loaded {len(self.leave_records)} leave records", 'success')

            return True

        except Exception as e:
            self.log_message(f"Error loading leave data: {str(e)}", 'error')
            return False

    def parse_leave_matrix(self, df):
        """Parse leave data in matrix format - optimized for performance.
       ."""
        current_year = datetime.now().year

        crm_col = self.find_column(df, ['crm'], ['CRM'])

        if not crm_col:
            self.log_message("CRM column not found in leave sheet", 'warning')
            return

        month_map = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }

        # Pre-identify date columns for efficiency
        date_columns = []
        for col in df.columns:
            leave_date = None

            if isinstance(col, datetime):
                leave_date = col
            elif hasattr(col, 'date'):
                leave_date = datetime(col.year, col.month, col.day)
            else:
                col_str = str(col)
                if '-' in col_str:
                    parts = col_str.split('-')
                    if len(parts) == 2:
                        try:
                            day = int(parts[0])
                            month_str = parts[1].lower()[:3]
                            if month_str in month_map:
                                month = month_map[month_str]
                                leave_date = datetime(current_year, month, day)
                        except Exception:
                            pass

            if leave_date:
                date_columns.append((col, leave_date))

        if not date_columns:
            self.log_message("No date columns found in leave sheet", 'warning')
            return

        # Process rows efficiently
        crm_values = df[crm_col].astype(str).str.strip()

        for idx, crm in enumerate(crm_values):
            if not crm or crm == 'nan' or crm.lower() == 'crm':
                continue

            row = df.iloc[idx]

            for col, leave_date in date_columns:
                try:
                    leave_type = row[col]
                    if pd.notna(leave_type):
                        leave_str = str(leave_type).strip()
                        if leave_str:
                            self.leave_records.append({
                                'crm': crm,
                                'date': leave_date,
                                'leave_type': leave_str
                            })
                except Exception:
                    pass

    def parse_leave_vertical(self, df):
        """Parse leave data in vertical format -"""
        crm_col = self.find_column(df, ['crm', 'employee'], ['CRM', 'Employee CRM'])
        date_col = self.find_column(df, ['date'], ['Date'])
        type_col = self.find_column(df, ['type', 'leave'], ['Leave Type', 'Type'])

        if not all([crm_col, date_col]):
            self.log_message("Required columns not found in leave sheet", 'warning')
            return

        for idx, row in df.iterrows():
            crm = str(row[crm_col]).strip()
            if not crm or crm == 'nan':
                continue

            try:
                leave_date = pd.to_datetime(row[date_col])
                leave_type = str(row[type_col]).strip() if type_col and pd.notna(row[type_col]) else "Leave"

                self.leave_records.append({
                    'crm': crm,
                    'date': leave_date,
                    'leave_type': leave_type
                })
            except Exception:
                continue

    def get_filtered_employee_mapping(self, selected_depts=None, selected_crms=None):
        """Return employee mapping filtered by selected departments and CRMs"""
        if not selected_depts and not selected_crms:
            return self.employee_mapping

        all_depts = set(info['department'] for info in self.employee_mapping.values() if info['department'])
        all_crms = set(info['crm'] for info in self.employee_mapping.values() if info['crm'])

        selected_depts = set(selected_depts) if selected_depts else all_depts
        selected_crms = set(selected_crms) if selected_crms else all_crms

        filtered = {}
        for ac_no, info in self.employee_mapping.items():
            dept = info.get('department', '')
            crm = info.get('crm', '')

            dept_match = not dept or dept in selected_depts
            crm_match = not crm or crm in selected_crms

            if dept_match and crm_match:
                filtered[ac_no] = info

        return filtered

    def extract_date_from_filename(self, filename):
        """Extract date from attendance filename -"""
        try:
            # Strategy 1: Try to extract full date (YYYY-MM-DD format)
            date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', filename)
            if date_match:
                year, month, day = date_match.groups()
                return datetime(int(year), int(month), int(day))

            # Strategy 2: Try DD-MM-YYYY or DD_MM_YYYY
            date_match = re.search(r'(\d{1,2})[-_](\d{1,2})[-_](\d{4})', filename)
            if date_match:
                day, month, year = date_match.groups()
                return datetime(int(year), int(month), int(day))

            # Strategy 3: Try YYYYMMDD
            date_match = re.search(r'(\d{8})', filename)
            if date_match:
                try:
                    return datetime.strptime(date_match.group(1), '%Y%m%d')
                except ValueError:
                    pass

            # Strategy 4: Extract day only, use current month/year
            day_match = re.search(r'(\d{1,2})_report', filename)
            if day_match:
                day = int(day_match.group(1))
                now = datetime.now()
                return datetime(now.year, now.month, day)

        except Exception:
            pass

        return datetime.now()

    def get_first_clock_in(self, row):
        """Get first clock in time from row"""
        for i in range(1, 6):
            col_name = f'Clock In {i}'
            if col_name in row.index and pd.notna(row[col_name]):
                return row[col_name]
        return None

    def get_last_clock_out(self, row):
        """Get last clock out time from row"""
        for i in range(5, 0, -1):
            col_name = f'Clock Out {i}'
            if col_name in row.index and pd.notna(row[col_name]):
                return row[col_name]
        return None

    def determine_status(self, clock_in, clock_out, date, crm):
        """Determine attendance status -"""
        day_of_week = date.weekday()

        # Check if employee has resigned (before all other checks)
        for ac_no, info in self.employee_mapping.items():
            if info.get('crm') == crm:
                exit_date_str = info.get('exit_date', '')
                if exit_date_str:
                    try:
                        exit_date = datetime.strptime(exit_date_str, '%Y-%m-%d').date()
                        if date.date() > exit_date:
                            return "Departed", "Departed", "Departed"
                    except Exception:
                        pass
                break

        # Check if it's an OFF day FIRST (before leaves - ignore leaves on OFF days)
        off_days = self.config.get('off_days', [4])
        if day_of_week in off_days:
            if not clock_in and not clock_out:
                return "Weekend", "Weekend", "Weekend"
            else:
                return "Worked on Day Off", "Worked", "Worked"

        # Check for leave (only on working days, not OFF days)
        for leave in self.leave_records:
            if leave['crm'] == crm and leave['date'].date() == date.date():
                return leave['leave_type'], "On Leave", "On Leave"

        # Check for absent
        if not clock_in and not clock_out:
            return "Absent", "No Clock In", "No Clock Out"

        # Check for missing punch - differentiate between In and Out
        if not clock_in:
            return "Missing Punch In", "No Clock In", "..."

        if not clock_out:
            try:
                if isinstance(clock_in, str):
                    clock_in_time = datetime.strptime(clock_in, "%I:%M:%S %p").time()
                else:
                    clock_in_time = clock_in

                late_threshold_str = self.config.get('late_threshold', '12:00 PM')
                late_threshold = datetime.strptime(late_threshold_str, "%I:%M %p").time()
                in_status = "Late" if clock_in_time > late_threshold else "On Time"
            except Exception:
                in_status = "..."

            return "Missing Punch Out", in_status, "No Clock Out"

        # Determine if late
        try:
            if isinstance(clock_in, str):
                clock_in_time = datetime.strptime(clock_in, "%I:%M:%S %p").time()
            else:
                clock_in_time = clock_in

            late_threshold_str = self.config.get('late_threshold', '12:00 PM')
            late_threshold = datetime.strptime(late_threshold_str, "%I:%M %p").time()
            is_late = clock_in_time > late_threshold

            if is_late:
                return "Late", "Late", "On Time"
            else:
                return "Normal", "On Time", "On Time"
        except Exception:
            return "Normal", "On Time", "On Time"

    def process_attendance_files(self, files, selected_depts=None, selected_crms=None):
        """Process all attendance files -"""
        self.processed_data = []
        filtered_mapping = self.get_filtered_employee_mapping(selected_depts, selected_crms)
        skipped_files = 0

        for file_obj, filename in files:
            try:
                self.log_message(f"Processing: {filename}")
                file_obj.seek(0)
                df = read_excel_file(file_obj, filename)

                # Validate file
                self.validate_attendance_file(df, filename)

                file_date = self.extract_date_from_filename(filename)

                ac_col = self.find_column(df, ['ac', 'no'], ['AC-No.', 'Ac-No.', 'AC No'])
                name_col = self.find_column(df, ['name'], ['Name'])
                date_col = self.find_column(df, ['date'], ['Date'])

                for idx, row in df.iterrows():
                    ac_no = self.normalize_id(row[ac_col]) if ac_col and pd.notna(row[ac_col]) else ""
                    if not ac_no or ac_no == 'nan':
                        continue

                    record_date = file_date
                    if date_col and pd.notna(row[date_col]):
                        try:
                            record_date = pd.to_datetime(row[date_col])
                            if hasattr(record_date, 'to_pydatetime'):
                                record_date = record_date.to_pydatetime()
                        except Exception:
                            record_date = file_date

                    clock_in = self.get_first_clock_in(row)
                    clock_out = self.get_last_clock_out(row)

                    # Get employee info - try multiple matching strategies
                    emp_info = filtered_mapping.get(ac_no)

                    # Strategy 2: Match by name if AC-No. didn't match
                    if not emp_info and name_col and pd.notna(row[name_col]):
                        row_name = str(row[name_col]).strip().lower()
                        for key, info in filtered_mapping.items():
                            if info['name'].lower() == row_name:
                                emp_info = info
                                break

                    if emp_info:
                        crm = emp_info['crm']
                        name = emp_info['name']
                        dept = emp_info['department']
                        position = emp_info['position']
                    else:
                        # Skip employees not in filtered mapping when filters active
                        if selected_depts or selected_crms:
                            continue
                        name = str(row[name_col]).strip() if name_col and pd.notna(row[name_col]) else "Unknown"
                        crm = name if name != "Unknown" else f"ID-{ac_no}"
                        dept = ""
                        position = ""

                    status, in_status, out_status = self.determine_status(clock_in, clock_out, record_date, crm)

                    self.processed_data.append({
                        'ac_no': ac_no,
                        'crm': crm,
                        'name': name,
                        'department': dept,
                        'position': position,
                        'date': record_date,
                        'day': record_date.strftime('%A'),
                        'clock_in': clock_in,
                        'clock_out': clock_out,
                        'in_status': in_status,
                        'out_status': out_status,
                        'status': status
                    })

            except Exception as e:
                self.log_message(f"Skipped {filename} due to error: {str(e)}", 'warning')
                skipped_files += 1

        if skipped_files > 0:
            self.log_message(f"Skipped {skipped_files} file(s) due to errors", 'warning')

        self.log_message(f"Total records processed: {len(self.processed_data)}", 'success')

    def fill_leave_records(self, selected_depts=None, selected_crms=None):
        """Apply leave records to attendance data -"""
        self.log_message("Applying leave records to attendance data...")
        off_days = self.config.get('off_days', [4])

        if not self.processed_data:
            self.log_message("No processed data, skipping leave fill")
            return

        all_attendance_dates = [record['date'] for record in self.processed_data]
        min_date = min(all_attendance_dates).date()
        max_date = max(all_attendance_dates).date()
        self.log_message(f"Attendance date range: {min_date} to {max_date}")

        filtered_mapping = self.get_filtered_employee_mapping(selected_depts, selected_crms)
        valid_crms = set(info['crm'] for info in filtered_mapping.values())
        self.log_message(f"Valid CRMs from filtered mapping: {len(valid_crms)}")

        record_index = {}
        for idx, record in enumerate(self.processed_data):
            key = (record['crm'], record['date'].date())
            record_index[key] = idx

        crm_to_emp_info = {}
        for ac_no, info in filtered_mapping.items():
            crm_to_emp_info[info['crm']] = info

        self.conflict_records = []
        added_count = 0
        updated_count = 0
        skipped_friday = 0
        skipped_out_of_range = 0
        skipped_no_master = 0

        leave_lookup = {}
        for leave in self.leave_records:
            crm = leave['crm']
            leave_date = leave['date'].date()

            if crm not in valid_crms:
                skipped_no_master += 1
                continue

            if leave_date < min_date or leave_date > max_date:
                skipped_out_of_range += 1
                continue

            if leave['date'].weekday() in off_days:
                skipped_friday += 1
                continue

            key = (crm, leave_date)
            leave_lookup[key] = leave

        self.log_message(f"Leaves to apply: {len(leave_lookup)}")

        for key, leave in leave_lookup.items():
            crm, leave_date = key

            if key in record_index:
                idx = record_index[key]
                existing_record = self.processed_data[idx]
                existing_status = existing_record['status']

                # Track conflict BEFORE updating
                if existing_status not in ['Weekend', 'Departed', 'Absent']:
                    self.conflict_records.append({
                        'crm': crm,
                        'name': existing_record.get('name', ''),
                        'date': leave['date'],
                        'attendance_status': existing_status,
                        'leave_type': leave['leave_type'],
                        'clock_in': existing_record.get('clock_in'),
                        'clock_out': existing_record.get('clock_out')
                    })

                if existing_status in ['Absent', 'Missing Punch In', 'Missing Punch Out', 'Late', 'Normal']:
                    self.processed_data[idx]['status'] = leave['leave_type']
                    self.processed_data[idx]['in_status'] = 'On Leave'
                    self.processed_data[idx]['out_status'] = 'On Leave'
                    updated_count += 1
            else:
                emp_info = crm_to_emp_info.get(crm)
                if emp_info:
                    self.processed_data.append({
                        'ac_no': '',
                        'crm': crm,
                        'name': emp_info['name'],
                        'department': emp_info['department'],
                        'position': emp_info['position'],
                        'date': leave['date'],
                        'day': leave['date'].strftime('%A'),
                        'clock_in': None,
                        'clock_out': None,
                        'in_status': 'On Leave',
                        'out_status': 'On Leave',
                        'status': leave['leave_type']
                    })
                    added_count += 1

        self.log_message(f"Updated {updated_count} existing records with leave status")
        self.log_message(f"Added {added_count} new leave-only records")
        self.log_message(f"Skipped: {skipped_friday} Fridays, {skipped_out_of_range} out-of-range, {skipped_no_master} not in master")
        if self.conflict_records:
            self.log_message(f"Detected {len(self.conflict_records)} leave vs attendance conflicts", 'warning')
        self.log_message(f"Total records: {len(self.processed_data)}")

    def calculate_penalties(self):
        """Calculate penalties per employee based on attendance policy -"""
        self.log_message("Calculating penalties...")
        penalties_config = self.config.get('penalties', {})
        currency = penalties_config.get('currency', 'EGP')

        employee_stats = {}
        for record in self.processed_data:
            crm = record['crm']
            status = record['status']

            if crm not in employee_stats:
                national_id = ''
                vendor = ''
                ps_id = ''
                join_date = ''
                for ac_no, emp_info in self.employee_mapping.items():
                    if emp_info.get('crm') == crm:
                        national_id = emp_info.get('national_id', '')
                        vendor = emp_info.get('vendor', '')
                        ps_id = emp_info.get('ps_id', '')
                        join_date = emp_info.get('join_date', '')
                        break

                employee_stats[crm] = {
                    'name': record.get('name', ''),
                    'department': record.get('department', ''),
                    'position': record.get('position', ''),
                    'national_id': national_id,
                    'vendor': vendor,
                    'ps_id': ps_id,
                    'join_date': join_date,
                    'late_count': 0,
                    'missing_punch_count': 0,
                    'absence_count': 0,
                    'early_departure_count': 0,
                    'total_days': 0,
                    'working_days': 0
                }

            employee_stats[crm]['total_days'] += 1

            if status == 'Late':
                employee_stats[crm]['late_count'] += 1
                employee_stats[crm]['working_days'] += 1
            elif 'Missing Punch' in status:
                employee_stats[crm]['missing_punch_count'] += 1
                employee_stats[crm]['working_days'] += 1
            elif status == 'Absent':
                employee_stats[crm]['absence_count'] += 1
            elif status in ['Normal', 'Present']:
                employee_stats[crm]['working_days'] += 1

        penalties_summary = {}
        for crm, stats in employee_stats.items():
            late_count = stats['late_count']
            missing_count = stats['missing_punch_count']
            absence_count = stats['absence_count']

            # Calculate late penalties (cumulative: 1st=100, 2nd=200, 3rd=500, 4th+=500)
            late_penalty = 0
            late_warnings = 0
            for i in range(late_count):
                if i == 0:
                    late_penalty += penalties_config.get('late_1st', 100)
                elif i == 1:
                    late_penalty += penalties_config.get('late_2nd', 200)
                elif i == 2:
                    late_penalty += penalties_config.get('late_3rd', 500)
                    late_warnings += 1
                else:
                    late_penalty += penalties_config.get('late_4th_plus', 500)

            # Calculate missing punch deductions
            missing_threshold = penalties_config.get('missing_punch_threshold', 3)
            missing_warning_threshold = penalties_config.get('missing_punch_warning_threshold', 6)
            missing_deduction_rate = penalties_config.get('missing_punch_deduction', 0.5)

            missing_deduction = 0
            missing_warnings = 0
            if missing_count > missing_threshold:
                missing_deduction = (missing_count - missing_threshold) * missing_deduction_rate
            if missing_count > missing_warning_threshold:
                missing_warnings = 1

            # Calculate absence deductions (2 days per absence)
            absence_deduction_rate = penalties_config.get('absence_deduction', 2)
            absence_deduction = absence_count * absence_deduction_rate
            absence_warnings = 1 if absence_count > 0 else 0

            # Total calculations
            total_penalty_egp = late_penalty
            total_deduction_days = missing_deduction + absence_deduction
            total_warnings = late_warnings + missing_warnings + absence_warnings

            penalties_summary[crm] = {
                'name': stats['name'],
                'department': stats.get('department', ''),
                'position': stats.get('position', ''),
                'national_id': stats.get('national_id', ''),
                'vendor': stats.get('vendor', ''),
                'ps_id': stats.get('ps_id', ''),
                'join_date': stats.get('join_date', ''),
                'late_count': late_count,
                'late_penalty': late_penalty,
                'missing_punch_count': missing_count,
                'missing_deduction': missing_deduction,
                'absence_count': absence_count,
                'absence_deduction': absence_deduction,
                'total_penalty_egp': total_penalty_egp,
                'total_deduction_days': total_deduction_days,
                'total_warnings': total_warnings,
                'working_days': stats['working_days'],
                'total_days': stats['total_days']
            }

        self.penalties_data = penalties_summary
        self.log_message(f"Calculated penalties for {len(penalties_summary)} employees", 'success')
        return penalties_summary

    def create_excel_report(self):
        """Create the final Excel report -"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self.calculate_penalties()
        self.create_summary_sheet(wb)
        self.create_analytics_sheet(wb)
        self.create_alerts_sheet(wb)
        self.create_penalties_sheet(wb)
        self.create_refund_sheet(wb)
        self.create_duplicates_sheet(wb)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _apply_status_color(self, cell, status):
        """Apply color coding -"""
        # Light Green - Refund leaves (Refund) - check first before other rules
        if '(Refund)' in status:
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        # Light Purple - Backdated leaves (BD)
        elif '(BD)' in status:
            cell.fill = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
        # Green - No deduction statuses
        elif status in ['Normal', 'Present', 'Late (Approved)', 'Annual Leave', 'Casual Leave',
                        'Marriage Leave', 'Paternity Leave', 'Maternity Leave', 'Bereavement Leave',
                        'Military Call Leave', 'Early Departure (Approved)']:
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        # Yellow - Late with deduction
        elif status == 'Late':
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        # Red - Absent
        elif status == 'Absent':
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        # Pink - Missing Punch types
        elif 'Missing Punch' in status:
            cell.fill = PatternFill(start_color='FFD9E6', end_color='FFD9E6', fill_type='solid')
        # Light Blue - Leave types with deduction
        elif status in ['Sick Leave', 'Unpaid Leave', 'Unpaid leave']:
            cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        # Orange - Half day / Early departure / Early Leave (HD)
        elif status in ['Early Departure', 'Half Day', 'Early Leave (HD)', 'Early Leave (HD) (BD)']:
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        # Gray - Weekend
        elif status == 'Weekend':
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        # Light Silver - Departed (post exit date)
        elif status == 'Departed':
            cell.fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
            cell.font = Font(color='808080', italic=True)
        # Light Silver - Not Yet Hired
        elif status == 'Not Yet Hired':
            cell.fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
            cell.font = Font(color='808080', italic=True)
        # Light green - Worked on Day Off
        elif status == 'Worked on Day Off':
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        # Default light blue for other leave types
        elif 'Leave' in status:
            cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

    def create_summary_sheet(self, wb):
        """Create summary sheet -"""
        ws = wb.create_sheet("Summary Report", 0)

        crms = sorted(set(r['crm'] for r in self.processed_data))
        attendance_dates = [r['date'] for r in self.processed_data]
        if attendance_dates:
            min_date = min(attendance_dates)
            max_date = max(attendance_dates)
            dates = []
            current = min_date
            while current <= max_date:
                dates.append(current)
                current = current + timedelta(days=1)
        else:
            dates = sorted(set(r['date'] for r in self.processed_data))

        # Build matrix
        matrix = {crm: {date: '' for date in dates} for crm in crms}
        for record in self.processed_data:
            matrix[record['crm']][record['date']] = record['status']

        # Title
        ws.merge_cells('A1:' + get_column_letter(len(dates) + 3) + '1')
        title_cell = ws['A1']
        title_cell.value = "Enhanced Attendance Summary Report"
        title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Headers
        row = 3
        ws.cell(row, 1, "CRM")
        ws.cell(row, 2, "Normal Days")
        ws.cell(row, 3, "Abnormal Days")

        for i, date in enumerate(dates, start=4):
            ws.cell(row, i, date.strftime("%d-%b"))

        for col in range(1, len(dates) + 4):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        off_days = self.config.get('off_days', [4])

        # Justification options
        justification_options = [
            "Normal", "Late (Approved)", "Late", "Absent", "Missing Punch In",
            "Missing Punch In (Justified)", "Missing Punch Out", "Missing Punch Out (Justified)",
            "Early Departure (Approved)", "Early Departure", "Half Day", "Early Leave (HD)",
            "Sick Leave", "Annual Leave", "Casual Leave", "Marriage Leave", "Paternity Leave",
            "Maternity Leave", "Bereavement Leave", "Military Call Leave", "Unpaid Leave",
            "Weekend", "Departed",
            "Annual Leave (BD)", "Casual Leave (BD)", "Sick Leave (BD)", "Unpaid Leave (BD)",
            "Half Day (BD)", "Early Leave (HD) (BD)", "Early Departure (BD)", "Marriage Leave (BD)",
            "Paternity Leave (BD)", "Maternity Leave (BD)", "Bereavement Leave (BD)",
            # Refund variants - reversal of previous penalties
            "Annual Leave (Refund)", "Sick Leave (Refund)", "Half Day (Refund)"
        ]

        justification_list = ",".join(justification_options)
        dv = DataValidation(
            type="list",
            formula1=f'"{justification_list}"',
            allow_blank=True,
            showDropDown=False
        )
        dv.error = "Please select a valid justification from the list"
        dv.errorTitle = "Invalid Entry"
        dv.prompt = "Select attendance status/justification"
        dv.promptTitle = "Attendance Status"
        ws.add_data_validation(dv)

        # Build CRM → join_date lookup for "Not Yet Hired" detection
        crm_join_dates = {}
        for ac_no, emp_info in self.employee_mapping.items():
            crm_val = emp_info.get('crm', '')
            jd = emp_info.get('join_date', '')
            if crm_val and jd:
                try:
                    crm_join_dates[crm_val] = datetime.strptime(jd, '%Y-%m-%d').date() if isinstance(jd, str) else jd.date() if hasattr(jd, 'date') else jd
                except (ValueError, TypeError):
                    pass

        # Build CRM → exit_date lookup for "Departed" detection
        crm_exit_dates = {}
        for ac_no, emp_info in self.employee_mapping.items():
            crm_val = emp_info.get('crm', '')
            ed = emp_info.get('exit_date', '')
            if crm_val and ed:
                try:
                    crm_exit_dates[crm_val] = datetime.strptime(ed, '%Y-%m-%d').date() if isinstance(ed, str) else ed.date() if hasattr(ed, 'date') else ed
                except (ValueError, TypeError):
                    pass

        # Data rows
        row = 4
        first_data_row = row
        for crm in crms:
            ws.cell(row, 1, crm)

            # Count normal days - EXACT SAME LOGIC as desktop
            normal_statuses = ['Normal', 'Present', 'Weekend', 'Worked on Day Off', 'Late (Approved)',
                               'Annual Leave', 'Casual Leave', 'Marriage Leave', 'Paternity Leave',
                               'Maternity Leave', 'Bereavement Leave', 'Military Call Leave',
                               'Early Departure (Approved)',
                               'Annual Leave (Refund)', 'Sick Leave (Refund)', 'Half Day (Refund)',
                               'Not Yet Hired', 'Departed']
            normal_count = 0
            for d in dates:
                # Pre-hire dates count as normal (no penalty)
                join_dt = crm_join_dates.get(crm)
                exit_dt = crm_exit_dates.get(crm)
                d_date = d.date() if hasattr(d, 'date') else d
                if join_dt and d_date < join_dt:
                    normal_count += 1
                elif d.weekday() in off_days:
                    normal_count += 1
                # Post-exit dates count as normal (no penalty) — but weekends stay as Weekend
                elif exit_dt and d_date > exit_dt:
                    normal_count += 1
                else:
                    status = matrix[crm].get(d, '')
                    if status in normal_statuses:
                        normal_count += 1
            ws.cell(row, 2, normal_count)
            ws.cell(row, 3, len(dates) - normal_count)

            for i, date in enumerate(dates, start=4):
                status = matrix[crm].get(date, '')
                day_of_week = date.weekday()

                # Check if date is before employee's join date
                join_dt = crm_join_dates.get(crm)
                exit_dt = crm_exit_dates.get(crm)
                date_val = date.date() if hasattr(date, 'date') else date
                if join_dt and date_val < join_dt:
                    status = "Not Yet Hired"
                # Friday always Weekend
                elif day_of_week in off_days:
                    status = "Weekend"
                # Post-exit dates show as "Departed"
                elif exit_dt and date_val > exit_dt:
                    status = "Departed"
                elif not status:
                    status = "Absent"

                # Map old values
                status_mapping = {'Present': 'Normal'}
                status = status_mapping.get(status, status)

                cell = ws.cell(row, i, status)
                self._apply_status_color(cell, status)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            row += 1

        last_data_row = row - 1

        # Apply data validation
        if dates and crms:
            first_date_col = get_column_letter(4)
            last_date_col = get_column_letter(len(dates) + 3)
            dv.add(f"{first_date_col}{first_data_row}:{last_date_col}{last_data_row}")

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        for i in range(4, len(dates) + 4):
            ws.column_dimensions[get_column_letter(i)].width = 18

    def create_analytics_sheet(self, wb):
        """Create analytics sheet -"""
        ws = wb.create_sheet("Individual Analytics", 1)

        analytics = {}
        for crm in set(r['crm'] for r in self.processed_data):
            records = [r for r in self.processed_data if r['crm'] == crm]
            total = len(records)
            normal = sum(1 for r in records if r['status'] == 'Normal')
            late = sum(1 for r in records if r['status'] == 'Late')
            absent = sum(1 for r in records if r['status'] == 'Absent')
            missing = sum(1 for r in records if 'Missing Punch' in r['status'])

            analytics[crm] = {
                'name': records[0]['name'],
                'dept': records[0].get('department', ''),
                'position': records[0].get('position', ''),
                'total': total,
                'normal': normal,
                'late': late,
                'absent': absent,
                'missing': missing,
                'attendance_rate': round((total - absent) / total * 100, 1) if total > 0 else 0,
                'punctuality_rate': round(normal / total * 100, 1) if total > 0 else 0
            }

        # Title
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = "Individual Employee Analytics"
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Headers
        headers = ['CRM', 'Name', 'Department', 'Position', 'Total Days', 'Normal Days',
                   'Late Days', 'Absent Days', 'Missing Punch', 'Attendance Rate %', 'Punctuality Rate %']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        row = 4
        for crm, data in sorted(analytics.items()):
            ws.cell(row, 1, crm)
            ws.cell(row, 2, data['name'])
            ws.cell(row, 3, data['dept'])
            ws.cell(row, 4, data['position'])
            ws.cell(row, 5, data['total'])
            ws.cell(row, 6, data['normal'])
            ws.cell(row, 7, data['late'])
            ws.cell(row, 8, data['absent'])
            ws.cell(row, 9, data['missing'])
            ws.cell(row, 10, f"{data['attendance_rate']}%")
            ws.cell(row, 11, f"{data['punctuality_rate']}%")
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        for col in range(5, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def create_alerts_sheet(self, wb):
        """Create alerts sheet -"""
        ws = wb.create_sheet("Alerts & Warnings", 2)

        # Get thresholds
        thresholds = self.config.get('alert_thresholds', {})
        high_absences = thresholds.get('high_absences', 3)
        frequent_late = thresholds.get('frequent_late', 5)
        missing_punches = thresholds.get('missing_punches', 5)
        low_rate = thresholds.get('low_attendance_rate', 75.0)

        # Calculate analytics for alerts
        analytics = {}
        for crm in set(r['crm'] for r in self.processed_data):
            records = [r for r in self.processed_data if r['crm'] == crm]
            total = len(records)
            late = sum(1 for r in records if r['status'] == 'Late')
            absent = sum(1 for r in records if r['status'] == 'Absent')
            missing = sum(1 for r in records if 'Missing Punch' in r['status'])
            attendance_rate = (total - absent) / total * 100 if total > 0 else 0

            analytics[crm] = {
                'name': records[0]['name'],
                'late': late,
                'absent': absent,
                'missing': missing,
                'rate': attendance_rate
            }

        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "Attendance Alerts & Warnings"
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Headers
        headers = ['CRM', 'Name', 'Alert Type', 'Details']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Generate alerts
        row = 4
        alert_count = 0

        for crm, data in sorted(analytics.items()):
            if data['absent'] >= high_absences:
                ws.cell(row, 1, crm)
                ws.cell(row, 2, data['name'])
                ws.cell(row, 3, "High Absences")
                ws.cell(row, 4, f"{data['absent']} absences in period")
                for col in range(1, 5):
                    ws.cell(row, col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                row += 1
                alert_count += 1

            if data['late'] >= frequent_late:
                ws.cell(row, 1, crm)
                ws.cell(row, 2, data['name'])
                ws.cell(row, 3, "Frequent Late Arrivals")
                ws.cell(row, 4, f"{data['late']} late arrivals in period")
                for col in range(1, 5):
                    ws.cell(row, col).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                row += 1
                alert_count += 1

            if data['missing'] >= missing_punches:
                ws.cell(row, 1, crm)
                ws.cell(row, 2, data['name'])
                ws.cell(row, 3, "Missing Punches")
                ws.cell(row, 4, f"{data['missing']} missing punches (possible device issue)")
                for col in range(1, 5):
                    ws.cell(row, col).fill = PatternFill(start_color='FFD9E6', end_color='FFD9E6', fill_type='solid')
                row += 1
                alert_count += 1

            if data['rate'] < low_rate:
                ws.cell(row, 1, crm)
                ws.cell(row, 2, data['name'])
                ws.cell(row, 3, "Low Attendance Rate")
                ws.cell(row, 4, f"Attendance rate: {data['rate']:.1f}%")
                for col in range(1, 5):
                    ws.cell(row, col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                row += 1
                alert_count += 1

        if alert_count == 0:
            ws.merge_cells('A5:D5')
            cell = ws['A5']
            cell.value = "No attendance issues detected!"
            cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40

    def create_penalties_sheet(self, wb):
        """Create penalties sheet - (25 columns with formulas)"""
        ws = wb.create_sheet("Penalties", 3)
        currency = self.config.get('penalties', {}).get('currency', 'EGP')
        penalties_config = self.config.get('penalties', {})

        # Get summary sheet info for formula references
        summary_ws = wb['Summary Report']
        summary_last_col = summary_ws.max_column
        summary_last_col_letter = get_column_letter(summary_last_col)

        total_cols = 26

        # Title
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        title_cell = ws['A1']
        title_cell.value = f"Attendance Penalties Report ({currency})"
        title_cell.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
        subtitle = ws['A2']
        subtitle.value = "Based on Attendance and Discipline Policy 2026 - Section 7 (Linked to Summary Report)"
        subtitle.font = Font(size=9, italic=True)
        subtitle.alignment = Alignment(horizontal='center')

        # Headers (26 columns)
        headers = [
            'CRM', 'Name', 'National ID', 'Vendor', 'PS ID', 'Department', 'Join Date',
            'Late Count', f'Late Penalty ({currency})',
            'Missing Punches', 'Punch Ded. (-)',
            'Absences', 'Absence Ded. (-)',
            'Early Dep.', 'Early Dep. Ded. (-)',
            'Half Day', 'Half Day Ded. (-)',
            'Sick Leave', 'Sick Ded. (-)',
            'Unpaid Leave', 'Unpaid Ded. (-)',
            f'Total Penalty ({currency})', 'Gross Ded. (days)', 'Net Ded. (days)',
            'Warnings', 'Backdated Leaves'
        ]

        header_row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill(start_color='495057', end_color='495057', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        ws.row_dimensions[header_row].height = 40

        # Get penalty config values
        late_1st = penalties_config.get('late_1st', 100)
        late_2nd = penalties_config.get('late_2nd', 200)
        late_3rd = penalties_config.get('late_3rd', 500)
        late_4th_plus = penalties_config.get('late_4th_plus', 500)
        missing_threshold = penalties_config.get('missing_punch_threshold', 3)
        missing_deduction_rate = penalties_config.get('missing_punch_deduction', 0.5)
        absence_deduction_rate = penalties_config.get('absence_deduction', 2)
        early_dep_deduction = 0.5
        half_day_deduction = 0.5
        sick_leave_deduction = 0.25
        unpaid_leave_deduction = 1.0

        # Determine attendance period end date (for tenure check)
        period_end = max((r['date'] for r in self.processed_data), default=datetime.now())
        if hasattr(period_end, 'date'):
            period_end = period_end.date()

        # Build CRM to summary row mapping
        crm_to_summary_row = {}
        for r in range(4, summary_ws.max_row + 1):
            crm_val = summary_ws.cell(r, 1).value
            if crm_val:
                crm_to_summary_row[crm_val] = r

        data_row = 5
        for crm, data in sorted(self.penalties_data.items()):
            summary_row = crm_to_summary_row.get(crm, None)
            sr = summary_row
            rng = f"'Summary Report'!$D${sr}:${summary_last_col_letter}${sr}" if sr else None

            # Column A: CRM
            ws.cell(data_row, 1, crm)

            # Columns B-G: Static employee info
            ws.cell(data_row, 2, data['name'])
            ws.cell(data_row, 3, data.get('national_id', ''))
            ws.cell(data_row, 4, data.get('vendor', ''))
            ws.cell(data_row, 5, data.get('ps_id', ''))
            ws.cell(data_row, 6, data.get('department', ''))
            ws.cell(data_row, 7, data.get('join_date', ''))

            # Column H: Late Count
            if sr:
                ws.cell(data_row, 8, f'=COUNTIF({rng},"Late")')
            else:
                ws.cell(data_row, 8, data['late_count'])

            # Column I: Late Penalty formula
            late_col = "H"
            penalty_formula = (
                f"=IF({late_col}{data_row}=0,0,"
                f"IF({late_col}{data_row}=1,{late_1st},"
                f"IF({late_col}{data_row}=2,{late_1st}+{late_2nd},"
                f"IF({late_col}{data_row}=3,{late_1st}+{late_2nd}+{late_3rd},"
                f"{late_1st}+{late_2nd}+{late_3rd}+({late_col}{data_row}-3)*{late_4th_plus}))))"
            )
            ws.cell(data_row, 9, penalty_formula)

            # Column J: Missing Punches
            if sr:
                missing_formula = (
                    f'=COUNTIF({rng},"Missing Punch In")'
                    f'+COUNTIF({rng},"Missing Punch In (Justified)")'
                    f'+COUNTIF({rng},"Missing Punch Out")'
                    f'+COUNTIF({rng},"Missing Punch Out (Justified)")'
                )
                ws.cell(data_row, 10, missing_formula)
            else:
                ws.cell(data_row, 10, data['missing_punch_count'])

            # Column K: Punch Deduction (negative)
            ws.cell(data_row, 11, f"=-IF(J{data_row}>{missing_threshold},(J{data_row}-{missing_threshold})*{missing_deduction_rate},0)")

            # Column L: Absences
            if sr:
                ws.cell(data_row, 12, f'=COUNTIF({rng},"Absent")')
            else:
                ws.cell(data_row, 12, data['absence_count'])

            # Column M: Absence Deduction (negative)
            ws.cell(data_row, 13, f"=-L{data_row}*{absence_deduction_rate}")

            # Column N: Early Departure Count
            if sr:
                ws.cell(data_row, 14, f'=COUNTIF({rng},"Early Departure")+COUNTIF({rng},"Early Departure (BD)")')
            else:
                ws.cell(data_row, 14, 0)

            # Column O: Early Departure Deduction (negative)
            ws.cell(data_row, 15, f"=-N{data_row}*{early_dep_deduction}")

            # Column P: Half Day Count
            if sr:
                ws.cell(data_row, 16, f'=COUNTIF({rng},"Half Day")+COUNTIF({rng},"Half Day (BD)")+COUNTIF({rng},"Early Leave (HD)")+COUNTIF({rng},"Early Leave (HD) (BD)")')
            else:
                ws.cell(data_row, 16, 0)

            # Column Q: Half Day Deduction (negative)
            # Employees hired > 6 months before period end are exempt from half day deduction
            join_date_str = data.get('join_date', '')
            half_day_exempt = False
            if join_date_str:
                try:
                    join_dt = datetime.strptime(join_date_str, '%Y-%m-%d').date() if isinstance(join_date_str, str) else join_date_str
                    tenure_days = (period_end - join_dt).days
                    half_day_exempt = tenure_days > 183  # ~6 months
                except (ValueError, TypeError):
                    pass
            if half_day_exempt:
                ws.cell(data_row, 17, 0)
            else:
                ws.cell(data_row, 17, f"=-P{data_row}*{half_day_deduction}")

            # Column R: Sick Leave Count
            if sr:
                ws.cell(data_row, 18, f'=COUNTIF({rng},"Sick Leave")+COUNTIF({rng},"Sick Leave (BD)")')
            else:
                ws.cell(data_row, 18, 0)

            # Column S: Sick Leave Deduction (negative)
            ws.cell(data_row, 19, f"=-R{data_row}*{sick_leave_deduction}")

            # Column T: Unpaid Leave Count
            if sr:
                ws.cell(data_row, 20, f'=COUNTIF({rng},"Unpaid Leave")+COUNTIF({rng},"Unpaid Leave (BD)")')
            else:
                ws.cell(data_row, 20, 0)

            # Column U: Unpaid Leave Deduction (negative)
            ws.cell(data_row, 21, f"=-T{data_row}*{unpaid_leave_deduction}")

            # Column V: Total Penalty (only late penalty contributes to EGP)
            ws.cell(data_row, 22, f"=I{data_row}")

            # Column W: Gross Deduction (sum of all negative day deductions)
            ws.cell(data_row, 23, f"=K{data_row}+M{data_row}+O{data_row}+Q{data_row}+S{data_row}+U{data_row}")

            # Column X: Net Deduction (gross deductions only, refunds shown separately)
            ws.cell(data_row, 24, f"=W{data_row}")

            # Column Y: Warnings
            warnings_formula = (
                f"=IF(H{data_row}>=3,1,0)+"
                f"IF(J{data_row}>6,1,0)+"
                f"IF(L{data_row}>0,1,0)"
            )
            ws.cell(data_row, 25, warnings_formula)

            # Column Z: Backdated Leaves
            if sr:
                ws.cell(data_row, 26, f'=COUNTIF({rng},"*(BD)")')
            else:
                ws.cell(data_row, 26, 0)

            # Apply borders
            for col in range(1, total_cols + 1):
                cell = ws.cell(data_row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            data_row += 1

        # Totals row
        totals_row = data_row + 1
        first_data_row = 5
        last_data_row = data_row - 1

        ws.cell(totals_row, 1, "TOTAL").font = Font(bold=True)

        sum_columns = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26]
        for col in sum_columns:
            col_letter = get_column_letter(col)
            ws.cell(totals_row, col, f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")

        # Highlight totals
        ws.cell(totals_row, 22).fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        ws.cell(totals_row, 22).font = Font(bold=True, color='FFFFFF', size=12)
        ws.cell(totals_row, 23).fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        ws.cell(totals_row, 23).font = Font(bold=True, color='FFFFFF', size=12)
        ws.cell(totals_row, 24).fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        ws.cell(totals_row, 24).font = Font(bold=True, color='FFFFFF', size=12)
        ws.cell(totals_row, 25).font = Font(bold=True)

        for col in range(1, total_cols + 1):
            ws.cell(totals_row, col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='double'), bottom=Side(style='thin')
            )

        # Column widths (26 columns)
        widths = [15, 18, 16, 12, 10, 15, 11, 8, 14, 10, 10, 8, 10, 8, 10, 8, 10, 8, 8, 10, 10, 14, 12, 12, 8, 12]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Apply highlighting to deduction columns
        for r in range(5, data_row):
            ws.cell(r, 9).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            ws.cell(r, 11).fill = PatternFill(start_color='FFD9E6', end_color='FFD9E6', fill_type='solid')
            ws.cell(r, 13).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws.cell(r, 15).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            ws.cell(r, 17).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            ws.cell(r, 19).fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            ws.cell(r, 21).fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            # Total Penalty
            ws.cell(r, 22).fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            ws.cell(r, 22).font = Font(bold=True)
            # Gross Deduction
            ws.cell(r, 23).fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            ws.cell(r, 23).font = Font(bold=True)
            # Net Deduction
            ws.cell(r, 24).fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            ws.cell(r, 24).font = Font(bold=True)
            # Warnings
            ws.cell(r, 25).fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            ws.cell(r, 25).font = Font(bold=True, color='856404')
            # Backdated Leaves
            ws.cell(r, 26).fill = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
            ws.cell(r, 26).font = Font(bold=True, color='6B3FA0')

        # Policy legend
        legend_row = totals_row + 3
        ws.cell(legend_row, 1, "Deduction Policy Reference:").font = Font(bold=True)
        legend_row += 1
        ws.cell(legend_row, 1, f"Late: {currency} {late_1st} (1st), {currency} {late_2nd} (2nd), {currency} {late_3rd}+ (3rd+) + Warning | Late (Approved): No Deduction")
        legend_row += 1
        ws.cell(legend_row, 1, f"Missing Punch (all types): {missing_deduction_rate} day after {missing_threshold} occurrences, warning after 6")
        legend_row += 1
        ws.cell(legend_row, 1, f"Absent: {absence_deduction_rate} days | Early Departure: {early_dep_deduction} day | Half Day: {half_day_deduction} day (exempt if hired >6 months) | Sick Leave: {sick_leave_deduction} day | Unpaid Leave: {unpaid_leave_deduction} day")
        legend_row += 1
        ws.cell(legend_row, 1, "No Deduction: Normal, Late (Approved), Early Departure (Approved), Annual/Casual/Marriage/Paternity/Maternity/Bereavement/Military Call Leave")
        legend_row += 1
        ws.cell(legend_row, 1, "REFUND LEAVES: Shown separately in the 'Refund' sheet - Annual Leave (Refund) +1 day, Sick Leave (Refund) +0.75 days, Half Day (Refund) +0.5 days")
        ws.cell(legend_row, 1).font = Font(bold=True, color='375623')
        legend_row += 1
        ws.cell(legend_row, 1, "SIGN CONVENTION: Deductions shown as negative (-). Net Ded. = Gross Ded. (refunds excluded - see Refund sheet)")
        legend_row += 1
        ws.cell(legend_row, 1, "BACKDATED LEAVES (BD): Leaves marked with (BD) suffix are transferred from previous months - shown in purple, HR to review manually")
        ws.cell(legend_row, 1).font = Font(bold=True, color='6B3FA0')
        legend_row += 1
        ws.cell(legend_row, 1, "Note: This sheet is linked to Summary Report - use the dropdown to change status and penalties will auto-update")
        ws.cell(legend_row, 1).font = Font(italic=True, color='0066CC')

    def create_refund_sheet(self, wb):
        """Create refund sheet - shows refund leave credits separately"""
        ws = wb.create_sheet("Refund", 4)
        currency = self.config.get('penalties', {}).get('currency', 'EGP')

        summary_ws = wb['Summary Report']
        summary_last_col = summary_ws.max_column
        summary_last_col_letter = get_column_letter(summary_last_col)

        total_cols = 14

        # Title
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        title_cell = ws['A1']
        title_cell.value = "Attendance Refund Report"
        title_cell.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
        subtitle = ws['A2']
        subtitle.value = "Refund credits for reversed deductions (linked to Summary Report)"
        subtitle.font = Font(size=9, italic=True)
        subtitle.alignment = Alignment(horizontal='center')

        # Headers
        headers = [
            'CRM', 'Name', 'National ID', 'Vendor', 'PS ID', 'Department', 'Join Date',
            'Annual Leave (Refund)', 'Annual Refund Days',
            'Sick Leave (Refund)', 'Sick Refund Days',
            'Half Day (Refund)', 'Half Day Refund Days',
            'Total Refund Days'
        ]

        header_row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        ws.row_dimensions[header_row].height = 40

        # Build CRM to summary row mapping
        crm_to_summary_row = {}
        for r in range(4, summary_ws.max_row + 1):
            crm_val = summary_ws.cell(r, 1).value
            if crm_val:
                crm_to_summary_row[crm_val] = r

        data_row = 5
        for crm, data in sorted(self.penalties_data.items()):
            summary_row = crm_to_summary_row.get(crm, None)
            sr = summary_row
            rng = f"'Summary Report'!$D${sr}:${summary_last_col_letter}${sr}" if sr else None

            # Columns A-G: Static employee info
            ws.cell(data_row, 1, crm)
            ws.cell(data_row, 2, data['name'])
            ws.cell(data_row, 3, data.get('national_id', ''))
            ws.cell(data_row, 4, data.get('vendor', ''))
            ws.cell(data_row, 5, data.get('ps_id', ''))
            ws.cell(data_row, 6, data.get('department', ''))
            ws.cell(data_row, 7, data.get('join_date', ''))

            # Column H: Annual Leave (Refund) Count
            if sr:
                ws.cell(data_row, 8, f'=COUNTIF({rng},"Annual Leave (Refund)")')
            else:
                ws.cell(data_row, 8, 0)

            # Column I: Annual Leave Refund Days (+1 per occurrence)
            ws.cell(data_row, 9, f"=H{data_row}*1")

            # Column J: Sick Leave (Refund) Count
            if sr:
                ws.cell(data_row, 10, f'=COUNTIF({rng},"Sick Leave (Refund)")')
            else:
                ws.cell(data_row, 10, 0)

            # Column K: Sick Leave Refund Days (+0.75 per occurrence)
            ws.cell(data_row, 11, f"=J{data_row}*0.75")

            # Column L: Half Day (Refund) Count
            if sr:
                ws.cell(data_row, 12, f'=COUNTIF({rng},"Half Day (Refund)")')
            else:
                ws.cell(data_row, 12, 0)

            # Column M: Half Day Refund Days (+0.5 per occurrence)
            ws.cell(data_row, 13, f"=L{data_row}*0.5")

            # Column N: Total Refund Days
            ws.cell(data_row, 14, f"=I{data_row}+K{data_row}+M{data_row}")

            # Apply borders and formatting
            for col in range(1, total_cols + 1):
                cell = ws.cell(data_row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            # Color refund count/days columns green
            for col in [8, 9, 10, 11, 12, 13, 14]:
                ws.cell(data_row, col).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            ws.cell(data_row, 9).font = Font(bold=True, color='375623')
            ws.cell(data_row, 11).font = Font(bold=True, color='375623')
            ws.cell(data_row, 13).font = Font(bold=True, color='375623')
            ws.cell(data_row, 14).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            ws.cell(data_row, 14).font = Font(bold=True, color='375623')

            data_row += 1

        # Totals row
        totals_row = data_row + 1
        first_data_row = 5
        last_data_row = data_row - 1

        ws.cell(totals_row, 1, "TOTAL").font = Font(bold=True)
        for col in [8, 9, 10, 11, 12, 13, 14]:
            col_letter = get_column_letter(col)
            ws.cell(totals_row, col, f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
            ws.cell(totals_row, col).fill = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
            ws.cell(totals_row, col).font = Font(bold=True, color='FFFFFF', size=12)
        for col in range(1, total_cols + 1):
            ws.cell(totals_row, col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='double'), bottom=Side(style='thin')
            )

        # Column widths
        widths = [15, 18, 16, 12, 10, 15, 11, 16, 16, 14, 14, 14, 16, 16]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Legend
        legend_row = totals_row + 3
        ws.cell(legend_row, 1, "Refund Policy Reference:").font = Font(bold=True, color='375623')
        legend_row += 1
        ws.cell(legend_row, 1, "Annual Leave (Refund): +1 day | Sick Leave (Refund): +0.75 days | Half Day (Refund): +0.5 days")
        legend_row += 1
        ws.cell(legend_row, 1, "Note: Refund leaves reverse previously applied deductions. This sheet is linked to Summary Report.")
        ws.cell(legend_row, 1).font = Font(italic=True, color='0066CC')

    def create_duplicates_sheet(self, wb):
        """Create duplicates sheet - shows leave vs attendance conflicts"""
        ws = wb.create_sheet("Duplicates", 5)

        ws.merge_cells('A1:G1')
        title = ws['A1']
        title.value = "Leave vs Attendance Conflicts"
        title.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        title.fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:G2')
        subtitle = ws['A2']
        subtitle.value = "Employees with both attendance and leave on the same date"
        subtitle.font = Font(size=9, italic=True)
        subtitle.alignment = Alignment(horizontal='center')

        headers = ['CRM', 'Name', 'Date', 'Original Attendance Status', 'Leave Type Applied', 'Clock In', 'Clock Out']
        header_row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(header_row, col, header)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        ws.row_dimensions[header_row].height = 25

        if self.conflict_records:
            data_row = 5
            for conflict in sorted(self.conflict_records, key=lambda x: (x['crm'], x['date'])):
                clock_in = conflict.get('clock_in')
                clock_out = conflict.get('clock_out')
                clock_in_str = str(clock_in) if clock_in else '-'
                clock_out_str = str(clock_out) if clock_out else '-'

                date_val = conflict['date']
                date_str = date_val.strftime('%d-%b-%Y') if hasattr(date_val, 'strftime') else str(date_val)

                ws.cell(data_row, 1, conflict['crm'])
                ws.cell(data_row, 2, conflict['name'])
                ws.cell(data_row, 3, date_str)
                ws.cell(data_row, 4, conflict['attendance_status'])
                ws.cell(data_row, 5, conflict['leave_type'])
                ws.cell(data_row, 6, clock_in_str)
                ws.cell(data_row, 7, clock_out_str)

                for col in range(1, 8):
                    cell = ws.cell(data_row, col)
                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                data_row += 1

            summary_row = data_row + 1
            ws.cell(summary_row, 1, f"Total Conflicts: {len(self.conflict_records)}")
            ws.cell(summary_row, 1).font = Font(bold=True)
        else:
            ws.merge_cells('A5:G5')
            no_conflict = ws['A5']
            no_conflict.value = "No conflicts detected - All leave records are consistent with attendance data"
            no_conflict.font = Font(size=11, color='28A745')
            no_conflict.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[5].height = 30

        for i, width in enumerate([15, 20, 12, 25, 20, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = width


# Streamlit App
def main():
    st.set_page_config(
        page_title="Attendance Dashboard",
        page_icon="📊",
        layout="wide"
    )

    if 'processor' not in st.session_state:
        st.session_state.processor = AttendanceProcessor()
    if 'master_loaded' not in st.session_state:
        st.session_state.master_loaded = False
    if 'attendance_loaded' not in st.session_state:
        st.session_state.attendance_loaded = False
    if 'leave_loaded' not in st.session_state:
        st.session_state.leave_loaded = False
    if 'attendance_files' not in st.session_state:
        st.session_state.attendance_files = []

    st.image("Logo.png", width=120)
    st.title("Attendance Dashboard v2.2")
    st.markdown("*Web-based attendance processing*")

    with st.sidebar:
        st.header("Settings")

        # Work Hours
        st.subheader("Work Hours")
        work_start = st.text_input("Start Time", value=st.session_state.processor.config.get('work_start_time', '12:00 PM'))
        work_end = st.text_input("End Time", value=st.session_state.processor.config.get('work_end_time', '9:00 PM'))

        # Generate 5-minute interval time options (12:00 AM to 11:55 PM)
        time_options = []
        for h in range(24):
            for m in range(0, 60, 5):
                t = datetime(2000, 1, 1, h, m).strftime("%I:%M %p")
                time_options.append(t)
        current_late = st.session_state.processor.config.get('late_threshold', '12:00 PM')
        late_index = time_options.index(current_late) if current_late in time_options else 0
        late_threshold_str = st.selectbox("Late Threshold", options=time_options, index=late_index)

        off_days = st.multiselect(
            "Off Days",
            options=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
            default=["Friday"]
        )

        day_map = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3,
                   "Friday": 4, "Saturday": 5, "Sunday": 6}
        off_day_nums = [day_map[d] for d in off_days]

        st.session_state.processor.config['work_start_time'] = work_start
        st.session_state.processor.config['work_end_time'] = work_end
        st.session_state.processor.config['late_threshold'] = late_threshold_str
        st.session_state.processor.config['off_days'] = off_day_nums

        # Alert Thresholds
        st.divider()
        st.subheader("Alert Thresholds")
        current_thresholds = st.session_state.processor.config.get('alert_thresholds', {})

        high_absences = st.number_input("High Absences (days)",
                                        value=current_thresholds.get('high_absences', 3),
                                        min_value=1, step=1)
        frequent_late = st.number_input("Frequent Late (count)",
                                        value=current_thresholds.get('frequent_late', 5),
                                        min_value=1, step=1)
        missing_punches_threshold = st.number_input("Missing Punches (count)",
                                                     value=current_thresholds.get('missing_punches', 5),
                                                     min_value=1, step=1)
        low_attendance_rate = st.number_input("Low Attendance Rate (%)",
                                              value=float(current_thresholds.get('low_attendance_rate', 75.0)),
                                              min_value=0.0, max_value=100.0, step=5.0)

        st.session_state.processor.config['alert_thresholds'] = {
            'high_absences': high_absences,
            'frequent_late': frequent_late,
            'missing_punches': missing_punches_threshold,
            'low_attendance_rate': low_attendance_rate
        }

        st.divider()

        if st.session_state.master_loaded:
            st.header("Filters")

            depts = sorted(set(info['department'] for info in st.session_state.processor.employee_mapping.values()
                               if info['department']))
            crms = sorted(set(info['crm'] for info in st.session_state.processor.employee_mapping.values()
                              if info['crm']))

            selected_depts = st.multiselect("Departments", options=depts, default=depts)
            selected_crms = st.multiselect("CRMs", options=crms, default=crms)

            st.session_state.selected_depts = selected_depts
            st.session_state.selected_crms = selected_crms

            filtered_count = len(st.session_state.processor.get_filtered_employee_mapping(selected_depts, selected_crms))
            total_count = len(st.session_state.processor.employee_mapping)
            st.info(f"Showing {filtered_count} of {total_count} employees")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.subheader("Master Data")
        master_file = st.file_uploader("Upload Master Data", type=['xlsx', 'xls'], key='master_upload')

        if master_file is not None and not st.session_state.master_loaded:
            with st.spinner("Loading master data..."):
                if st.session_state.processor.load_master_data(master_file, master_file.name):
                    st.session_state.master_loaded = True
                    st.success(f"Loaded {len(st.session_state.processor.employee_mapping)} employees")
                else:
                    for log in st.session_state.processor.logs:
                        if log['level'] == 'error':
                            st.error(log['message'])

        if st.session_state.master_loaded:
            st.success(f"{len(st.session_state.processor.employee_mapping)} employees loaded")

    with col2:
        st.subheader("Attendance Files")
        attendance_files = st.file_uploader("Upload Attendance Files", type=['xlsx', 'xls'],
                                            accept_multiple_files=True, key='attendance_upload')

        if attendance_files:
            st.session_state.attendance_files = [(f, f.name) for f in attendance_files]
            st.session_state.attendance_loaded = True
            st.success(f"{len(attendance_files)} file(s) uploaded")

    with col3:
        st.subheader("Leave Sheet")
        leave_file = st.file_uploader("Upload Leave Sheet", type=['xlsx', 'xls'], key='leave_upload')

        if leave_file is not None and not st.session_state.leave_loaded:
            with st.spinner("Loading leave data..."):
                if st.session_state.processor.load_leave_data(leave_file, leave_file.name):
                    st.session_state.leave_loaded = True
                    st.success(f"Loaded {len(st.session_state.processor.leave_records)} leave records")
                else:
                    for log in st.session_state.processor.logs:
                        if log['level'] == 'error':
                            st.error(log['message'])

        if st.session_state.leave_loaded:
            st.success(f"{len(st.session_state.processor.leave_records)} leave records loaded")

    st.divider()

    ready = st.session_state.master_loaded and st.session_state.attendance_loaded

    if st.button("GENERATE REPORT", type="primary", disabled=not ready, use_container_width=True):
        with st.spinner("Processing attendance data..."):
            progress_bar = st.progress(0)
            status_text = st.empty()

            selected_depts = getattr(st.session_state, 'selected_depts', None)
            selected_crms = getattr(st.session_state, 'selected_crms', None)

            # Validate filter selection
            if selected_depts is not None or selected_crms is not None:
                filtered = st.session_state.processor.get_filtered_employee_mapping(selected_depts, selected_crms)
                if not filtered:
                    st.warning("No employees match the current filter selection. Please adjust your Department/CRM filters.")
                    st.stop()

            status_text.text("Processing attendance files...")
            progress_bar.progress(10)
            st.session_state.processor.process_attendance_files(
                st.session_state.attendance_files, selected_depts, selected_crms
            )

            progress_bar.progress(30)

            if st.session_state.processor.leave_records:
                status_text.text("Applying leave records...")
                progress_bar.progress(50)
                st.session_state.processor.fill_leave_records(selected_depts, selected_crms)

            status_text.text("Calculating penalties...")
            progress_bar.progress(70)

            status_text.text("Generating Excel report...")
            progress_bar.progress(80)
            excel_data = st.session_state.processor.create_excel_report()

            progress_bar.progress(100)
            status_text.text("Report generated successfully!")

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="Download Report",
                data=excel_data,
                file_name=f"Attendance_Report_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

            with st.expander("Processing Log"):
                for log in st.session_state.processor.logs:
                    if log['level'] == 'error':
                        st.error(log['message'])
                    elif log['level'] == 'warning':
                        st.warning(log['message'])
                    elif log['level'] == 'success':
                        st.success(log['message'])
                    else:
                        st.info(log['message'])

            if st.session_state.processor.conflict_records:
                st.warning(f"Detected {len(st.session_state.processor.conflict_records)} leave vs attendance conflicts")
                with st.expander("View Conflicts"):
                    conflict_df = pd.DataFrame(st.session_state.processor.conflict_records)
                    st.dataframe(conflict_df)

    if not ready:
        st.info("Please upload Master Data and at least one Attendance file to generate a report.")

    if st.button("Reset All"):
        st.session_state.processor = AttendanceProcessor()
        st.session_state.master_loaded = False
        st.session_state.attendance_loaded = False
        st.session_state.leave_loaded = False
        st.session_state.attendance_files = []
        st.rerun()

    st.divider()
    st.caption("Attendance Dashboard v2.2 - Web Edition")


if __name__ == "__main__":
    main()
