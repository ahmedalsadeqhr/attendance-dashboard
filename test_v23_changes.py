"""
Tests for v2.3 changes:
  - No Show status (counted as Absent)
  - O(1) crm_to_info lookup
  - _safe_cell formula injection prevention
  - Pre-flight summary
  - Bare except logging
"""
import io
import sys
import pandas as pd
from datetime import datetime, time
from attendance_dashboard_streamlit import AttendanceProcessor, DEFAULT_CONFIG

PASS = 0
FAIL = 0


def check(label, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  [+] {label}" + (f"  ({detail})" if detail else ""))
    else:
        FAIL += 1
        print(f"  [FAIL] {label}" + (f"  ({detail})" if detail else ""))


def section(title):
    print()
    print("=" * 60)
    print(f"  {title}")
    print("=" * 60)


def make_processor_with_employees(employees):
    """Build a processor with a known employee set without file I/O."""
    p = AttendanceProcessor()
    for emp in employees:
        p.employee_mapping[emp['ac_no']] = {
            'crm': emp['crm'],
            'name': emp.get('name', emp['crm']),
            'department': emp.get('department', ''),
            'position': emp.get('position', ''),
            'national_id': emp.get('national_id', ''),
            'vendor': emp.get('vendor', ''),
            'ps_id': emp.get('ps_id', ''),
            'join_date': emp.get('join_date', ''),
            'exit_date': emp.get('exit_date', ''),
        }
    p.crm_to_info = {
        info['crm']: info
        for info in p.employee_mapping.values()
        if info['crm']
    }
    return p


# ── TEST 1: No Show in determine_status ───────────────────────────────────────
section("TEST 1: No Show from leave sheet applied correctly")

p = make_processor_with_employees([{'ac_no': '100', 'crm': 'EMP001'}])
date = datetime(2026, 3, 10)  # Monday
p.leave_records = [{'crm': 'EMP001', 'date': date, 'leave_type': 'No Show'}]
status, in_s, out_s = p.determine_status(None, None, date, 'EMP001')
check("No Show leave → status is 'No Show'", status == 'No Show', f"got '{status}'")
check("No Show leave → in_status is 'On Leave'", in_s == 'On Leave', f"got '{in_s}'")


# ── TEST 2: No Show counted as absence in calculate_penalties ─────────────────
section("TEST 2: No Show counted as Absent in calculate_penalties")

p2 = make_processor_with_employees([{'ac_no': '200', 'crm': 'EMP002'}])
p2.processed_data = [
    {'crm': 'EMP002', 'name': 'Test', 'department': '', 'position': '',
     'date': datetime(2026, 3, 3), 'status': 'No Show'},
    {'crm': 'EMP002', 'name': 'Test', 'department': '', 'position': '',
     'date': datetime(2026, 3, 4), 'status': 'No Show'},
    {'crm': 'EMP002', 'name': 'Test', 'department': '', 'position': '',
     'date': datetime(2026, 3, 5), 'status': 'Normal'},
]
penalties = p2.calculate_penalties()
check("No Show x2 → absence_count = 2",
      penalties['EMP002']['absence_count'] == 2,
      f"got {penalties['EMP002']['absence_count']}")
expected_ded = 2 * DEFAULT_CONFIG['penalties']['absence_deduction']
check("No Show x2 → absence_deduction = 4 days",
      penalties['EMP002']['absence_deduction'] == expected_ded,
      f"got {penalties['EMP002']['absence_deduction']}")


# ── TEST 3: Mixed Absent + No Show both counted ───────────────────────────────
section("TEST 3: Absent and No Show both count as absences")

p3 = make_processor_with_employees([{'ac_no': '300', 'crm': 'EMP003'}])
p3.processed_data = [
    {'crm': 'EMP003', 'name': 'Test', 'department': '', 'position': '',
     'date': datetime(2026, 3, 3), 'status': 'Absent'},
    {'crm': 'EMP003', 'name': 'Test', 'department': '', 'position': '',
     'date': datetime(2026, 3, 4), 'status': 'No Show'},
    {'crm': 'EMP003', 'name': 'Test', 'department': '', 'position': '',
     'date': datetime(2026, 3, 5), 'status': 'Normal'},
]
penalties3 = p3.calculate_penalties()
check("Absent + No Show → absence_count = 2",
      penalties3['EMP003']['absence_count'] == 2,
      f"got {penalties3['EMP003']['absence_count']}")


# ── TEST 4: No Show color in apply_status_color ───────────────────────────────
section("TEST 4: No Show gets red fill (same as Absent)")

from openpyxl import Workbook
from openpyxl.styles import PatternFill

p4 = AttendanceProcessor()
wb = Workbook()
ws = wb.active

cell_no_show = ws.cell(1, 1, 'No Show')
p4._apply_status_color(cell_no_show, 'No Show')

cell_absent = ws.cell(2, 1, 'Absent')
p4._apply_status_color(cell_absent, 'Absent')

check("No Show fill color matches Absent fill color",
      cell_no_show.fill.fgColor.rgb == cell_absent.fill.fgColor.rgb,
      f"No Show={cell_no_show.fill.fgColor.rgb}, Absent={cell_absent.fill.fgColor.rgb}")
check("No Show fill is red (FFC7CE)",
      'FFC7CE' in cell_no_show.fill.fgColor.rgb,
      f"got {cell_no_show.fill.fgColor.rgb}")


# ── TEST 5: crm_to_info built correctly on load_master_data ──────────────────
section("TEST 5: crm_to_info O(1) lookup built after master load")

p5 = AttendanceProcessor()
master_df = pd.DataFrame({
    'AC-No.': [111, 222, 333],
    'CRM': ['A001', 'A002', 'A003'],
    'Name': ['Alice', 'Bob', 'Carol'],
    'Department': ['HR', 'IT', 'HR'],
})
buf = io.BytesIO()
master_df.to_excel(buf, index=False)
buf.seek(0)
p5.load_master_data(buf, 'master.xlsx')

check("crm_to_info exists after load", hasattr(p5, 'crm_to_info'))
check("crm_to_info has correct count", len(p5.crm_to_info) == 3, f"got {len(p5.crm_to_info)}")
check("crm_to_info lookup by CRM works", p5.crm_to_info.get('A001', {}).get('name') == 'Alice',
      f"got {p5.crm_to_info.get('A001', {}).get('name')}")
check("Unknown CRM returns None", p5.crm_to_info.get('UNKNOWN') is None)


# ── TEST 6: _safe_cell formula injection prevention ───────────────────────────
section("TEST 6: _safe_cell prevents Excel formula injection")

p6 = AttendanceProcessor()
check("Normal name unchanged", p6._safe_cell("Ahmed") == "Ahmed")
check("= prefix gets quoted", p6._safe_cell("=CMD()").startswith("'"))
check("+ prefix gets quoted", p6._safe_cell("+1234").startswith("'"))
check("- prefix gets quoted", p6._safe_cell("-1234").startswith("'"))
check("@ prefix gets quoted", p6._safe_cell("@SUM").startswith("'"))
check("Non-string value untouched", p6._safe_cell(12345) == 12345)
check("Empty string untouched", p6._safe_cell("") == "")
check("CRM like 51Ahmed unchanged", p6._safe_cell("51Ahmed") == "51Ahmed")


# ── TEST 7: get_preflight_summary counts unmatched IDs ───────────────────────
section("TEST 7: Pre-flight summary detects unmatched IDs")

p7 = AttendanceProcessor()
# Load master with 2 employees
master_df7 = pd.DataFrame({
    'AC-No.': [1, 2],
    'CRM': ['CRM1', 'CRM2'],
    'Name': ['Emp1', 'Emp2'],
})
buf7 = io.BytesIO()
master_df7.to_excel(buf7, index=False)
buf7.seek(0)
p7.load_master_data(buf7, 'master.xlsx')

# Attendance with 3 rows: 2 match, 1 doesn't
att_df7 = pd.DataFrame({
    'AC-No.': [1, 2, 999],
    'Name': ['Emp1', 'Emp2', 'Ghost'],
    'Clock In 1': ['12:00:00 PM', '11:00:00 AM', '10:00:00 AM'],
    'Clock Out 1': ['09:00:00 PM', '09:00:00 PM', '09:00:00 PM'],
    'Date': ['2026-03-10', '2026-03-10', '2026-03-10'],
})
att_buf7 = io.BytesIO()
att_df7.to_excel(att_buf7, index=False)
att_buf7.seek(0)

summary7 = p7.get_preflight_summary([(att_buf7, 'att_2026-03-10.xlsx')])
check("master_employees = 2", summary7['master_employees'] == 2, f"got {summary7['master_employees']}")
check("attendance_rows = 3", summary7['attendance_rows'] == 3, f"got {summary7['attendance_rows']}")
check("unmatched_ac_nos = 1 (ID 999)", summary7['unmatched_ac_nos'] == 1,
      f"got {summary7['unmatched_ac_nos']}")


# ── TEST 8: No Show in justification dropdown list ────────────────────────────
section("TEST 8: No Show appears in justification_options (dropdown)")

# Generate a minimal report and check the validation list
p8 = make_processor_with_employees([{'ac_no': '400', 'crm': 'EMP004'}])
p8.processed_data = [
    {'crm': 'EMP004', 'name': 'Test', 'department': '', 'position': '',
     'date': datetime(2026, 3, 10), 'day': 'Monday',
     'clock_in': None, 'clock_out': None,
     'in_status': 'On Leave', 'out_status': 'On Leave', 'status': 'No Show'},
]

wb8 = Workbook()
wb8.remove(wb8.active)
p8.calculate_penalties()
p8.create_summary_sheet(wb8)

ws8 = wb8['Summary Report']
found_no_show = False
for dv in ws8.data_validations.dataValidation:
    if dv.formula1 and 'No Show' in str(dv.formula1):
        found_no_show = True
        break

check("'No Show' in summary sheet dropdown validation", found_no_show)


# ── TEST 9: Bare except logging — no silent failures ─────────────────────────
section("TEST 9: Errors are logged, not silently swallowed")

p9 = AttendanceProcessor()
p9.crm_to_info = {}

# Trigger bad exit date parsing
p9.employee_mapping['BAD'] = {
    'crm': 'BADCRM', 'name': 'X', 'department': '', 'position': '',
    'national_id': '', 'vendor': '', 'ps_id': '',
    'join_date': '', 'exit_date': 'not-a-date'
}
p9.crm_to_info['BADCRM'] = p9.employee_mapping['BAD']

initial_log_count = len(p9.logs)
p9.determine_status(None, None, datetime(2026, 3, 10), 'BADCRM')
check("Bad exit date logs a warning (not silent)",
      len(p9.logs) > initial_log_count,
      f"logs before={initial_log_count}, after={len(p9.logs)}")
check("Warning level set",
      any(l['level'] == 'warning' for l in p9.logs[initial_log_count:]))


# ── TEST 10: No Show weekend interaction ──────────────────────────────────────
section("TEST 10: No Show on a Friday (weekend) is ignored")

p10 = make_processor_with_employees([{'ac_no': '500', 'crm': 'EMP005'}])
friday = datetime(2026, 3, 13)  # Friday
p10.leave_records = [{'crm': 'EMP005', 'date': friday, 'leave_type': 'No Show'}]
status10, _, _ = p10.determine_status(None, None, friday, 'EMP005')
check("No Show on Friday → Weekend (off day takes priority)",
      status10 == 'Weekend', f"got '{status10}'")


# ── FINAL RESULTS ─────────────────────────────────────────────────────────────
print()
print("=" * 60)
print("  FINAL RESULTS")
print("=" * 60)
print()
print(f"  Passed : {PASS}/{PASS+FAIL}")
print(f"  Failed : {FAIL}/{PASS+FAIL}")
print()
if FAIL == 0:
    print("  ALL TESTS PASSED")
else:
    print("  SOME TESTS FAILED — review output above")
print()
sys.exit(0 if FAIL == 0 else 1)
